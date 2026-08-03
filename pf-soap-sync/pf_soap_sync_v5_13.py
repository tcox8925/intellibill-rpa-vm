from __future__ import annotations

"""
Practice Fusion appointment -> Facesheet/SOAP PDF sync (Playwright/CDP)

Storage used today
------------------
- Appointment report: CSV/XLSX (pulled by Playwright or supplied manually)
- Persistent queue + patient mappings + run history: JSON
- Patient registry: CSV/JSON/XLSX (temporary until moved to a database table)
- Output charts: PDF

Main commands
-------------
- self-test          Local synthetic test; no Practice Fusion/login required.
- doctor             Validate configs, Chrome paths, and local inputs.
- pull-report        Pull Appointment & eligibility report for a date range.
- ingest             Upsert a report file into the JSON queue.
- match-patients     Resolve appointment patients against the patient registry.
- resolve-patient    Manually resolve a needs_attention row and save the mapping.
- process            Process ready/review rows and create PDFs.
- refresh            Refresh one appointment/encounter or the newest encounter for a patient.
- nightly            pull-report -> ingest -> match-patients -> process.
- status             Show queue counts and unresolved rows.
- reset              Put selected rows back into ready for retesting.

Authentication/profile behavior
-------------------------------
- Uses Playwright attached to normal Google Chrome over CDP.
- The real Profile 11 may be cloned once into a dedicated automation directory.
- The dedicated profile is reused on later runs, preserving cookies/device trust.
- Username/password can be supplied with PF_USERNAME and PF_PASSWORD.
- Playwright enters them through normal keyboard/input events and clicks Log in.
- If Practice Fusion requires OTP/security verification, the worker waits for the
  authenticated EHR to appear; OTP still requires an external source or manual entry.
"""

import argparse
BUILD_ID = "PF-SOAP-SYNC-v5.13.0-summary-only-appointment-check"

import base64
import csv
import difflib
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple


def configure_csv_field_limit() -> int:
    """Raise Python's CSV cell-size limit for PF exports containing large JSON fields.

    Patient registry exports may include raw_patient_json, summary, or note columns that
    exceed csv's default 131,072-character ceiling. Use the largest platform-supported
    value, reducing it only when the local C runtime rejects the integer.
    """
    candidate = sys.maxsize
    while candidate > 131_072:
        try:
            csv.field_size_limit(candidate)
            return candidate
        except OverflowError:
            candidate //= 10
    csv.field_size_limit(131_072)
    return 131_072


CSV_FIELD_LIMIT = configure_csv_field_limit()
from zoneinfo import ZoneInfo

from playwright.sync_api import BrowserContext, Locator, Page, TimeoutError as PWTimeout
from playwright.sync_api import sync_playwright


LOGIN_URL = "https://static.practicefusion.com/apps/ehr/index.html#/login"
EHR_BASE_URL = "https://static.practicefusion.com/apps/ehr/index.html"
PF_USERNAME_SELECTOR = "#inputUsername, input[name='inputUsername']"
PF_PASSWORD_SELECTOR = "#inputPswd, input[name='inputPswdTest']"
PF_LOGIN_BUTTON_SELECTOR = "#loginButton"
DEFAULT_TIMEOUT = 30_000
SHORT_TIMEOUT = 5_000
# v5.4: the practice timezone drives default report dates. NWARK Internal Medicine
# is in Rogers, Arkansas (Central). The previous hardcoded America/Detroit value was
# a leftover from another practice and made an unattended nightly run between
# 11:00 PM and midnight Central pull the following day's report.
PRACTICE_TZ_NAME = os.environ.get("PF_PRACTICE_TIMEZONE", "America/Chicago").strip() or "America/Chicago"
try:
    PRACTICE_TZ = ZoneInfo(PRACTICE_TZ_NAME)
except Exception:  # pragma: no cover - invalid tz name supplied by operator
    print(
        f"WARNING: unknown PF_PRACTICE_TIMEZONE {PRACTICE_TZ_NAME!r}; falling back to America/Chicago.",
        flush=True,
    )
    PRACTICE_TZ_NAME = "America/Chicago"
    PRACTICE_TZ = ZoneInfo(PRACTICE_TZ_NAME)

PATIENT_NAME_SELECTOR = "[data-element='full-name']"
PATIENT_RECORD_NUMBER_SELECTOR = "[data-element='prn-text']"

DEFAULT_IGNORED_STATUSES = {
    "cancelled",
    "canceled",
    "no show",
    "no-show",
    "rescheduled",
    "deleted",
    "void",
}

DEFAULT_SEEN_STATUSES = {
    "seen",
    "completed",
    "complete",
    "checked out",
    "checked-out",
    "signed",
}

PROFILE_CACHE_IGNORE = shutil.ignore_patterns(
    "Cache",
    "Code Cache",
    "GPUCache",
    "GraphiteDawnCache",
    "DawnCache",
    "DawnGraphiteCache",
    "DawnWebGPUCache",
    "ShaderCache",
    "GrShaderCache",
    "Crashpad",
    "component_crx_cache",
    "optimization_guide_model_store",
    "*.tmp",
    "*.log",
    "SingletonLock",
    "SingletonCookie",
    "SingletonSocket",
    "lockfile",
)

_CHROME_PROC: Optional[subprocess.Popen] = None


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class QueueRecord:
    row_id: str
    practice: str = ""

    appointment_id: str = ""
    appointment_date: str = ""
    appointment_status: str = ""
    appointment_type: str = ""
    provider: str = ""

    # Identity from the appointment report.
    patient_name: str = ""
    patient_dob: str = ""
    patient_phone: str = ""
    patient_phone_normalized: str = ""

    # Resolved Practice Fusion patient identity.
    patient_id: str = ""
    ehr_patient_guid: str = ""
    patient_match_status: str = "unmatched"
    patient_match_method: str = ""
    patient_match_score: float = 0.0
    patient_match_message: str = ""
    patient_candidates: List[Dict[str, Any]] = field(default_factory=list)

    # Encounter discovered from the authenticated patient Summary/timeline.
    encounter_id: str = ""  # Native encounter ID if a future source provides it.
    encounter_key: str = ""  # Stable derived key from visible encounter content.
    encounter_date: str = ""
    encounter_type: str = ""
    encounter_code: str = ""
    encounter_chief_complaint: str = ""
    encounter_source: str = ""

    # Queue state.
    status: str = "ready"
    status_reason: str = ""
    message: str = ""
    attempt_count: int = 0
    review_count: int = 0
    refresh_count: int = 0

    source_report_name: str = ""
    source_row_json: Dict[str, Any] = field(default_factory=dict)

    created_at: str = ""
    updated_at: str = ""
    first_ready_at: str = ""
    processing_started_at: str = ""
    last_checked_at: str = ""
    processed_at: str = ""

    selected_soap_note_text: str = ""
    # v5.5: record what was actually printed, so a chart PDF can be audited after the
    # fact without re-driving the browser.
    selected_sections: List[str] = field(default_factory=list)
    notes_selection_mode: str = ""
    pdf_path: str = ""
    elapsed_seconds: float = 0.0
    error_message: str = ""
    scrape_run_id: str = ""

    @classmethod
    def from_dict(cls, value: Dict[str, Any]) -> "QueueRecord":
        allowed = cls.__dataclass_fields__.keys()
        return cls(**{key: value.get(key) for key in allowed if key in value})


@dataclass
class DetectedEncounter:
    encounter_date: str
    encounter_type: str = ""
    encounter_code: str = ""
    chief_complaint: str = ""
    display_text: str = ""
    encounter_key: str = ""
    source: str = "summary"


class EncounterNotFoundError(LookupError):
    pass


class SoapNoteNotFoundError(LookupError):
    pass


@dataclass
class SyncConfig:
    # Confirmed patient Summary / Print Chart controls.
    print_chart_button_selector: str = "[data-element='btn-print-chart-control-bar']"

    # v5.5: the chart page contains TWO elements carrying
    # data-element="print-patient-content-modal": an outer
    # `carbon-content-modal-component` wrapper that stays hidden, and the real
    # `.content-modal` dialog. Waiting on `.first` locked onto the hidden wrapper and
    # timed out after 30s ("63 x locator resolved to hidden ..."). Ordered
    # most-specific-first, and every wait now scans all matches for a visible one.
    print_modal_ready_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='print-patient-content-modal'].content-modal",
            "[data-element='print-patient-content-modal'].modal",
            "div[data-element='print-patient-content-modal']:not(.carbon-content-modal-component)",
            "[data-element='print-patient-content-modal']",
        ]
    )
    print_modal_close_selector: str = "[data-element='concent-model-close-btn']"
    print_modal_cancel_selector: str = "[data-element='btn-print-modal-cancel']"

    # "Select: all | none" links inside the modal header.
    print_modal_select_none_selector: str = "[data-element='print-modal-select-none']"
    print_modal_select_all_selector: str = "[data-element='print-modal-select-all']"

    notes_dropdown_selector: str = "[data-element='print-chart-notes-dropdown']"
    notes_dropdown_toggle_selector: str = (
        "[data-element='print-chart-notes-dropdown'] "
        "[data-element='checkbox-dropdown-grouping__toggle']"
    )
    # The group checkbox in the notes row selects/clears every note at once.
    notes_group_checkbox_selector: str = (
        "[data-element='print-chart-notes-dropdown'] input[type='checkbox']"
    )
    # Toggle label text when nothing is selected, used to verify the selection took.
    notes_empty_label_text: str = "No notes selected"

    # The notes menu is tethered outside the modal, so it cannot be scoped to the modal.
    # Tried in order; falls back to a page-wide search filtered by date token.
    note_dropdown_menu_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='checkbox-dropdown-grouping__menu']",
            ".checkbox-dropdown-grouping__menu",
            ".tether-element .dropdown-menu",
            ".tether-element",
        ]
    )

    # Sections to CHECK. Everything else in the modal is explicitly cleared first, so
    # this list is the complete intended selection, not a set of additions.
    facesheet_checkbox_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='chk-patient-demographics'] input[type='checkbox']",
            "[data-element='print-insurance-options'] input[type='checkbox']",
            "[data-element='chk-diagnoses'] input[type='checkbox']",
        ]
    )

    # Every option the Print Chart modal is known to render, so the exact-selection pass
    # can clear the ones that are not wanted and detect any new option PF adds. The notes
    # dropdown is deliberately absent: it is selected separately and per record.
    facesheet_known_option_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='chk-patient-demographics'] input[type='checkbox']",
            "[data-element='print-insurance-options'] input[type='checkbox']",
            "[data-element='chk-flowsheet-Vitals'] input[type='checkbox']",
            "[data-element='growth-charts'] input[type='checkbox']",
            "[data-element='chk-diagnoses'] input[type='checkbox']",
            "[data-element='chk-allergies'] input[type='checkbox']",
            "[data-element='chk-medications'] input[type='checkbox']",
            "[data-element='chk-immunizations'] input[type='checkbox']",
            "[data-element='chk-social-history'] input[type='checkbox']",
            "[data-element='chk-past-medical-history'] input[type='checkbox']",
            "[data-element='chk-family-history'] input[type='checkbox']",
            "[data-element='chk-advanced-directives'] input[type='checkbox']",
            "[data-element='print-implantable-devices-checkbox'] input[type='checkbox']",
            "[data-element='chk-health-concerns'] input[type='checkbox']",
            "[data-element='chk-goals'] input[type='checkbox']",
            "[data-element='chk-sia'] input[type='checkbox']",
        ]
    )
    # Fail rather than print a chart whose section selection could not be verified.
    enforce_exact_facesheet_selection: bool = True

    # "auto"  - all notes on the first PDF for a patient, then only the appointment date
    # "all"   - always every note
    # "date"  - always only the appointment date
    notes_selection_mode: str = "auto"

    note_option_selector: str = "input[type='checkbox']"
    generate_pdf_button_selector: str = "[data-element='btn-print-modal-print']"
    printable_preview_ready_selector: str = "a.print-link[title='Print']"
    pdf_min_bytes: int = 1024

    # Summary and View All encounters.
    encounter_item_selector: str = "[data-element^='encounter-item-']"
    encounter_date_selector: str = ".text-color-link"
    encounter_type_selector: str = "[data-element='encounter-type']"
    encounter_code_selector: str = "[data-element='code-type-and-code-value']"
    encounter_chief_complaint_selector: str = ".chief-complaint"
    encounter_timeline_link_selector: str = "a[href*='/timeline/encounter']"
    encounter_timeline_item_selector: str = (
        "[data-element^='encounter-item-'], li[title*='SOAP Note']"
    )
    encounter_soap_title_text: str = "SOAP Note"

    download_filename_template: str = (
        "{patient_id}_{appointment_date}_{encounter_id}.pdf"
    )
    note_date_formats: List[str] = field(
        default_factory=lambda: [
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y-%m-%d",
            "%b %d, %Y",
            "%B %d, %Y",
        ]
    )
    ignored_statuses: List[str] = field(
        default_factory=lambda: sorted(DEFAULT_IGNORED_STATUSES)
    )
    seen_statuses: List[str] = field(
        default_factory=lambda: sorted(DEFAULT_SEEN_STATUSES)
    )

    @classmethod
    def load(cls, path: str) -> "SyncConfig":
        if not path or not os.path.exists(path):
            return cls()
        with open(path, "r", encoding="utf-8") as handle:
            raw = json.load(handle)
        allowed = cls.__dataclass_fields__.keys()
        defaults = asdict(cls())

        # v5.5: print_modal_ready_selector became a list. Accept the old singular key.
        sync_aliases = {
            "print_modal_ready_selector": "print_modal_ready_selectors",
            "note_dropdown_menu_selector": "note_dropdown_menu_selectors",
        }
        for old, new in sync_aliases.items():
            if raw.get(old) and not raw.get(new):
                raw[new] = [raw[old]]

        # Never allow stale placeholder values ("", null, or []) to erase
        # confirmed built-in selectors. Only meaningful config values override
        # the dataclass defaults. This also makes older config drafts safe.
        overrides = {}
        unknown: List[str] = []
        healed: List[str] = []
        for key, value in raw.items():
            if key not in allowed:
                if key not in sync_aliases:
                    unknown.append(key)
                continue
            if value is None:
                healed.append(key)
                continue
            if isinstance(value, str) and not value.strip():
                healed.append(key)
                continue
            if isinstance(value, (list, dict)) and not value:
                healed.append(key)
                continue
            overrides[key] = value

        # v5.4: unknown and placeholder keys used to be dropped in silence, so a
        # misspelled key looked identical to a working one and an empty placeholder gave
        # no hint that the built-in default was doing the work instead of the file.
        if unknown:
            print(
                f"WARNING: ignoring unrecognized keys in {path}: {sorted(unknown)}",
                flush=True,
            )
        if healed:
            print(
                f"NOTE: empty/null keys in {path} fell back to built-in defaults: "
                f"{sorted(set(healed))}",
                flush=True,
            )

        merged = {**defaults, **overrides}
        return cls(**merged)


@dataclass
class AppointmentReportConfig:
    # Confirmed selectors supplied by the user.
    reports_menu_selector: str = "a[data-tracking='Reports'][href='#/PF/reporting']"
    appointment_report_link_selector: str = "a[data-element='appointment-report']"
    report_start_date_selector: str = "input[data-element='start-date']"
    report_end_date_selector: str = "input[data-element='end-date']"

    # PF commonly uses run-report-button; text fallbacks make the flow testable if
    # this report uses another stable data-element.
    run_report_button_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='run-report-button']",
            "button:has-text('Run report')",
            "button:has-text('Run Report')",
            "button:has-text('Run')",
        ]
    )
    report_ready_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='pager-label']",
            "tr[data-element^='data-table-row-']",
            "table tbody tr",
            "[role='row']",
        ]
    )
    export_report_button_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element*='export']",
            "button:has-text('Export')",
            "a:has-text('Export')",
            "button:has-text('Download')",
            "a:has-text('Download')",
        ]
    )
    csv_export_option_selectors: List[str] = field(
        default_factory=lambda: [
            "[role='menuitem']:has-text('CSV')",
            "li:has-text('CSV')",
            "a:has-text('CSV')",
            "button:has-text('CSV')",
        ]
    )

    # Generic table fallback when PF does not expose a direct CSV download.
    table_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element*='results-table']",
            "[data-element*='report-table']",
            "table",
            "[role='table']",
        ]
    )
    row_selectors: List[str] = field(
        default_factory=lambda: [
            "tr[data-element^='data-table-row-']",
            "tbody tr",
            "[role='row']",
        ]
    )
    scroller_selectors: List[str] = field(
        default_factory=lambda: [
            "[data-element='data-table-scroller']",
            ".data-table-scroller",
        ]
    )
    next_page_selectors: List[str] = field(
        default_factory=lambda: [
            "button[data-element*='next']:not([disabled])",
            "a[data-element*='next']:not([disabled])",
            "button[aria-label*='Next']:not([disabled])",
            "a[aria-label*='Next']:not([disabled])",
            "button:has-text('Next'):not([disabled])",
            "a:has-text('Next')",
        ]
    )
    no_results_selectors: List[str] = field(
        default_factory=lambda: [
            "text=No results were found",
            "text=No results",
        ]
    )

    @classmethod
    def load(cls, path: str) -> "AppointmentReportConfig":
        if not path or not os.path.exists(path):
            return cls()
        with open(path, "r", encoding="utf-8") as handle:
            raw = json.load(handle)
        defaults = asdict(cls())
        allowed = cls.__dataclass_fields__.keys()
        # Backward-compatible singular keys from earlier config drafts.
        aliases = {
            "run_report_button_selector": "run_report_button_selectors",
            "report_ready_selector": "report_ready_selectors",
            "export_report_button_selector": "export_report_button_selectors",
            "report_row_selector": "row_selectors",
            "report_next_page_selector": "next_page_selectors",
        }
        for old, new in aliases.items():
            if raw.get(old) and not raw.get(new):
                raw[new] = [raw[old]]
        overrides = {}
        unknown: List[str] = []
        healed: List[str] = []
        for key, value in raw.items():
            if key not in allowed:
                if key not in aliases:
                    unknown.append(key)
                continue
            if value is None:
                healed.append(key)
                continue
            if isinstance(value, str) and not value.strip():
                healed.append(key)
                continue
            if isinstance(value, (list, dict)) and not value:
                healed.append(key)
                continue
            overrides[key] = value

        if unknown:
            print(
                f"WARNING: ignoring unrecognized keys in {path}: {sorted(unknown)}",
                flush=True,
            )
        if healed:
            print(
                f"NOTE: empty/null keys in {path} fell back to built-in defaults: "
                f"{sorted(set(healed))}",
                flush=True,
            )

        merged = {**defaults, **overrides}
        return cls(**merged)


# ---------------------------------------------------------------------------
# Generic helpers and persistent JSON store
# ---------------------------------------------------------------------------


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def practice_today() -> date:
    return datetime.now(PRACTICE_TZ).date()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: str) -> str:
    """Normalize a source column header to a lowercase space-separated token.

    v5.4: Practice Fusion's CSV export emits camelCase headers (AppointmentTime,
    MobilePhone, AppointmentType, AppointmentStatus, SeenBy) while the on-screen
    report renders spaced titles. Collapsing only non-alphanumerics left the
    camelCase forms as single tokens ("appointmenttime"), so six of thirteen target
    fields silently failed to map. Splitting on the lower->upper boundary makes both
    the exported and the scraped header forms normalize to the same token.
    """
    text = clean(value)
    # AppointmentTime -> Appointment Time ; DOBValue -> DOB Value
    text = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", text)
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def normalize_status(value: str) -> str:
    return clean(value).lower().replace("_", " ")


def status_matches(value: str, configured_statuses: Iterable[str]) -> bool:
    normalized = normalize_status(value)
    if not normalized:
        return False
    for configured in configured_statuses:
        token = normalize_status(configured)
        if token and (normalized == token or token in normalized):
            return True
    return False


# ---------------------------------------------------------------------------
# Single-definition appointment status gates
#
# Every code path that decides whether an appointment is skipped or considered
# clinically complete must call these two functions. Duplicating the status lists
# is how a gate fix lands in one path and misses its twin.
# ---------------------------------------------------------------------------


def is_ignored_status(value: str, config: "SyncConfig") -> bool:
    return status_matches(value, config.ignored_statuses)


def is_seen_status(value: str, config: "SyncConfig") -> bool:
    return status_matches(value, config.seen_statuses)


def parse_date(value: str) -> Optional[date]:
    text = clean(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    formats = (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%b %d, %Y",
        "%B %d, %Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    # Appointment exports commonly append a time.
    date_prefixes = [
        text.split(" ")[0],
        " ".join(text.split(" ")[:2]),
        " ".join(text.split(" ")[:3]),
    ]
    for prefix in date_prefixes:
        for fmt in formats:
            try:
                return datetime.strptime(prefix, fmt).date()
            except ValueError:
                pass
    return None


def require_date(value: str, label: str) -> date:
    parsed = parse_date(value)
    if parsed is None:
        raise ValueError(f"Could not parse {label}: {value!r}")
    return parsed


def safe_filename(value: str) -> str:
    value = re.sub(r"[<>:\"/\\|?*]+", "_", value)
    value = re.sub(r"\s+", "_", value).strip("._")
    return value or "practice_fusion_file"


def atomic_write_json(path: str, payload: Any) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_suffix(target.suffix + ".tmp")
    with temp.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp, target)


def empty_store() -> Dict[str, Any]:
    return {
        "schema_version": 3,
        "updated_at": now_iso(),
        "counts": {},
        "patient_mappings": [],
        "runs": [],
        "rows": [],
    }


def load_store(path: str) -> Dict[str, Any]:
    if not os.path.exists(path):
        return empty_store()
    with open(path, "r", encoding="utf-8") as handle:
        raw = json.load(handle)
    if isinstance(raw, list):
        store = empty_store()
        store["rows"] = raw
        return store
    if not isinstance(raw, dict):
        raise ValueError("Queue JSON must be an object or a list of rows.")
    store = empty_store()
    store.update(raw)
    store.setdefault("rows", [])
    store.setdefault("patient_mappings", [])
    store.setdefault("runs", [])
    return store


def store_rows(store: Dict[str, Any]) -> List[QueueRecord]:
    return [
        QueueRecord.from_dict(item)
        for item in store.get("rows", [])
        if isinstance(item, dict)
    ]


def save_store(path: str, store: Dict[str, Any], rows: Optional[Sequence[QueueRecord]] = None) -> None:
    if rows is not None:
        store["rows"] = [asdict(row) for row in rows]
    counts: Dict[str, int] = {}
    for item in store.get("rows", []):
        status = clean(item.get("status") if isinstance(item, dict) else "") or "unknown"
        counts[status] = counts.get(status, 0) + 1
    store["schema_version"] = 3
    store["updated_at"] = now_iso()
    store["counts"] = counts
    store["runs"] = list(store.get("runs", []))[-100:]
    atomic_write_json(path, store)


def append_run(store: Dict[str, Any], command: str, details: Dict[str, Any]) -> str:
    run_id = str(uuid.uuid4())
    store.setdefault("runs", []).append(
        {
            "run_id": run_id,
            "command": command,
            "started_at": now_iso(),
            **details,
        }
    )
    return run_id


def finish_run(store: Dict[str, Any], run_id: str, status: str, details: Dict[str, Any]) -> None:
    for run in reversed(store.get("runs", [])):
        if run.get("run_id") == run_id:
            run.update({"finished_at": now_iso(), "status": status, **details})
            return


# ---------------------------------------------------------------------------
# Tabular file handling
# ---------------------------------------------------------------------------


def read_tabular_rows(path: str) -> List[Dict[str, Any]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".json":
        with open(path, "r", encoding="utf-8") as handle:
            raw = json.load(handle)
        if isinstance(raw, dict):
            raw = raw.get("rows", raw.get("data", []))
        if not isinstance(raw, list):
            raise ValueError(f"JSON tabular input must contain a list of rows: {path}")
        return [dict(item) for item in raw if isinstance(item, dict)]

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read XLSX appointment reports") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean(value) or f"column_{index + 1}" for index, value in enumerate(rows[0])]
        return [
            {headers[index]: clean(value) for index, value in enumerate(row)}
            for row in rows[1:]
            if any(clean(value) for value in row)
        ]

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"Tabular file has no header row: {path}")
        return [dict(row) for row in reader]


def write_csv(path: str, rows: Sequence[Dict[str, Any]]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            key = clean(key)
            if key and key not in seen:
                headers.append(key)
                seen.add(key)
    if not headers:
        headers = ["message"]
        rows = [{"message": "No results"}]
    with target.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: clean(row.get(key)) for key in headers})


# ---------------------------------------------------------------------------
# Appointment report ingestion
# ---------------------------------------------------------------------------


COLUMN_ALIASES: Dict[str, Tuple[str, ...]] = {
    "appointment_id": (
        "appointment_id", "appointment id", "appt_id", "appt id", "confirmation number"
    ),
    "encounter_id": (
        "encounter_id", "encounter id", "visit_id", "visit id"
    ),
    "patient_id": (
        "patient_id", "patient id", "record_number", "record number", "prn"
    ),
    "ehr_patient_guid": (
        "ehr_patient_guid", "patient_guid", "patient guid", "pf_patient_guid"
    ),
    "patient_name": (
        "patient_name", "patient name", "name", "patient"
    ),
    "patient_first_name": (
        "patient first name", "first name", "patient_first_name"
    ),
    "patient_last_name": (
        "patient last name", "last name", "patient_last_name"
    ),
    "patient_dob": (
        "patient_dob", "patient dob", "dob", "date of birth", "birth date"
    ),
    "patient_phone": (
        "patient_phone", "patient phone", "phone", "phone number", "mobile phone",
        "home phone", "preferred contact"
    ),
    "appointment_date": (
        # "appointment time" is the header PF actually exports (AppointmentTime).
        "appointment_date", "appointment date", "appointment date time",
        "appointment time", "appointmenttime", "appt time",
        "appt date", "appt date time", "date/time", "date time", "date",
        "date of service", "service date", "scheduled date", "start time",
        "appointment start"
    ),
    "appointment_status": (
        "appointment_status", "appointment status", "appt status", "status"
    ),
    "appointment_type": (
        "appointment_type", "appointment type", "appt type", "visit type", "reason"
    ),
    "provider": (
        # "seen by" is the header PF actually exports (SeenBy).
        "provider", "provider_name", "provider name", "seen by provider",
        "seen by", "seenby", "rendering provider",
        "resource", "staff"
    ),
}

# Fields that must map to a non-empty value on at least one row of a real report.
# Anything listed here is treated as a hard mapping failure rather than a blank column.
REQUIRED_APPOINTMENT_FIELDS: Dict[str, str] = {
    "appointment_date": "appointment date/time",
    "patient_name": "patient name",
    "patient_dob": "patient DOB",
    "appointment_status": "appointment status",
    "appointment_type": "appointment type",
    "provider": "provider (seen by)",
}

# Fields that are expected but not fatal; a warning keeps a silent blank visible.
OPTIONAL_APPOINTMENT_FIELDS: Dict[str, str] = {
    "patient_phone": "patient phone",
}


def alias_value(normalized: Dict[str, str], aliases: Iterable[str]) -> str:
    for alias in aliases:
        value = normalized.get(normalize_header(alias), "")
        if value:
            return value
    return ""


def map_appointment_row(source: Dict[str, Any]) -> Dict[str, str]:
    normalized: Dict[str, str] = {}
    for key, value in source.items():
        normalized_key = normalize_header(key)
        cleaned_value = clean(value)
        # Preserve a non-empty value when two source headers normalize to the same
        # token (for example PHONE and Phone in synthetic/merged files).
        if normalized_key not in normalized or cleaned_value:
            normalized[normalized_key] = cleaned_value
    mapped = {
        target: alias_value(normalized, aliases)
        for target, aliases in COLUMN_ALIASES.items()
    }
    if not mapped["patient_name"]:
        mapped["patient_name"] = clean(
            f"{mapped['patient_first_name']} {mapped['patient_last_name']}"
        )
    return mapped


def validate_appointment_report_mapping(
    source_rows: Sequence[Dict[str, Any]], mapped_rows: Sequence[Dict[str, str]]
) -> None:
    """Fail early when a PF export column was not recognized.

    This prevents a whole report from entering the queue with blank appointment dates
    or patient identities. Practice Fusion currently exports headers such as
    DATE/TIME, APPT. TYPE, APPT. STATUS, and SEEN BY PROVIDER.
    """
    if not source_rows:
        return
    headers = [clean(key) for key in source_rows[0].keys() if clean(key)]

    def any_populated(field_name: str) -> bool:
        return any(clean(row.get(field_name, "")) for row in mapped_rows)

    # v5.4: this used to check only date/name/DOB. A blank appointment_status is the
    # most dangerous silent failure in the pipeline: status_matches() returns False on
    # an empty string, so the ignored gate never fires and cancelled/no-show
    # appointments enter the queue as "ready" and get driven through the browser.
    missing = [
        label for field_name, label in REQUIRED_APPOINTMENT_FIELDS.items()
        if not any_populated(field_name)
    ]
    if missing:
        raise ValueError(
            "Appointment report column mapping failed for "
            + ", ".join(missing)
            + f". Actual CSV headers: {headers}"
            + f". Normalized headers: {sorted({normalize_header(h) for h in headers})}"
        )

    for field_name, label in OPTIONAL_APPOINTMENT_FIELDS.items():
        if not any_populated(field_name):
            print(
                f"WARNING: no source column mapped to {label}; "
                f"phone-based match tie-breaking will be unavailable.",
                flush=True,
            )


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", clean(value))
    if len(digits) > 10 and digits.startswith("1"):
        digits = digits[-10:]
    return digits


def normalize_person_name(value: str) -> str:
    text = clean(value).lower()
    # Convert Last, First into First Last for comparison.
    if "," in text:
        left, right = text.split(",", 1)
        text = f"{right} {left}"
    text = re.sub(r"\b(jr|sr|ii|iii|iv|mr|mrs|ms|dr)\b\.?", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return clean(text)


def name_similarity(left: str, right: str) -> float:
    a = normalize_person_name(left)
    b = normalize_person_name(right)
    if not a or not b:
        return 0.0
    direct = difflib.SequenceMatcher(None, a, b).ratio()
    token = difflib.SequenceMatcher(None, " ".join(sorted(a.split())), " ".join(sorted(b.split()))).ratio()
    return max(direct, token)


def name_token_containment(left: str, right: str) -> float:
    """Fraction of the shorter name's tokens present in the longer name.

    Character-ratio similarity punishes the dominant real-world mismatch at this
    practice: the appointment report drops a middle name or a second surname that the
    chart carries ("Marlene Revilla Gomez" vs "Marlene Del Carmen Revilla Gomez",
    "Elizabeth Vazquez Martinez" vs "Elizabeth Vazquez"). Those score 0.75-0.79 on
    difflib and fall under the 0.82 threshold despite an exact DOB match.

    The subset direction is only trusted when the shorter name carries at least two
    distinct tokens. A single-token name cannot establish identity on its own: the real
    registry contains a malformed row "Peyton Peyton", whose token set is just
    {peyton}, and an unguarded shorter-side ratio scored that 1.0 against "Peyton
    Hicks" -- silently attaching the wrong chart. One shared given name is never
    sufficient evidence, even inside a matching DOB bucket.
    """
    a = set(normalize_person_name(left).split())
    b = set(normalize_person_name(right).split())
    if not a or not b:
        return 0.0
    shared = len(a & b)
    if shared < 2:
        return 0.0
    smaller = min(len(a), len(b))
    if smaller < 2:
        return 0.0
    return shared / smaller


def identity_score(appointment_name: str, registry_name: str) -> float:
    return max(
        name_similarity(appointment_name, registry_name),
        name_token_containment(appointment_name, registry_name),
    )


def index_registry_by_dob(registry: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    buckets: Dict[str, List[Dict[str, Any]]] = {}
    for patient in registry:
        dob = clean(patient.get("dob"))
        if dob:
            buckets.setdefault(dob, []).append(patient)
    return buckets


def record_key(mapped: Dict[str, str], practice: str) -> str:
    if mapped.get("appointment_id"):
        return f"{practice}|appointment|{mapped['appointment_id']}"
    if mapped.get("encounter_id"):
        return f"{practice}|encounter|{mapped['encounter_id']}"
    fallback = "|".join(
        [
            practice,
            normalize_person_name(mapped.get("patient_name", "")),
            parse_date(mapped.get("patient_dob", "")).isoformat()
            if parse_date(mapped.get("patient_dob", "")) else "",
            clean(mapped.get("appointment_date", "")),
            clean(mapped.get("provider", "")),
        ]
    )
    digest = hashlib.sha256(fallback.encode("utf-8")).hexdigest()[:24]
    return f"{practice}|fallback|{digest}"


def ingest_appointments(
    appointments_file: str,
    queue_json: str,
    practice: str,
    source_report_name: str = "",
    reset_existing: bool = False,
    config: Optional[SyncConfig] = None,
) -> Dict[str, int]:
    # v5.4: ingest previously read the module-level DEFAULT_IGNORED_STATUSES directly
    # while process() read config.ignored_statuses. Editing ignored_statuses in
    # pf_pdf_sync_config.json therefore changed only half the pipeline. Both paths now
    # resolve the gate through is_ignored_status() against the same config.
    if config is None:
        config = SyncConfig()
    store = empty_store() if reset_existing else load_store(queue_json)
    rows = [] if reset_existing else store_rows(store)
    by_id = {row.row_id: row for row in rows}
    source_rows = read_tabular_rows(appointments_file)
    mapped_rows = [map_appointment_row(source) for source in source_rows]
    validate_appointment_report_mapping(source_rows, mapped_rows)
    if source_rows:
        actual_headers = [clean(key) for key in source_rows[0].keys() if clean(key)]
        sample = mapped_rows[0]
        print(f"Appointment report headers: {actual_headers}", flush=True)
        print(
            "Mapped sample: "
            f"date={sample.get('appointment_date', '')!r}, "
            f"patient={sample.get('patient_name', '')!r}, "
            f"status={sample.get('appointment_status', '')!r}, "
            f"type={sample.get('appointment_type', '')!r}, "
            f"provider={sample.get('provider', '')!r}",
            flush=True,
        )

    source_name = source_report_name or os.path.basename(appointments_file)
    # v5.1 could create unusable queue rows because PF's DATE/TIME header was not
    # recognized. Remove only those unprocessed, date-less rows from this same
    # source report before re-ingesting the corrected export.
    malformed_ids = {
        row.row_id for row in rows
        if not clean(row.appointment_date)
        and clean(row.source_report_name) == clean(source_name)
        and row.status != "processed"
    }
    if malformed_ids:
        rows = [row for row in rows if row.row_id not in malformed_ids]
        by_id = {row.row_id: row for row in rows}

    counts = {
        "inserted": 0, "updated": 0, "ignored": 0,
        "removed_malformed_missing_date": len(malformed_ids),
    }
    timestamp = now_iso()

    for source, mapped in zip(source_rows, mapped_rows):
        row_id = record_key(mapped, practice)
        appointment_status = normalize_status(mapped["appointment_status"])
        ignored = is_ignored_status(appointment_status, config)
        prior = by_id.get(row_id)

        if prior is None:
            prior = QueueRecord(
                row_id=row_id,
                practice=practice,
                appointment_id=mapped["appointment_id"],
                appointment_date=mapped["appointment_date"],
                appointment_status=mapped["appointment_status"],
                appointment_type=mapped["appointment_type"],
                provider=mapped["provider"],
                patient_name=mapped["patient_name"],
                patient_dob=mapped["patient_dob"],
                patient_phone=mapped["patient_phone"],
                patient_phone_normalized=normalize_phone(mapped["patient_phone"]),
                patient_id=mapped["patient_id"],
                ehr_patient_guid=mapped["ehr_patient_guid"],
                encounter_id=mapped["encounter_id"],
                status="ignored" if ignored else "ready",
                status_reason=(
                    f"ignored_appointment_status:{appointment_status}"
                    if ignored else "appointment_report_loaded"
                ),
                patient_match_status=(
                    "matched" if mapped["patient_id"] and mapped["ehr_patient_guid"] else "unmatched"
                ),
                patient_match_method=(
                    "appointment_report" if mapped["patient_id"] and mapped["ehr_patient_guid"] else ""
                ),
                source_report_name=source_name,
                source_row_json={clean(k): clean(v) for k, v in source.items()},
                created_at=timestamp,
                updated_at=timestamp,
                first_ready_at=timestamp if not ignored else "",
            )
            by_id[row_id] = prior
            counts["inserted"] += 1
        else:
            # Preserve successful encounter/PDF history but refresh report facts.
            prior.practice = practice or prior.practice
            prior.appointment_id = mapped["appointment_id"] or prior.appointment_id
            prior.appointment_date = mapped["appointment_date"] or prior.appointment_date
            prior.appointment_status = mapped["appointment_status"] or prior.appointment_status
            prior.appointment_type = mapped["appointment_type"] or prior.appointment_type
            prior.provider = mapped["provider"] or prior.provider
            prior.patient_name = mapped["patient_name"] or prior.patient_name
            prior.patient_dob = mapped["patient_dob"] or prior.patient_dob
            prior.patient_phone = mapped["patient_phone"] or prior.patient_phone
            prior.patient_phone_normalized = normalize_phone(prior.patient_phone)
            prior.patient_id = mapped["patient_id"] or prior.patient_id
            prior.ehr_patient_guid = mapped["ehr_patient_guid"] or prior.ehr_patient_guid
            prior.encounter_id = mapped["encounter_id"] or prior.encounter_id
            prior.source_report_name = source_name
            prior.source_row_json = {clean(k): clean(v) for k, v in source.items()}
            prior.updated_at = timestamp

            if ignored:
                prior.status = "ignored"
                prior.status_reason = f"ignored_appointment_status:{appointment_status}"
            elif prior.status == "ignored":
                prior.status = "ready"
                prior.status_reason = "appointment_reactivated"
            elif prior.status not in {"processed", "processing"}:
                # Review/failed/unmatched rows are eligible to be reconsidered.
                if prior.patient_match_status == "needs_attention":
                    prior.status = "needs_attention"
                else:
                    prior.status = "ready"
                    prior.status_reason = "appointment_report_reloaded"
                prior.error_message = ""
                if not prior.first_ready_at:
                    prior.first_ready_at = timestamp
            counts["updated"] += 1

        if ignored:
            counts["ignored"] += 1

    final_rows = list(by_id.values())
    final_rows.sort(
        key=lambda item: (
            parse_date(item.appointment_date) or date.min,
            normalize_person_name(item.patient_name),
            item.row_id,
        )
    )
    run_id = append_run(
        store,
        "ingest",
        {
            "appointments_file": str(Path(appointments_file).resolve()),
            "practice": practice,
            "source_rows": len(source_rows),
        },
    )
    finish_run(store, run_id, "success", counts)
    save_store(queue_json, store, final_rows)
    return counts


# ---------------------------------------------------------------------------
# Patient registry matching and persistent manual mappings
# ---------------------------------------------------------------------------


PATIENT_ALIASES: Dict[str, Tuple[str, ...]] = {
    "patient_id": (
        "patient_id", "patient id", "record_number", "record number", "prn"
    ),
    "ehr_patient_guid": (
        "ehr_patient_guid", "patient_guid", "patient guid", "pf_patient_guid", "id"
    ),
    "ehr_patient_url": (
        "ehr_patient_url", "patient url", "profile_url", "summary_url"
    ),
    "patient_name": (
        "patient_name", "patient name", "full_name", "full name", "name"
    ),
    "first_name": ("first_name", "first name"),
    "last_name": ("last_name", "last name"),
    "dob": ("dob", "date of birth", "birth date"),
    "mobile_phone": ("mobile_phone", "mobile phone", "cell phone"),
    "home_phone": ("home_phone", "home phone"),
    "work_phone": ("work_phone", "work phone"),
    "phone": ("phone", "phone number", "preferred contact"),
    "patient_status": ("patient_status", "patient status", "status"),
}


def parse_guid_from_url(value: str) -> str:
    match = re.search(r"/patients/([0-9a-fA-F-]{20,})", clean(value))
    return match.group(1) if match else ""


def map_patient_registry_row(source: Dict[str, Any]) -> Dict[str, Any]:
    normalized = {normalize_header(k): clean(v) for k, v in source.items()}
    mapped = {
        target: alias_value(normalized, aliases)
        for target, aliases in PATIENT_ALIASES.items()
    }
    if not mapped["patient_name"]:
        mapped["patient_name"] = clean(f"{mapped['first_name']} {mapped['last_name']}")
    if not mapped["ehr_patient_guid"]:
        mapped["ehr_patient_guid"] = parse_guid_from_url(mapped["ehr_patient_url"])
    phones = {
        normalize_phone(mapped[key])
        for key in ("mobile_phone", "home_phone", "work_phone", "phone")
        if normalize_phone(mapped[key])
    }
    return {
        "patient_id": mapped["patient_id"],
        "ehr_patient_guid": mapped["ehr_patient_guid"],
        "patient_name": mapped["patient_name"],
        "normalized_name": normalize_person_name(mapped["patient_name"]),
        "dob": parse_date(mapped["dob"]).isoformat() if parse_date(mapped["dob"]) else "",
        "phones": sorted(phones),
        "patient_status": mapped["patient_status"],
    }


def is_inactive_patient(patient: Dict[str, Any]) -> bool:
    """Return True only for registry rows explicitly marked inactive.

    Blank or unfamiliar statuses remain eligible so older exports without a status
    column continue to work. Practice Fusion values such as ``Inactive`` and
    ``Inactive patient`` are excluded before DOB/name/phone scoring.
    """
    status = re.sub(r"[^a-z]+", " ", clean(patient.get("patient_status")).casefold()).strip()
    return status == "inactive" or status.startswith("inactive ")


def load_patient_registry(path: str) -> List[Dict[str, Any]]:
    """Load only patient-matching fields, streaming CSV rows when possible.

    Existing PF patient exports can contain very large raw JSON/note columns. Those
    columns are irrelevant to identity matching, so they are not retained in memory.
    """
    suffix = Path(path).suffix.lower()
    patients: List[Dict[str, Any]] = []

    if suffix not in {".json", ".xlsx", ".xlsm"}:
        with open(path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError(f"Tabular file has no header row: {path}")
            for row in reader:
                patient = map_patient_registry_row(row)
                if (patient["patient_id"] or patient["ehr_patient_guid"]) and not is_inactive_patient(patient):
                    patients.append(patient)
        return patients

    for row in read_tabular_rows(path):
        patient = map_patient_registry_row(row)
        if (patient["patient_id"] or patient["ehr_patient_guid"]) and not is_inactive_patient(patient):
            patients.append(patient)
    return patients


def mapping_identity(record: QueueRecord) -> Dict[str, str]:
    return {
        "normalized_name": normalize_person_name(record.patient_name),
        "dob": parse_date(record.patient_dob).isoformat() if parse_date(record.patient_dob) else "",
        "phone": normalize_phone(record.patient_phone),
    }


def find_saved_mapping(record: QueueRecord, mappings: Sequence[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    identity = mapping_identity(record)
    candidates: List[Tuple[float, Dict[str, Any]]] = []
    for mapping in mappings:
        mapping_dob = clean(mapping.get("dob"))
        if identity["dob"] and mapping_dob and identity["dob"] != mapping_dob:
            continue
        similarity = name_similarity(record.patient_name, clean(mapping.get("patient_name")))
        if similarity < 0.90:
            continue
        mapped_phone = normalize_phone(clean(mapping.get("phone")))
        if identity["phone"] and mapped_phone and identity["phone"] != mapped_phone:
            # A phone mismatch does not automatically invalidate a manually confirmed
            # name+DOB mapping, but it reduces its priority.
            similarity -= 0.10
        candidates.append((similarity, mapping))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    best_score = candidates[0][0]
    best = [item[1] for item in candidates if abs(item[0] - best_score) < 0.001]
    distinct = {(item.get("patient_id"), item.get("ehr_patient_guid")) for item in best}
    return best[0] if len(distinct) == 1 else None


def add_or_update_mapping(
    mappings: List[Dict[str, Any]],
    record: QueueRecord,
    patient: Dict[str, Any],
    source: str,
) -> None:
    identity = mapping_identity(record)
    timestamp = now_iso()
    for mapping in mappings:
        if (
            clean(mapping.get("patient_id")) == clean(patient.get("patient_id"))
            and clean(mapping.get("normalized_name")) == identity["normalized_name"]
            and clean(mapping.get("dob")) == identity["dob"]
        ):
            mapping.update(
                {
                    "ehr_patient_guid": clean(patient.get("ehr_patient_guid")),
                    "patient_name": record.patient_name or clean(patient.get("patient_name")),
                    "normalized_name": identity["normalized_name"],
                    "dob": identity["dob"],
                    "phone": identity["phone"],
                    "source": source,
                    "updated_at": timestamp,
                }
            )
            return
    mappings.append(
        {
            "mapping_id": str(uuid.uuid4()),
            "patient_id": clean(patient.get("patient_id")),
            "ehr_patient_guid": clean(patient.get("ehr_patient_guid")),
            "patient_name": record.patient_name or clean(patient.get("patient_name")),
            "normalized_name": identity["normalized_name"],
            "dob": identity["dob"],
            "phone": identity["phone"],
            "source": source,
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    )


def apply_patient_match(
    record: QueueRecord,
    patient: Dict[str, Any],
    method: str,
    score: float,
) -> None:
    record.patient_id = clean(patient.get("patient_id")) or record.patient_id
    record.ehr_patient_guid = clean(patient.get("ehr_patient_guid")) or record.ehr_patient_guid
    record.patient_match_status = "matched"
    record.patient_match_method = method
    record.patient_match_score = round(float(score), 4)
    record.patient_match_message = ""
    record.patient_candidates = []
    record.message = ""
    if record.status == "needs_attention":
        record.status = "ready"
        record.status_reason = "patient_match_resolved"
    record.updated_at = now_iso()


def candidate_summary(patient: Dict[str, Any], score: float) -> Dict[str, Any]:
    return {
        "patient_id": patient.get("patient_id", ""),
        "ehr_patient_guid": patient.get("ehr_patient_guid", ""),
        "patient_name": patient.get("patient_name", ""),
        "dob": patient.get("dob", ""),
        "phones": patient.get("phones", []),
        "name_score": round(score, 4),
    }


def match_patients(
    queue_json: str,
    patients_file: str,
    fuzzy_threshold: float = 0.82,
    rematch_all: bool = False,
    dob_match_threshold: float = 0.85,
) -> Dict[str, int]:
    store = load_store(queue_json)
    rows = store_rows(store)
    registry = load_patient_registry(patients_file)
    dob_buckets = index_registry_by_dob(registry)
    mappings = store.setdefault("patient_mappings", [])
    counts = {
        "matched": 0,
        "reused_mapping": 0,
        "needs_attention": 0,
        "already_matched": 0,
        "ignored": 0,
    }

    for record in rows:
        if record.status == "ignored":
            counts["ignored"] += 1
            continue
        # v5.4: no longer requires patient_id. Records for the 631 registry patients
        # with no PRN never satisfied that condition and were re-scored from scratch on
        # every run.
        if (
            not rematch_all
            and record.patient_match_status == "matched"
            and record.ehr_patient_guid
        ):
            counts["already_matched"] += 1
            continue

        saved = find_saved_mapping(record, mappings)
        if saved:
            apply_patient_match(record, saved, "saved_mapping", 1.0)
            counts["reused_mapping"] += 1
            continue

        appointment_dob = (
            parse_date(record.patient_dob).isoformat() if parse_date(record.patient_dob) else ""
        )

        # v5.4: DOB first, then name.
        #
        # The previous order scored the appointment name against all 8288 registry rows
        # and only then filtered by DOB, so an exact DOB match could never rescue a name
        # that landed under fuzzy_threshold. Ten of twelve unresolved rows in the
        # 2026-07-25..30 window were in the registry with an exact DOB match, differing
        # only by a dropped middle name or second surname. Scoring inside the DOB bucket
        # lets a lower name threshold apply safely, because DOB has already done the
        # discriminating -- and it turns an 8288-row scan into a 1-3 row scan.
        if appointment_dob and appointment_dob in dob_buckets:
            bucket = dob_buckets[appointment_dob]
            scored = [
                (identity_score(record.patient_name, patient["patient_name"]), patient)
                for patient in bucket
            ]
            candidates = [
                (score, patient) for score, patient in scored if score >= dob_match_threshold
            ]
            candidates.sort(key=lambda item: item[0], reverse=True)
            name_candidates = sorted(scored, key=lambda item: item[0], reverse=True)
        else:
            # No usable DOB, or no chart carries it. Fall back to the whole registry at
            # the stricter name threshold.
            scored = [
                (identity_score(record.patient_name, patient["patient_name"]), patient)
                for patient in registry
            ]
            name_candidates = [
                (score, patient) for score, patient in scored if score >= fuzzy_threshold
            ]
            name_candidates.sort(key=lambda item: item[0], reverse=True)
            candidates = list(name_candidates)
        if len(candidates) > 1 and record.patient_phone_normalized:
            phone_candidates = [
                (score, patient)
                for score, patient in candidates
                if record.patient_phone_normalized in patient.get("phones", [])
            ]
            if phone_candidates:
                candidates = phone_candidates

        # Lowering the in-bucket threshold lets more than one chart clear the bar, so an
        # unambiguously better score should still resolve rather than escalate. Charts
        # that are genuinely indistinguishable (a duplicated chart with the same DOB and
        # the same phone) stay ambiguous and reach a human, which is the intent.
        if len(candidates) > 1:
            best_score = candidates[0][0]
            runner_up = candidates[1][0]
            leaders = [item for item in candidates if abs(item[0] - best_score) < 0.001]
            if len(leaders) == 1 and (best_score - runner_up) >= 0.08:
                candidates = leaders

        if len(candidates) == 1:
            score, patient = candidates[0]
            method = "fuzzy_name_dob_phone" if record.patient_phone_normalized else "fuzzy_name_dob"
            apply_patient_match(record, patient, method, score)
            add_or_update_mapping(mappings, record, patient, method)
            counts["matched"] += 1
            continue

        record.patient_match_status = "needs_attention"
        record.patient_match_method = ""
        record.patient_match_score = 0.0
        record.status = "needs_attention"
        record.status_reason = "patient_match_ambiguous" if candidates else "patient_match_not_found"
        if candidates:
            message = (
                f"More than one patient matched name/DOB/phone for {record.patient_name}. "
                "Assign the visible Practice Fusion patient ID manually."
            )
            display_candidates = candidates[:10]
        else:
            message = (
                f"No unique patient match was found for {record.patient_name} "
                f"DOB {record.patient_dob or '<missing>'}. Assign the patient manually."
            )
            display_candidates = name_candidates[:10]
        record.patient_match_message = message
        record.message = message
        record.patient_candidates = [
            candidate_summary(patient, score) for score, patient in display_candidates
        ]
        record.updated_at = now_iso()
        counts["needs_attention"] += 1

    run_id = append_run(
        store,
        "match-patients",
        {
            "patients_file": str(Path(patients_file).resolve()),
            "registry_rows": len(registry),
            "fuzzy_threshold": fuzzy_threshold,
        },
    )
    finish_run(store, run_id, "success", counts)
    save_store(queue_json, store, rows)
    return counts


def select_queue_rows(
    rows: Sequence[QueueRecord],
    row_id: str = "",
    appointment_id: str = "",
    patient_id: str = "",
    encounter_id: str = "",
) -> List[QueueRecord]:
    selected = []
    for record in rows:
        if row_id and record.row_id != row_id:
            continue
        if appointment_id and record.appointment_id != appointment_id:
            continue
        if patient_id and record.patient_id != patient_id:
            continue
        if encounter_id and encounter_id not in {record.encounter_id, record.encounter_key}:
            continue
        selected.append(record)
    return selected


def resolve_patient_manually(
    queue_json: str,
    patient_id: str,
    ehr_patient_guid: str,
    row_id: str = "",
    appointment_id: str = "",
    patients_file: str = "",
    resolved_patient_name: str = "",
) -> Dict[str, int]:
    store = load_store(queue_json)
    rows = store_rows(store)
    selected = select_queue_rows(rows, row_id=row_id, appointment_id=appointment_id)
    if not selected:
        raise ValueError("No queue row matched --row-id/--appointment-id.")

    patient: Dict[str, Any] = {
        "patient_id": patient_id,
        "ehr_patient_guid": ehr_patient_guid,
        "patient_name": resolved_patient_name,
    }
    if patients_file:
        registry = load_patient_registry(patients_file)
        matches: List[Dict[str, Any]] = []
        if patient_id:
            matches = [item for item in registry if item["patient_id"] == patient_id]
        # v5.4: allow lookup by GUID alone. 631 registry patients carry no PRN, so
        # requiring --patient-id made those charts impossible to resolve by either route.
        if len(matches) != 1 and ehr_patient_guid:
            guid_matches = [
                item for item in (matches or registry)
                if item["ehr_patient_guid"] == ehr_patient_guid
            ]
            if len(guid_matches) == 1:
                matches = guid_matches
        if len(matches) == 1:
            patient = matches[0]

    patient["patient_id"] = clean(patient.get("patient_id")) or patient_id
    patient["ehr_patient_guid"] = clean(patient.get("ehr_patient_guid")) or ehr_patient_guid
    if not patient["ehr_patient_guid"]:
        raise ValueError(
            "The patient GUID is required for chart navigation. Supply --ehr-patient-guid, "
            "or --patient-id together with --patients-file containing that patient ID."
        )
    if not patient["patient_id"]:
        # Not fatal: the GUID drives chart navigation and PDF naming falls back to it.
        print(
            "WARNING: resolving without a patient_id/PRN. The GUID is sufficient for "
            "chart navigation; generated PDFs will be named using the GUID.",
            flush=True,
        )

    mappings = store.setdefault("patient_mappings", [])
    primary = selected[0]
    add_or_update_mapping(mappings, primary, patient, "manual")

    # Apply the confirmed mapping to every appointment row representing the same
    # appointment identity, not only the one selected for manual resolution.
    identity = mapping_identity(primary)
    applied = 0
    for record in rows:
        record_identity = mapping_identity(record)
        # v5.4: a matching name alone is no longer enough to fan the resolution out.
        # When the resolved row had no DOB, the previous condition applied the mapping to
        # every same-name row regardless of DOB, which could attach one chart to a
        # different patient sharing a name (juniors, twins).
        same_name_dob = (
            record_identity["normalized_name"] == identity["normalized_name"]
            and bool(identity["dob"])
            and record_identity["dob"] == identity["dob"]
        )
        if record in selected or same_name_dob:
            apply_patient_match(record, patient, "manual", 1.0)
            applied += 1

    save_store(queue_json, store, rows)
    return {"resolved_rows": applied}


# ---------------------------------------------------------------------------
# Chrome profile reuse + Playwright/CDP
# ---------------------------------------------------------------------------


def find_chrome_exe(explicit: str = "") -> str:
    if explicit and os.path.isfile(explicit):
        return explicit
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/usr/bin/google-chrome",
    ]
    for candidate in candidates:
        if candidate and os.path.isfile(candidate):
            return candidate
    raise SystemExit("Could not locate Google Chrome. Pass --chrome-exe explicitly.")


def clone_profile_if_needed(args: argparse.Namespace) -> None:
    if args.attach:
        return
    destination_root = os.path.abspath(args.chrome_user_data_dir)
    destination_profile = os.path.join(destination_root, "Default")
    if os.path.isdir(destination_profile) and not args.refresh_profile:
        print(f"Reusing saved Practice Fusion Chrome profile: {destination_profile}")
        return

    os.makedirs(destination_root, exist_ok=True)
    if not args.source_user_data_dir:
        print(
            "No source profile supplied. A dedicated profile will be retained after "
            "the first manual Practice Fusion login."
        )
        return

    source_profile = os.path.join(
        os.path.abspath(args.source_user_data_dir), args.source_profile
    )
    if not os.path.isdir(source_profile):
        raise SystemExit(f"Source Chrome profile does not exist: {source_profile}")
    if args.refresh_profile and os.path.isdir(destination_profile):
        shutil.rmtree(destination_profile, ignore_errors=True)

    print(f"Cloning Chrome profile:\n  from: {source_profile}\n  to:   {destination_profile}")
    try:
        shutil.copytree(
            source_profile,
            destination_profile,
            ignore=PROFILE_CACHE_IGNORE,
            dirs_exist_ok=True,
        )
    except PermissionError as exc:
        raise SystemExit(
            "Chrome profile files are locked. Close every Chrome window and try again."
        ) from exc

    source_local_state = os.path.join(args.source_user_data_dir, "Local State")
    if os.path.isfile(source_local_state):
        shutil.copy2(source_local_state, os.path.join(destination_root, "Local State"))
    print("Profile clone complete. Later runs will reuse this dedicated profile.")


def wait_devtools(endpoint: str, timeout_seconds: float = 30.0) -> None:
    import urllib.request

    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(endpoint + "/json/version", timeout=1) as response:
                if response.status == 200:
                    return
        except Exception:
            time.sleep(0.4)
    raise SystemExit(f"Chrome DevTools endpoint did not start: {endpoint}")


def is_logged_in(page: Page) -> bool:
    try:
        url = (page.url or "").lower()
        if "securitycheck" in url or "/login" in url:
            return False
        if "#/pf" in url:
            return True
        if page.locator("a[data-tracking='Reports']").count():
            return True
    except Exception:
        pass
    return False


def find_logged_in_page(context: BrowserContext) -> Optional[Page]:
    for candidate in list(context.pages):
        if is_logged_in(candidate):
            return candidate
    return None


def _type_login_value(locator: Locator, value: str, delay_ms: int) -> None:
    """Enter a login value through regular keyboard/input events.

    The small fixed delay is for Ember-controlled input reliability; this function
    does not alter browser fingerprints or bypass Practice Fusion security checks.
    """
    locator.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    locator.click()
    locator.press("Control+A")
    locator.press("Delete")
    locator.type(value, delay=max(0, int(delay_ms)))
    locator.press("Tab")


def _find_login_page(context: BrowserContext) -> Optional[Page]:
    for candidate in list(context.pages):
        try:
            if candidate.locator(PF_PASSWORD_SELECTOR).count():
                return candidate
        except Exception:
            continue
    return None


def wait_for_pf_login(context: BrowserContext, page: Page, args: argparse.Namespace) -> Page:
    """Reuse an authenticated tab or perform the PF username/password login.

    Credentials come from --username/--password or PF_USERNAME/PF_PASSWORD. The
    function automatically continues when any Chrome tab exposes the authenticated
    Practice Fusion EHR. If PF requests OTP, it waits without requiring a console
    ENTER; OTP itself is not bypassed.
    """
    found = find_logged_in_page(context)
    if found is not None:
        found.bring_to_front()
        return found

    username = (getattr(args, "username", "") or os.getenv("PF_USERNAME", "")).strip()
    password = getattr(args, "password", "") or os.getenv("PF_PASSWORD", "")
    timeout_seconds = max(30, int(getattr(args, "login_timeout_seconds", 900)))
    typing_delay_ms = max(0, int(getattr(args, "typing_delay_ms", 65)))

    try:
        page.goto(LOGIN_URL, wait_until="domcontentloaded")
    except Exception:
        pass

    deadline = time.time() + timeout_seconds
    submitted = False
    submitted_at = 0.0
    last_status = 0.0

    while time.time() < deadline:
        found = find_logged_in_page(context)
        if found is not None:
            found.bring_to_front()
            print(f"Practice Fusion authenticated. Active tab: {found.url}")
            return found

        login_page = _find_login_page(context) or page
        login_url = (login_page.url or "").lower()

        try:
            password_field = login_page.locator(PF_PASSWORD_SELECTOR).first
            login_form_visible = password_field.count() and password_field.is_visible()
        except Exception:
            login_form_visible = False

        if login_form_visible and not submitted:
            if not username or not password:
                raise ValueError(
                    "Practice Fusion login form is visible, but credentials are missing. "
                    "Set PF_USERNAME and PF_PASSWORD before running, or pass --username "
                    "and --password (environment variables are safer)."
                )

            print(f"[{BUILD_ID}] Practice Fusion login form detected; entering saved credentials...")
            username_field = login_page.locator(PF_USERNAME_SELECTOR).first
            _type_login_value(username_field, username, typing_delay_ms)
            _type_login_value(password_field, password, typing_delay_ms)

            login_button = login_page.locator(PF_LOGIN_BUTTON_SELECTOR).first
            login_button.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
            try:
                login_button.click(timeout=DEFAULT_TIMEOUT)
            except Exception:
                login_button.evaluate("el => el.click()")
            submitted = True
            submitted_at = time.time()
            print("Credentials submitted; waiting for the authenticated EHR...")

        if "securitycheck" in login_url:
            now = time.time()
            if now - last_status >= 10:
                print(
                    "Practice Fusion security verification/OTP is open. "
                    "Waiting automatically for the authenticated EHR..."
                )
                last_status = now
        elif submitted and time.time() - submitted_at > 45 and login_form_visible:
            raise RuntimeError(
                "Practice Fusion remained on the login form after credentials were "
                "submitted. Verify PF_USERNAME/PF_PASSWORD and inspect the Chrome window "
                "for an error message or security challenge."
            )

        time.sleep(0.5)

    raise PWTimeout(
        f"Timed out after {timeout_seconds} seconds waiting for an authenticated "
        "Practice Fusion tab."
    )


def build_browser(args: argparse.Namespace):
    global _CHROME_PROC
    port = int(args.debug_port)
    endpoint = f"http://127.0.0.1:{port}"

    if args.attach:
        print(f"Attaching Playwright to existing Chrome at {endpoint}")
        wait_devtools(endpoint, timeout_seconds=8)
        playwright = sync_playwright().start()
        browser = playwright.chromium.connect_over_cdp(endpoint)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = find_logged_in_page(context) or (
            context.pages[0] if context.pages else context.new_page()
        )
        page.set_default_timeout(DEFAULT_TIMEOUT)
        page.set_default_navigation_timeout(DEFAULT_TIMEOUT)
        return playwright, context, page

    clone_profile_if_needed(args)
    user_data_dir = os.path.abspath(args.chrome_user_data_dir)
    chrome_exe = find_chrome_exe(args.chrome_exe)
    command = [
        chrome_exe,
        f"--remote-debugging-port={port}",
        f"--user-data-dir={user_data_dir}",
        "--profile-directory=Default",
        "--no-first-run",
        "--no-default-browser-check",
        "--disable-notifications",
        "--start-maximized",
        "about:blank",
    ]
    print("Launching Chrome with the reusable Practice Fusion profile...")
    _CHROME_PROC = subprocess.Popen(command)
    wait_devtools(endpoint)
    playwright = sync_playwright().start()
    browser = playwright.chromium.connect_over_cdp(endpoint)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = context.pages[0] if context.pages else context.new_page()
    page.set_default_timeout(DEFAULT_TIMEOUT)
    page.set_default_navigation_timeout(DEFAULT_TIMEOUT)
    return playwright, context, page


def close_browser(args: argparse.Namespace, playwright, context: BrowserContext, page: Page) -> None:
    global _CHROME_PROC
    if args.attach:
        print("Detaching Playwright; existing Chrome remains open.")
        try:
            playwright.stop()
        except Exception:
            pass
        return
    if args.keep_browser_open:
        print("Chrome left open because --keep-browser-open was supplied.")
        try:
            playwright.stop()
        except Exception:
            pass
        return
    print("Closing Chrome cleanly and preserving the Practice Fusion session...")
    try:
        cdp = context.new_cdp_session(page)
        cdp.send("Browser.close")
    except Exception:
        pass
    try:
        playwright.stop()
    except Exception:
        pass
    if _CHROME_PROC is not None:
        try:
            _CHROME_PROC.wait(timeout=15)
        except Exception:
            try:
                _CHROME_PROC.terminate()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Appointment report Playwright pull
# ---------------------------------------------------------------------------


def visible_match(page: Page, selector: str, max_candidates: int = 40) -> Optional[Locator]:
    """Return the first VISIBLE element matching selector, not merely the first element.

    v5.5: every visibility check in this worker used `page.locator(sel).first`, which
    resolves to the first element in DOM order regardless of visibility. The Print Chart
    page carries two elements with data-element="print-patient-content-modal" -- a hidden
    `carbon-content-modal-component` wrapper followed by the real `.content-modal` dialog
    -- so `.first` pinned the hidden wrapper and the 30s wait could never succeed. Any PF
    view that renders a hidden template alongside a live one hits the same trap.
    """
    if not clean(selector):
        return None
    try:
        locator = page.locator(selector)
        count = min(locator.count(), max_candidates)
    except Exception:
        return None
    for index in range(count):
        candidate = locator.nth(index)
        try:
            if candidate.is_visible():
                return candidate
        except Exception:
            continue
    return None


def first_visible_locator(
    page: Page,
    selectors: Sequence[str],
    timeout_ms: int = SHORT_TIMEOUT,
) -> Optional[Locator]:
    deadline = time.time() + timeout_ms / 1000
    while True:
        for selector in selectors:
            found = visible_match(page, selector)
            if found is not None:
                return found
        if time.time() >= deadline:
            return None
        time.sleep(0.2)


def require_visible_locator(
    page: Page,
    selectors: Sequence[str],
    timeout_ms: int,
    label: str,
) -> Locator:
    """Wait for a visible match or raise with the state that was actually found."""
    found = first_visible_locator(page, selectors, timeout_ms)
    if found is not None:
        return found
    diagnostics = []
    for selector in selectors:
        if not clean(selector):
            continue
        try:
            total = page.locator(selector).count()
        except Exception:
            total = -1
        diagnostics.append(f"{selector} -> {total} match(es), none visible")
    raise RuntimeError(
        f"{label} did not become visible within {timeout_ms}ms. " + "; ".join(diagnostics)
    )


def fill_date_input(page: Page, selector: str, value: date) -> None:
    locator = page.locator(selector).first
    locator.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    formatted = value.strftime("%m/%d/%Y")
    try:
        locator.click()
        locator.fill(formatted)
        locator.press("Tab")
    except Exception:
        locator.evaluate(
            """
            (el, value) => {
                el.value = value;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.blur();
            }
            """,
            formatted,
        )


def wait_report_ready(page: Page, config: AppointmentReportConfig) -> None:
    try:
        page.wait_for_load_state("networkidle", timeout=20_000)
    except Exception:
        pass
    if first_visible_locator(page, config.no_results_selectors, timeout_ms=2_000):
        return
    ready = first_visible_locator(page, config.report_ready_selectors, timeout_ms=20_000)
    if ready is None:
        raise RuntimeError(
            "The appointment report did not expose a recognized result element after Run."
        )


def copy_download_to_csv(download, output_csv: str) -> int:
    target = Path(output_csv)
    target.parent.mkdir(parents=True, exist_ok=True)
    suggested = clean(download.suggested_filename)
    temporary = target.with_name(target.stem + "_download" + Path(suggested).suffix)
    download.save_as(str(temporary))
    if temporary.suffix.lower() == ".csv":
        if temporary.resolve() != target.resolve():
            shutil.move(str(temporary), str(target))
    else:
        rows = read_tabular_rows(str(temporary))
        write_csv(str(target), rows)
        temporary.unlink(missing_ok=True)
    return len(read_tabular_rows(str(target)))


def try_export_report(
    page: Page,
    config: AppointmentReportConfig,
    output_csv: str,
) -> Optional[int]:
    export_button = first_visible_locator(
        page, config.export_report_button_selectors, timeout_ms=5_000
    )
    if export_button is None:
        return None

    try:
        with page.expect_download(timeout=7_000) as info:
            export_button.click()
        return copy_download_to_csv(info.value, output_csv)
    except Exception:
        # Some Export buttons open a format menu first.
        csv_option = first_visible_locator(
            page, config.csv_export_option_selectors, timeout_ms=3_000
        )
        if csv_option is None:
            return None
        try:
            with page.expect_download(timeout=10_000) as info:
                csv_option.click()
            return copy_download_to_csv(info.value, output_csv)
        except Exception:
            return None


def row_cells(row: Locator) -> List[str]:
    cell_selectors = "td, [role='cell'], [data-element^='data-table-cell-']"
    cells = row.locator(cell_selectors)
    values = []
    for index in range(cells.count()):
        try:
            values.append(clean(cells.nth(index).inner_text()))
        except Exception:
            values.append("")
    if values:
        return values
    try:
        text = clean(row.inner_text())
        return [text] if text else []
    except Exception:
        return []


def report_headers(page: Page, config: Optional[AppointmentReportConfig] = None) -> List[str]:
    """Read the report's column headers.

    v5.4: this previously collected every visible `th` on the whole page, so any other
    table PF happened to render (a filter panel, a summary strip) could prepend or
    interleave headers and silently shift every column in the scraped CSV. The scrape now
    searches inside config.table_selectors first -- a field that was defined but never
    referenced -- and only falls back to a page-wide search if no table matches.
    """
    header_selectors = [
        "[data-element^='data-table-header-']",
        "[role='columnheader']",
        "th",
    ]
    roots: List[Any] = []
    if config is not None:
        for table_selector in config.table_selectors:
            if not clean(table_selector):
                continue
            try:
                table = page.locator(table_selector).first
                if table.count():
                    roots.append(table)
            except Exception:
                continue
    roots.append(page)

    for root in roots:
        for selector in header_selectors:
            try:
                locator = root.locator(selector)
                count = locator.count()
            except Exception:
                continue
            headers = []
            for index in range(count):
                try:
                    if locator.nth(index).is_visible():
                        text = clean(locator.nth(index).inner_text())
                        if text:
                            headers.append(text)
                except Exception:
                    pass
            if headers:
                return headers
    return []


def collect_visible_report_rows(page: Page, config: AppointmentReportConfig) -> List[List[str]]:
    for selector in config.row_selectors:
        locator = page.locator(selector)
        rows = []
        for index in range(locator.count()):
            row = locator.nth(index)
            try:
                if not row.is_visible():
                    continue
            except Exception:
                continue
            cells = row_cells(row)
            if cells and any(cells):
                rows.append(cells)
        if rows:
            return rows
    return []


def _report_pager_state(page: Page) -> Tuple[Optional[int], Optional[int], Optional[int], str]:
    """Return (start, end, total, raw_text) for labels like
    '101 - 104 of 104 Appointments'. PF does not always expose a stable
    data-element on this report, so fall back to visible page text.
    """
    candidates = [
        "[data-element='pager-label']",
        ".pager-label",
        "[class*='pager']",
        "[class*='pagination']",
    ]
    texts: List[str] = []
    for selector in candidates:
        try:
            loc = page.locator(selector)
            for index in range(min(loc.count(), 20)):
                item = loc.nth(index)
                if item.is_visible():
                    value = clean(item.inner_text())
                    if value:
                        texts.append(value)
        except Exception:
            pass
    if not texts:
        try:
            texts.append(clean(page.locator("body").inner_text()))
        except Exception:
            pass

    pattern = re.compile(
        r"(?P<start>[0-9,]+)\s*[-–—]\s*(?P<end>[0-9,]+)\s+of\s+(?P<total>[0-9,]+)",
        flags=re.I,
    )
    for text in texts:
        match = pattern.search(text)
        if match:
            to_int = lambda value: int(value.replace(",", ""))
            return (
                to_int(match.group("start")),
                to_int(match.group("end")),
                to_int(match.group("total")),
                match.group(0),
            )
    return None, None, None, ""


def scrape_report_to_csv(page: Page, config: AppointmentReportConfig, output_csv: str) -> int:
    collected: Dict[str, List[str]] = {}
    headers = report_headers(page, config)
    page_guard = 0

    while page_guard < 300:
        page_guard += 1
        scroller = first_visible_locator(page, config.scroller_selectors, timeout_ms=1_000)
        if scroller is not None:
            try:
                scroller.evaluate("el => { el.scrollTop = 0; }")
            except Exception:
                pass
            stuck = 0
            previous_size = -1
            for _ in range(200):
                for cells in collect_visible_report_rows(page, config):
                    key = json.dumps(cells, ensure_ascii=False)
                    collected[key] = cells
                if len(collected) == previous_size:
                    stuck += 1
                else:
                    stuck = 0
                    previous_size = len(collected)
                try:
                    before = scroller.evaluate("el => el.scrollTop")
                    maximum = scroller.evaluate("el => el.scrollHeight - el.clientHeight")
                    scroller.evaluate(
                        "el => { el.scrollTop += Math.max(300, el.clientHeight * 0.7); }"
                    )
                    time.sleep(0.25)
                    after = scroller.evaluate("el => el.scrollTop")
                    if (after >= maximum - 5 or after == before) and stuck >= 2:
                        break
                except Exception:
                    break
        else:
            for cells in collect_visible_report_rows(page, config):
                collected[json.dumps(cells, ensure_ascii=False)] = cells

        shown_start, shown_end, total, pager_text = _report_pager_state(page)
        if shown_end is not None and total is not None:
            print(
                f"Appointment report page {shown_start}-{shown_end} of {total}; "
                f"collected {len(collected)} rows.",
                flush=True,
            )
            if shown_end >= total:
                break

        next_button = first_visible_locator(page, config.next_page_selectors, timeout_ms=1_500)
        if next_button is None:
            break
        try:
            if not next_button.is_enabled():
                break
        except Exception:
            pass

        previous_state = (shown_start, shown_end, total, pager_text)
        try:
            next_button.click()
        except Exception:
            break

        # Stop immediately if clicking the right-arrow does not advance the pager.
        advanced = False
        deadline = time.time() + 6
        while time.time() < deadline:
            time.sleep(0.25)
            current_state = _report_pager_state(page)
            if current_state[:3] != previous_state[:3] and current_state[1] is not None:
                advanced = True
                break
        if not advanced:
            print(
                "Appointment report pager did not advance; treating the current page as the last page.",
                flush=True,
            )
            break

    rows = list(collected.values())

    # Rows are deduplicated by full cell content because the virtual scroller re-serves
    # the same visible rows on every scroll step. Two genuinely identical appointments
    # (same patient, same slot, same provider) therefore collapse into one, and the PF
    # export carries no appointment ID to tell them apart. Surface the discrepancy rather
    # than letting the row silently disappear.
    _, _, reported_total, _ = _report_pager_state(page)
    if reported_total is not None and len(rows) < reported_total:
        print(
            f"WARNING: the report pager reported {reported_total} appointments but "
            f"{len(rows)} distinct rows were collected. Identical rows are collapsed by "
            "content because the export has no appointment ID; verify against PF before "
            "relying on this file.",
            flush=True,
        )

    width = max([len(headers), *(len(row) for row in rows)], default=0)
    if width == 0:
        write_csv(output_csv, [])
        return 0
    if len(headers) < width:
        headers = headers + [f"column_{index + 1}" for index in range(len(headers), width)]
    headers = [header or f"column_{index + 1}" for index, header in enumerate(headers[:width])]
    dictionaries = [
        {headers[index]: row[index] if index < len(row) else "" for index in range(width)}
        for row in rows
    ]
    write_csv(output_csv, dictionaries)
    return len(dictionaries)


def save_report_diagnostics(page: Page, output_csv: str) -> Tuple[str, str]:
    target = Path(output_csv)
    html_path = str(target.with_suffix(".diagnostic.html"))
    png_path = str(target.with_suffix(".diagnostic.png"))
    try:
        Path(html_path).write_text(page.content(), encoding="utf-8")
    except Exception:
        html_path = ""
    try:
        page.screenshot(path=png_path, full_page=True)
    except Exception:
        png_path = ""
    return html_path, png_path


def pull_appointment_report_on_page(
    page: Page,
    config: AppointmentReportConfig,
    start_date: date,
    end_date: date,
    output_csv: str,
) -> Dict[str, Any]:
    if start_date > end_date:
        raise ValueError("start_date cannot be after end_date")

    try:
        reports = page.locator(config.reports_menu_selector).first
        reports.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
        reports.click()
        report_link = page.locator(config.appointment_report_link_selector).first
        report_link.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
        report_link.click()

        page.locator(config.report_start_date_selector).first.wait_for(
            state="visible", timeout=DEFAULT_TIMEOUT
        )
        fill_date_input(page, config.report_start_date_selector, start_date)
        fill_date_input(page, config.report_end_date_selector, end_date)
        print(
            f"Appointment report date range: {start_date.strftime('%m/%d/%Y')} "
            f"through {end_date.strftime('%m/%d/%Y')}",
            flush=True,
        )

        run_button = first_visible_locator(
            page, config.run_report_button_selectors, timeout_ms=10_000
        )
        if run_button is None:
            raise RuntimeError("Could not find a visible Run Report button.")
        run_button.click()
        wait_report_ready(page, config)

        if first_visible_locator(page, config.no_results_selectors, timeout_ms=1_000):
            write_csv(output_csv, [])
            return {"rows": 0, "method": "no_results", "output_csv": output_csv}

        exported = try_export_report(page, config, output_csv)
        if exported is not None:
            return {"rows": exported, "method": "download", "output_csv": output_csv}

        print(
            "Export did not produce a browser download; scraping all report pages instead.",
            flush=True,
        )
        scraped = scrape_report_to_csv(page, config, output_csv)
        return {"rows": scraped, "method": "dom_scrape", "output_csv": output_csv}
    except Exception as exc:
        html_path, png_path = save_report_diagnostics(page, output_csv)
        raise RuntimeError(
            f"Appointment report pull failed: {exc}. Diagnostics: {html_path or '<none>'}, "
            f"{png_path or '<none>'}"
        ) from exc


# ---------------------------------------------------------------------------
# Encounter discovery, Print Chart selection, and PDF generation
# ---------------------------------------------------------------------------


def patient_summary_url(patient_guid: str) -> str:
    return f"{EHR_BASE_URL}#/PF/charts/patients/{patient_guid}/summary"


def checkbox_state_is(box: Locator, desired: bool) -> bool:
    try:
        return box.is_checked() == desired
    except Exception:
        return False


def checkbox_click_targets(box: Locator) -> List[Tuple[str, Locator]]:
    """Clickable stand-ins for a visually hidden checkbox input, best first."""
    targets: List[Tuple[str, Locator]] = []
    for name, selector in (
        ("sibling label", "xpath=following-sibling::label[1]"),
        ("parent label", "xpath=../label"),
        ("parent container", "xpath=.."),
    ):
        try:
            candidate = box.locator(selector).first
            if candidate.count() and candidate.is_visible():
                targets.append((name, candidate))
        except Exception:
            continue
    return targets


def checkbox_is_interactive(box: Locator) -> bool:
    """True when the box can be driven, whether or not the input itself renders."""
    try:
        if box.is_visible():
            return True
    except Exception:
        pass
    return bool(checkbox_click_targets(box))


def set_checkbox_state(box: Locator, desired: bool, label: str = "checkbox") -> None:
    """Force a styled checkbox into the desired state and verify it took.

    v5.6: Practice Fusion renders `check-box__input` inputs that are visually hidden and
    draws the control on the associated `check-box__label`. Playwright's check()/click()
    require visibility, so every section click failed with "element is not visible" even
    though the input resolved correctly.

    Clicking the label is not universally safe either: the notes row nests the dropdown
    toggle anchor INSIDE its label, so a center-click on that label opens the dropdown
    instead of toggling the group. So the input itself is driven first with force=True,
    then the label near its left edge where the glyph is drawn, then the center, then the
    DOM property with input/change events for Ember to observe. State is re-read after
    every attempt and the first success returns, so nothing is ever toggled twice.
    """
    box.wait_for(state="attached", timeout=DEFAULT_TIMEOUT)
    if checkbox_state_is(box, desired):
        return

    attempts: List[str] = []

    try:
        if desired:
            box.check(force=True, timeout=SHORT_TIMEOUT)
        else:
            box.uncheck(force=True, timeout=SHORT_TIMEOUT)
        if checkbox_state_is(box, desired):
            return
        attempts.append("force check/uncheck on input: state unchanged")
    except Exception as exc:
        attempts.append(f"force check/uncheck on input: {type(exc).__name__}")

    for name, target in checkbox_click_targets(box):
        # The glyph is drawn at the label's left edge; the center may be a nested link.
        for description, position in (("glyph", {"x": 6, "y": 8}), ("center", None)):
            try:
                if position is not None:
                    target.click(position=position, timeout=SHORT_TIMEOUT)
                else:
                    target.click(timeout=SHORT_TIMEOUT)
                if checkbox_state_is(box, desired):
                    return
                attempts.append(f"{name} {description} click: state unchanged")
            except Exception as exc:
                attempts.append(f"{name} {description} click: {type(exc).__name__}")

    try:
        box.evaluate(
            """
            (el, want) => {
                if (el.checked !== want) {
                    el.checked = want;
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                    el.dispatchEvent(new Event('click', {bubbles: true}));
                }
            }
            """,
            desired,
        )
        if checkbox_state_is(box, desired):
            return
        attempts.append("DOM property + input/change events: state unchanged")
    except Exception as exc:
        attempts.append(f"DOM property + input/change events: {type(exc).__name__}")

    raise RuntimeError(
        f"CHECKBOX_NOT_SETTABLE: could not set {label} to checked={desired}. "
        "Tried -> " + "; ".join(attempts)
    )


def ensure_checked(locator: Locator, label: str = "checkbox") -> None:
    set_checkbox_state(locator, True, label)


def ensure_unchecked(locator: Locator, label: str = "checkbox") -> None:
    set_checkbox_state(locator, False, label)


def locator_text(locator: Locator) -> str:
    try:
        return clean(locator.inner_text())
    except Exception:
        try:
            return clean(locator.text_content() or "")
        except Exception:
            return ""


def read_encounters(
    page: Page,
    config: SyncConfig,
    patient_guid: str,
    item_selector: str,
    source: str,
) -> List[DetectedEncounter]:
    try:
        page.locator(item_selector).first.wait_for(state="attached", timeout=10_000)
    except Exception:
        pass
    items = page.locator(item_selector)
    output: List[DetectedEncounter] = []
    for index in range(items.count()):
        item = items.nth(index)
        try:
            title = clean(item.get_attribute("title") or "")
            display = locator_text(item)
        except Exception:
            continue
        soap_marker = clean(config.encounter_soap_title_text).lower()
        if soap_marker and soap_marker not in f"{title} {display}".lower():
            continue
        encounter_date_text = locator_text(item.locator(config.encounter_date_selector).first)
        encounter_date = parse_date(encounter_date_text)
        if encounter_date is None:
            continue
        encounter_type = locator_text(item.locator(config.encounter_type_selector).first)
        encounter_code = locator_text(item.locator(config.encounter_code_selector).first)
        complaint = locator_text(
            item.locator(config.encounter_chief_complaint_selector).first
        )
        key_source = "|".join(
            [
                patient_guid.lower(),
                encounter_date.isoformat(),
                encounter_type.lower(),
                encounter_code.lower(),
                complaint.lower(),
            ]
        )
        output.append(
            DetectedEncounter(
                encounter_date=encounter_date.isoformat(),
                encounter_type=encounter_type,
                encounter_code=encounter_code,
                chief_complaint=complaint,
                display_text=display,
                encounter_key=hashlib.sha256(key_source.encode("utf-8")).hexdigest(),
                source=source,
            )
        )
    return output


def read_summary_encounters(page: Page, config: SyncConfig, patient_guid: str) -> List[DetectedEncounter]:
    return read_encounters(
        page, config, patient_guid, config.encounter_item_selector, "summary"
    )


def open_all_encounters_timeline(
    page: Page, config: SyncConfig, patient_guid: str
) -> bool:
    link = page.locator(config.encounter_timeline_link_selector).first
    if link.count() == 0:
        return False
    href = clean(link.get_attribute("href") or "")
    if href.startswith("#"):
        page.goto(EHR_BASE_URL + href, wait_until="domcontentloaded")
    else:
        link.click()
    deadline = time.time() + 15
    while time.time() < deadline:
        if "/timeline/encounter" in (page.url or ""):
            return True
        time.sleep(0.2)
    return "/timeline/encounter" in (page.url or "")


def all_patient_encounters(
    page: Page,
    config: SyncConfig,
    patient_guid: str,
    include_timeline: bool = True,
) -> List[DetectedEncounter]:
    encounters = read_summary_encounters(page, config, patient_guid)
    if include_timeline and open_all_encounters_timeline(page, config, patient_guid):
        timeline = read_encounters(
            page,
            config,
            patient_guid,
            config.encounter_timeline_item_selector,
            "timeline",
        )
        by_key = {item.encounter_key: item for item in encounters}
        for item in timeline:
            by_key[item.encounter_key] = item
        encounters = list(by_key.values())
    encounters.sort(
        key=lambda item: parse_date(item.encounter_date) or date.min,
        reverse=True,
    )
    return encounters


def find_encounter_for_appointment(
    page: Page,
    config: SyncConfig,
    patient_guid: str,
    appointment_date: str,
) -> DetectedEncounter:
    """Find the appointment-date encounter from the patient Summary only.

    Nightly appointment processing must never navigate to Timeline as a fallback.
    Practice Fusion can leave the automation waiting on the Timeline route until a
    manual return to Summary. If the encounter is absent from Summary, the queue row
    is left ready/review for a later poll and no PDF is generated.
    """
    requested_date = require_date(appointment_date, "appointment date")
    matches = [
        item
        for item in read_summary_encounters(page, config, patient_guid)
        if parse_date(item.encounter_date) == requested_date
    ]
    if not matches:
        raise EncounterNotFoundError(
            "ENCOUNTER_NOT_FOUND_FOR_APPOINTMENT_DATE_AFTER_SUMMARY_CHECK: "
            f"appointment_date={appointment_date}"
        )
    return matches[0]


def open_print_chart(page: Page, config: SyncConfig) -> Locator:
    """Click Print Chart and return the visible modal container."""
    button = page.locator(config.print_chart_button_selector).first
    button.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    button.click()
    return require_visible_locator(
        page,
        config.print_modal_ready_selectors,
        DEFAULT_TIMEOUT,
        "Print Chart modal",
    )


def close_print_chart(page: Page, config: SyncConfig) -> None:
    """Dismiss the Print Chart modal if it is still open.

    Best effort by design -- navigation is still the primary reset. Tries Cancel before
    the header X, since Cancel is the documented dismissal control.
    """
    for selector in (config.print_modal_cancel_selector, config.print_modal_close_selector):
        if not clean(selector):
            continue
        try:
            control = visible_match(page, selector)
            if control is None:
                continue
            control.click()
            for _ in range(10):
                if first_visible_locator(page, config.print_modal_ready_selectors, 200) is None:
                    return
                time.sleep(0.2)
        except Exception:
            continue


def modal_checkbox_state(modal: Locator, config: SyncConfig) -> Dict[str, bool]:
    """Map every option checkbox in the modal to its checked state.

    Keyed by the owning data-element so mismatches can be reported in PF's own
    vocabulary rather than as opaque selectors. The notes group checkbox is excluded:
    notes are selected separately and per record.
    """
    state: Dict[str, bool] = {}
    try:
        boxes = modal.locator("input[type='checkbox']")
        count = boxes.count()
    except Exception:
        return state
    for index in range(count):
        box = boxes.nth(index)
        try:
            owner = box.evaluate(
                """
                el => {
                    const host = el.closest('[data-element]');
                    return host ? (host.getAttribute('data-element') || '') : '';
                }
                """
            )
            owner = clean(owner)
            if not owner:
                continue
            # Notes are selected separately and per record.
            if "notes-dropdown" in owner:
                continue
            state[owner] = bool(box.is_checked())
        except Exception:
            continue
    return state


def clear_all_facesheet_sections(page: Page, config: SyncConfig, modal: Locator) -> bool:
    """Clear every section using the modal's own 'none' link, falling back to unchecking."""
    link = visible_match(page, config.print_modal_select_none_selector)
    if link is not None:
        try:
            link.click()
            time.sleep(0.3)
            if not any(modal_checkbox_state(modal, config).values()):
                return True
        except Exception:
            pass
    # Fallback: uncheck whatever is currently checked.
    cleared = True
    for owner, checked in modal_checkbox_state(modal, config).items():
        if not checked:
            continue
        try:
            box = modal.locator(f"[data-element='{owner}'] input[type='checkbox']").first
            ensure_unchecked(box, owner)
        except Exception:
            cleared = False
    return cleared


def select_facesheet_sections(page: Page, config: SyncConfig, modal: Optional[Locator] = None) -> List[str]:
    """Select exactly the configured sections and nothing else.

    v5.5: this used to only CHECK the configured selectors, leaving anything PF had
    pre-selected -- or anything a previous record left behind -- silently included in the
    printed chart. The modal's own "Select: none" link now clears the form first, the
    configured set is checked, and the resulting state is verified against the intended
    set. An option PF adds in a future release therefore defaults to OFF and surfaces as a
    named mismatch instead of quietly appearing in patient charts.
    """
    if not config.facesheet_checkbox_selectors:
        raise RuntimeError("No Facesheet checkbox selectors are configured.")
    if modal is None:
        modal = require_visible_locator(
            page, config.print_modal_ready_selectors, SHORT_TIMEOUT, "Print Chart modal"
        )

    if not clear_all_facesheet_sections(page, config, modal):
        print("  WARNING: could not confirm all Print Chart sections were cleared.", flush=True)

    selected: List[str] = []
    for selector in config.facesheet_checkbox_selectors:
        if not clean(selector):
            continue
        checkbox = modal.locator(selector).first
        if checkbox.count() == 0:
            raise RuntimeError(f"Facesheet checkbox not found in the modal: {selector}")
        ensure_checked(checkbox, selector)
        selected.append(selector)

    intended = set()
    for selector in config.facesheet_checkbox_selectors:
        match = re.search(r"data-element='([^']+)'", selector)
        if match:
            intended.add(match.group(1))

    state = modal_checkbox_state(modal, config)
    missing = sorted(owner for owner in intended if not state.get(owner, False))
    unexpected = sorted(owner for owner, checked in state.items() if checked and owner not in intended)
    if missing or unexpected:
        message = (
            "Print Chart section selection does not match the configured set. "
            f"expected_checked={sorted(intended)} missing={missing} unexpected_checked={unexpected}"
        )
        if config.enforce_exact_facesheet_selection:
            raise RuntimeError("FACESHEET_SELECTION_MISMATCH: " + message)
        print(f"  WARNING: {message}", flush=True)
    return selected


def note_date_tokens(appointment_date: str, formats: Iterable[str]) -> List[str]:
    """Build the date strings to look for in Print Chart note labels.

    v5.4: the previous de-zeroing used token.replace("/0", "/"), which can never touch a
    leading-zero month because there is no preceding slash. "07/27/2026" produced
    "07/27/2026" and "07/27/26" but never "7/27/2026", so a PF note label rendered
    without a leading zero raised SoapNoteNotFoundError.
    """
    parsed = require_date(appointment_date, "appointment date")
    tokens: List[str] = []
    for fmt in formats:
        token = parsed.strftime(fmt)
        variants = {
            token,
            token.replace("/0", "/").replace(" 0", " "),
            # Strip a leading zero on the first component as well.
            re.sub(r"^0(?=\d)", "", token),
            re.sub(r"^0(?=\d)", "", token).replace("/0", "/").replace(" 0", " "),
        }
        tokens.extend(variants)
    return sorted({clean(token) for token in tokens if clean(token)}, key=len, reverse=True)


def checkbox_display_text(checkbox: Locator) -> str:
    try:
        return clean(
            checkbox.evaluate(
                """
                el => {
                    const id = el.id || '';
                    let label = null;
                    if (id && window.CSS && CSS.escape) {
                        label = document.querySelector(`label[for="${CSS.escape(id)}"]`);
                    }
                    const container = label || el.closest(
                        'label, li, [role="option"], .checkbox-row, .check-box'
                    );
                    return container ? (container.innerText || container.textContent || '') : '';
                }
                """
            )
        )
    except Exception:
        return ""


def notes_toggle_label(page: Page, config: SyncConfig) -> str:
    toggle = visible_match(page, config.notes_dropdown_toggle_selector)
    if toggle is None:
        return ""
    return locator_text(toggle)


def notes_selection_is_empty(page: Page, config: SyncConfig) -> bool:
    """True when the notes toggle still reads as an empty selection."""
    label = notes_toggle_label(page, config).lower()
    empty_text = clean(config.notes_empty_label_text).lower()
    return bool(empty_text) and empty_text in label


def open_notes_dropdown(page: Page, config: SyncConfig) -> None:
    toggle = require_visible_locator(
        page, [config.notes_dropdown_toggle_selector], DEFAULT_TIMEOUT, "Notes dropdown toggle"
    )
    toggle.click()
    time.sleep(0.5)


def note_option_checkboxes(page: Page, config: SyncConfig) -> List[Locator]:
    """Visible note checkboxes from the tethered dropdown panel.

    The panel is tethered outside the modal, so it cannot be scoped to the modal
    container. Scoping to a known panel selector is preferred; the page-wide fallback is
    the pre-v5.5 behavior and relies on the date-token filter to exclude the modal's own
    section checkboxes.
    """
    containers: List[Any] = []
    for selector in config.note_dropdown_menu_selectors:
        panel = visible_match(page, selector)
        if panel is not None:
            containers.append(panel)
            break
    if not containers:
        containers.append(page)

    option_selector = config.note_option_selector or "input[type='checkbox']"
    found: List[Locator] = []
    for container in containers:
        try:
            boxes = container.locator(option_selector)
            count = boxes.count()
        except Exception:
            continue
        for index in range(count):
            box = boxes.nth(index)
            # v5.6: filter on whether the box is DRIVABLE, not whether the input renders.
            # Note options use the same hidden-input/visible-label pattern as the modal
            # sections, so an is_visible() filter discarded every note.
            if checkbox_is_interactive(box):
                found.append(box)
    return found


def select_all_notes(page: Page, config: SyncConfig) -> str:
    """Check every note via the notes row's group checkbox.

    The notes row is an Ember select-all-checkbox-dropdown: the checkbox beside the
    "No notes selected" toggle applies to the whole group, so one click selects every
    note without enumerating the tethered panel. Verified by reading the toggle label,
    and falls back to checking each option individually if the group click does not take.
    """
    # v5.6: the group checkbox is a visually hidden input, so a visibility-based lookup
    # could never find it and this path always fell through to enumeration.
    group = page.locator(config.notes_group_checkbox_selector).first
    try:
        if group.count():
            ensure_checked(group, "notes group checkbox")
            time.sleep(0.4)
            if not notes_selection_is_empty(page, config):
                return clean(notes_toggle_label(page, config)) or "all notes"
    except Exception as exc:
        print(f"  notes group checkbox unavailable ({type(exc).__name__}); enumerating instead", flush=True)

    # Fallback: enumerate the dropdown and check everything in it.
    open_notes_dropdown(page, config)
    boxes = note_option_checkboxes(page, config)
    labels: List[str] = []
    for box in boxes:
        text = checkbox_display_text(box)
        if not text:
            continue
        try:
            ensure_checked(box, f"note option {text[:60]!r}")
            labels.append(clean(text))
        except Exception:
            continue
    if not labels:
        raise SoapNoteNotFoundError(
            "SOAP_NOTE_NOT_FOUND_ANY: the notes dropdown exposed no selectable notes."
        )
    if notes_selection_is_empty(page, config):
        raise SoapNoteNotFoundError(
            "SOAP_NOTE_SELECTION_NOT_APPLIED: notes were clicked but the toggle still "
            f"reads {notes_toggle_label(page, config)!r}."
        )
    unique: List[str] = []
    for label in labels:
        if label not in unique:
            unique.append(label)
    return " | ".join(unique)


def select_soap_note_for_date(page: Page, config: SyncConfig, appointment_date: str) -> str:
    open_notes_dropdown(page, config)
    tokens = note_date_tokens(appointment_date, config.note_date_formats)
    matches: List[Tuple[Locator, str]] = []
    for checkbox in note_option_checkboxes(page, config):
        text = checkbox_display_text(checkbox)
        if text and any(token.lower() in text.lower() for token in tokens):
            matches.append((checkbox, text))
    if not matches:
        raise SoapNoteNotFoundError(
            f"SOAP_NOTE_NOT_FOUND_FOR_DATE: appointment_date={appointment_date}; tried={tokens}"
        )
    selected = []
    seen = set()
    for checkbox, text in matches:
        ensure_checked(checkbox, f"note option {text[:60]!r}")
        normalized = clean(text)
        if normalized and normalized not in seen:
            selected.append(normalized)
            seen.add(normalized)
    if notes_selection_is_empty(page, config):
        raise SoapNoteNotFoundError(
            "SOAP_NOTE_SELECTION_NOT_APPLIED: matching notes were clicked but the toggle "
            f"still reads {notes_toggle_label(page, config)!r}."
        )
    return " | ".join(selected)


def patient_has_prior_pdf(record: QueueRecord, all_rows: Sequence[QueueRecord]) -> bool:
    """Has a chart PDF already been produced for this patient?

    Keyed on ehr_patient_guid because that is the identifier every matched row carries;
    631 registry patients have no PRN. The record being processed is excluded so a retry
    of the same appointment does not count as prior history.
    """
    guid = clean(record.ehr_patient_guid)
    if not guid:
        return False
    for other in all_rows:
        if other.row_id == record.row_id:
            continue
        if clean(other.ehr_patient_guid) != guid:
            continue
        if other.status == "processed" and clean(other.pdf_path):
            return True
    return False


def resolve_notes_mode(
    record: QueueRecord,
    config: SyncConfig,
    all_rows: Sequence[QueueRecord],
) -> str:
    """Decide between every note and only the appointment date's note.

    "auto" selects every note the first time a patient is printed, so the initial chart
    carries the full note history, then narrows to the appointment date on later runs so
    each subsequent PDF adds only the new visit.
    """
    mode = clean(config.notes_selection_mode).lower() or "auto"
    if mode in {"all", "date"}:
        return mode
    if mode != "auto":
        print(
            f"  WARNING: unknown notes_selection_mode {mode!r}; treating it as 'auto'.",
            flush=True,
        )
    return "date" if patient_has_prior_pdf(record, all_rows) else "all"


def select_notes_for_record(
    page: Page,
    config: SyncConfig,
    record: QueueRecord,
    all_rows: Sequence[QueueRecord],
) -> Tuple[str, str]:
    mode = resolve_notes_mode(record, config, all_rows)
    if mode == "all":
        print("  notes: first chart for this patient -> selecting all SOAP notes", flush=True)
        return mode, select_all_notes(page, config)
    print(
        f"  notes: prior chart exists -> selecting notes dated {record.appointment_date}",
        flush=True,
    )
    return mode, select_soap_note_for_date(page, config, record.appointment_date)


def format_pdf_name(record: QueueRecord, config: SyncConfig) -> str:
    values = {
        "patient_id": record.patient_id or record.ehr_patient_guid,
        "appointment_date": (
            parse_date(record.appointment_date).isoformat()
            if parse_date(record.appointment_date) else record.appointment_date
        ),
        "encounter_id": record.encounter_id or record.encounter_key or "encounter",
        "appointment_id": record.appointment_id or "appointment",
    }
    try:
        name = config.download_filename_template.format(**values)
    except Exception:
        name = f"{values['patient_id']}_{values['appointment_date']}_{values['encounter_id']}.pdf"
    name = safe_filename(name)
    return name if name.lower().endswith(".pdf") else name + ".pdf"


def _mark_current_visible_print_links(page: Page, selector: str, marker: str) -> int:
    """Mark print links that already exist before PF creates the chart preview.

    The patient Summary page itself contains several generic ``a.print-link`` icons.
    Waiting on ``a.print-link[title='Print']`` therefore matched the Summary page before
    the printable chart overlay had rendered, causing CDP to save the wrong screen.
    Only links visible before clicking the modal Print button are marked, so a preview
    link that is inserted later -- or was pre-rendered but hidden -- remains eligible.
    """
    try:
        return int(
            page.evaluate(
                """
                ({selector, marker}) => {
                    let count = 0;
                    for (const el of document.querySelectorAll(selector)) {
                        const style = getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        const visible = style.display !== 'none' &&
                            style.visibility !== 'hidden' &&
                            rect.width > 0 && rect.height > 0;
                        if (visible) {
                            el.setAttribute('data-pf-sync-preexisting-print-link', marker);
                            count += 1;
                        }
                    }
                    return count;
                }
                """,
                {"selector": selector, "marker": marker},
            )
        )
    except Exception:
        return 0


def _new_visible_print_preview_link(
    page: Page, selector: str, marker: str, timeout_ms: int
) -> Locator:
    """Wait for the new print control belonging to PF's printable chart overlay."""
    deadline = time.time() + timeout_ms / 1000.0
    while time.time() < deadline:
        links = page.locator(selector)
        try:
            count = links.count()
        except Exception:
            count = 0
        for index in range(count):
            link = links.nth(index)
            try:
                if not link.is_visible():
                    continue
                if link.get_attribute('data-pf-sync-preexisting-print-link') == marker:
                    continue
                return link
            except Exception:
                continue
        time.sleep(0.2)
    raise PWTimeout(
        "Printable chart preview did not appear. Existing Summary-page print icons were "
        "ignored, but no new preview Print control became visible."
    )


def _find_and_mark_print_document(
    page: Page,
    record: QueueRecord,
    config: SyncConfig,
    timeout_ms: int = DEFAULT_TIMEOUT,
):
    """Locate the actual PF printable SOAP document, not the Summary page.

    PF keeps the browser URL on ``/summary`` and renders the chart preview as a
    body-level overlay (and, in some builds, inside a same-origin iframe).  The blue
    printer link is only a toolbar control and is not necessarily a descendant of the
    printable document.  Therefore, finding an ancestor of that link is unreliable.

    Instead, search every same-origin frame for the smallest visible element containing
    the distinctive printable-chart text markers and the encounter date.  The Summary
    page cannot satisfy this guard because it does not contain the combined
    PATIENT/FACILITY/ENCOUNTER/NOTE TYPE headings plus SOAP body headings.
    """
    date_tokens = note_date_tokens(record.appointment_date, config.note_date_formats)
    date_tokens = [clean(token).upper() for token in date_tokens if clean(token)]
    deadline = time.time() + timeout_ms / 1000.0
    last_diagnostics: List[Dict[str, Any]] = []

    finder_js = r"""
        ({dateTokens}) => {
            const normalize = value => String(value || '')
                .replace(/\u00a0/g, ' ')
                .replace(/\s+/g, ' ')
                .trim()
                .toUpperCase();

            document.querySelectorAll('[data-pf-sync-print-document-root]')
                .forEach(el => el.removeAttribute('data-pf-sync-print-document-root'));

            const selectors = [
                'main', 'article', 'section', 'form', 'table',
                'div', 'body'
            ].join(',');
            const candidates = [];
            const viewportArea = Math.max(1, innerWidth * innerHeight);

            for (const el of document.querySelectorAll(selectors)) {
                const style = getComputedStyle(el);
                const rect = el.getBoundingClientRect();
                if (style.display === 'none' || style.visibility === 'hidden' ||
                    rect.width < 250 || rect.height < 250) {
                    continue;
                }

                const text = normalize(el.innerText || el.textContent || '');
                if (text.length < 350) continue;

                const hasPatient = /(^|\s)PATIENT(\s|$)/.test(text);
                const hasFacility = /(^|\s)FACILITY(\s|$)/.test(text);
                const hasEncounter = /(^|\s)ENCOUNTER(\s|$)/.test(text);
                const hasNoteType = text.includes('NOTE TYPE') || text.includes('SOAP NOTE');
                const hasClinicalBody = [
                    'SUBJECTIVE', 'OBJECTIVE', 'ASSESSMENT', 'PLAN',
                    'CHIEF COMPLAINT', 'VITALS FOR THIS ENCOUNTER'
                ].some(token => text.includes(token));
                const hasDate = !dateTokens.length || dateTokens.some(token => text.includes(token));

                // This conjunction deliberately excludes the underlying patient Summary.
                if (!(hasPatient && hasFacility && hasEncounter && hasNoteType &&
                      hasClinicalBody && hasDate)) {
                    continue;
                }

                const position = style.position || '';
                const z = Number.parseInt(style.zIndex || '0', 10) || 0;
                const area = Math.max(1, rect.width * rect.height);
                let score = 1000;
                score += hasClinicalBody ? 200 : 0;
                score += hasDate ? 150 : 0;
                score += (position === 'fixed' || position === 'absolute') ? 60 : 0;
                score += z > 100 ? 40 : 0;
                score += style.backgroundColor && style.backgroundColor !== 'rgba(0, 0, 0, 0)' ? 15 : 0;
                // Prefer the tightest element that still contains the whole print document.
                score -= Math.min(200, Math.round((area / viewportArea) * 10));
                if (el === document.body) score -= 300;

                candidates.push({el, score, textLength: text.length, rect, z, position, text});
            }

            candidates.sort((a, b) =>
                (b.score - a.score) ||
                (a.textLength - b.textLength) ||
                ((a.rect.width * a.rect.height) - (b.rect.width * b.rect.height))
            );

            if (!candidates.length) {
                const bodyText = normalize(document.body?.innerText || '');
                return {
                    found: false,
                    url: location.href,
                    bodyTextLength: bodyText.length,
                    bodyHasFacility: bodyText.includes('FACILITY'),
                    bodyHasNoteType: bodyText.includes('NOTE TYPE'),
                    bodyHasSubjective: bodyText.includes('SUBJECTIVE'),
                    bodyHasObjective: bodyText.includes('OBJECTIVE'),
                    dateTokens
                };
            }

            const chosen = candidates[0];

            // The tightest matching element is commonly one DIV.print-section.
            // For an "all SOAP notes" print, PF renders many sibling print sections
            // inside one scrollable modal. Promote the match to the smallest ancestor
            // containing every available print section so PDF generation captures the
            // whole chart rather than only the first encounter.
            let promoted = chosen.el;
            let cursor = chosen.el.parentElement;
            while (cursor && cursor !== document.body && cursor !== document.documentElement) {
                const sectionCount = cursor.querySelectorAll('.print-section').length;
                if (sectionCount >= 2) {
                    promoted = cursor;
                    break;
                }
                cursor = cursor.parentElement;
            }

            // If PF has only one section in this print job, still promote to the nearest
            // scroll-clamping ancestor/modal rather than cloning the leaf section.
            if (promoted === chosen.el) {
                cursor = chosen.el.parentElement;
                while (cursor && cursor !== document.body && cursor !== document.documentElement) {
                    const style = getComputedStyle(cursor);
                    const hasScrollClamp = cursor.scrollHeight > cursor.clientHeight + 8 ||
                        ['auto', 'scroll', 'hidden'].includes(style.overflowY);
                    const hasPreviewToolbar = !!cursor.querySelector(
                        'a.print-link[title="Print"], .glyphicon-print, .glyphicon-remove, .icon-go-away'
                    );
                    if (hasScrollClamp || hasPreviewToolbar) {
                        promoted = cursor;
                        break;
                    }
                    cursor = cursor.parentElement;
                }
            }

            const promotedRect = promoted.getBoundingClientRect();
            const promotedStyle = getComputedStyle(promoted);
            const promotedText = normalize(promoted.innerText || promoted.textContent || '');
            const sectionCount = promoted.querySelectorAll('.print-section').length ||
                (promoted.matches('.print-section') ? 1 : 0);

            promoted.setAttribute('data-pf-sync-print-document-root', 'true');
            return {
                found: true,
                url: location.href,
                tag: promoted.tagName,
                className: String(promoted.className || ''),
                width: Math.round(promotedRect.width),
                height: Math.round(promotedRect.height),
                scrollHeight: Math.round(promoted.scrollHeight || promotedRect.height),
                clientHeight: Math.round(promoted.clientHeight || promotedRect.height),
                sectionCount,
                textLength: promotedText.length,
                zIndex: Number.parseInt(promotedStyle.zIndex || '0', 10) || 0,
                position: promotedStyle.position || '',
                textPreview: promotedText.slice(0, 240)
            };
        }
    """

    while time.time() < deadline:
        last_diagnostics = []
        for frame in list(page.frames):
            try:
                result = frame.evaluate(finder_js, {"dateTokens": date_tokens})
            except Exception as exc:
                last_diagnostics.append({"url": getattr(frame, "url", ""), "error": str(exc)[:200]})
                continue
            if isinstance(result, dict):
                last_diagnostics.append(result)
            if isinstance(result, dict) and result.get("found"):
                root = frame.locator("[data-pf-sync-print-document-root='true']").first
                root.wait_for(state="visible", timeout=SHORT_TIMEOUT)
                return frame, root, result
        time.sleep(0.25)

    # Save evidence that is useful if PF changes the preview structure again.
    debug_dir = Path("pf_sync_debug")
    debug_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        page.screenshot(path=str(debug_dir / f"print_preview_not_found_{stamp}.png"), full_page=True)
    except Exception:
        pass
    try:
        (debug_dir / f"print_preview_not_found_{stamp}.html").write_text(
            page.content(), encoding="utf-8"
        )
    except Exception:
        pass
    try:
        (debug_dir / f"print_preview_not_found_{stamp}.json").write_text(
            json.dumps(last_diagnostics, indent=2), encoding="utf-8"
        )
    except Exception:
        pass

    raise RuntimeError(
        "PRINT_DOCUMENT_NOT_FOUND: Practice Fusion showed the print toolbar, but no "
        "visible document containing PATIENT/FACILITY/ENCOUNTER, SOAP headings, and "
        f"the appointment date was found. Diagnostics were saved in {debug_dir.resolve()}."
    )


def _install_print_document_isolation(page: Page, target_frame) -> None:
    """Make Chromium print only the marked SOAP document.

    The printable document can live in the main document or a same-origin iframe.  In
    either case we hide the underlying EHR with print-only CSS and expand the marked
    document so scroll containers do not clip later SOAP pages.
    """
    inner_style_js = r"""
        () => {
            document.getElementById('pf-sync-print-document-style')?.remove();
            const style = document.createElement('style');
            style.id = 'pf-sync-print-document-style';
            style.textContent = `
                @media print {
                    html, body {
                        margin: 0 !important;
                        padding: 0 !important;
                        width: 100% !important;
                        height: auto !important;
                        overflow: visible !important;
                        background: white !important;
                    }
                    body * { visibility: hidden !important; }
                    [data-pf-sync-print-document-root='true'],
                    [data-pf-sync-print-document-root='true'] * {
                        visibility: visible !important;
                    }
                    [data-pf-sync-print-document-root='true'] {
                        position: absolute !important;
                        inset: 0 auto auto 0 !important;
                        width: 100% !important;
                        min-width: 0 !important;
                        max-width: none !important;
                        height: auto !important;
                        max-height: none !important;
                        overflow: visible !important;
                        margin: 0 !important;
                        padding: 0 !important;
                        transform: none !important;
                        z-index: 2147483647 !important;
                        background: white !important;
                    }
                    [data-pf-sync-print-document-root='true'] .print-link,
                    [data-pf-sync-print-document-root='true'] [title='Print'],
                    [data-pf-sync-print-document-root='true'] .glyphicon-print,
                    [data-pf-sync-print-document-root='true'] .glyphicon-remove,
                    [data-pf-sync-print-document-root='true'] .icon-go-away {
                        display: none !important;
                    }
                }
            `;
            document.head.appendChild(style);
        }
    """
    target_frame.evaluate(inner_style_js)

    if target_frame == page.main_frame:
        return

    # If PF places the document in an iframe, expose only that iframe in the parent.
    frame_element = target_frame.frame_element()
    frame_element.evaluate(
        "el => el.setAttribute('data-pf-sync-print-document-frame', 'true')"
    )
    page.evaluate(
        r"""
        () => {
            document.getElementById('pf-sync-print-frame-style')?.remove();
            const style = document.createElement('style');
            style.id = 'pf-sync-print-frame-style';
            style.textContent = `
                @media print {
                    body * { visibility: hidden !important; }
                    iframe[data-pf-sync-print-document-frame='true'] {
                        visibility: visible !important;
                        position: absolute !important;
                        inset: 0 auto auto 0 !important;
                        width: 100% !important;
                        height: 100vh !important;
                        border: 0 !important;
                    }
                }
            `;
            document.head.appendChild(style);
        }
        """
    )


def _cleanup_print_preview_markers(page: Page) -> None:
    for frame in list(page.frames):
        try:
            frame.evaluate(
                """
                () => {
                    document.querySelectorAll('[data-pf-sync-preexisting-print-link]')
                        .forEach(el => el.removeAttribute('data-pf-sync-preexisting-print-link'));
                    document.querySelectorAll('[data-pf-sync-print-preview-root]')
                        .forEach(el => el.removeAttribute('data-pf-sync-print-preview-root'));
                    document.querySelectorAll('[data-pf-sync-print-document-root]')
                        .forEach(el => el.removeAttribute('data-pf-sync-print-document-root'));
                    document.getElementById('pf-sync-print-isolation-style')?.remove();
                    document.getElementById('pf-sync-print-document-style')?.remove();
                }
                """
            )
        except Exception:
            pass
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('[data-pf-sync-print-document-frame]')
                    .forEach(el => el.removeAttribute('data-pf-sync-print-document-frame'));
                document.getElementById('pf-sync-print-frame-style')?.remove();
            }
            """
        )
    except Exception:
        pass


def _expand_and_materialize_print_modal(target_frame) -> Dict[str, Any]:
    """Scroll the PF preview and remove nested scroll clamps before cloning.

    Practice Fusion keeps the printable chart in a viewport-sized modal.  Long output
    lives inside one or more nested scrollers, and some sections can render only after
    those scrollers move.  Chromium therefore sees only the visible modal height unless
    we first traverse the scrollers and tag them for expansion in the isolated clone.
    """
    result = target_frame.evaluate(
        r"""
        async () => {
            const root = document.querySelector('[data-pf-sync-print-document-root="true"]');
            if (!root) throw new Error('Marked print document root is missing');

            const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
            const candidates = [root, ...root.querySelectorAll('*')].filter(el => {
                if (!(el instanceof HTMLElement)) return false;
                const style = getComputedStyle(el);
                const overflow = `${style.overflow} ${style.overflowY}`;
                return el.scrollHeight > el.clientHeight + 8 || /auto|scroll|hidden/.test(overflow);
            });

            let traversed = 0;
            for (const el of candidates) {
                const maxTop = Math.max(0, el.scrollHeight - el.clientHeight);
                if (maxTop <= 0) continue;
                el.setAttribute('data-pf-sync-expand-scroll', 'true');
                const step = Math.max(250, Math.floor(Math.max(1, el.clientHeight) * 0.75));
                let top = 0;
                let guard = 0;
                while (top < maxTop && guard < 80) {
                    top = Math.min(maxTop, top + step);
                    el.scrollTop = top;
                    el.dispatchEvent(new Event('scroll', {bubbles: true}));
                    await sleep(35);
                    guard += 1;
                }
                await sleep(80);
                el.scrollTop = 0;
                el.dispatchEvent(new Event('scroll', {bubbles: true}));
                traversed += 1;
            }

            // Also exercise the document scroller in case the preview is not the direct
            // scrolling element.  Return to the top afterward so the source UI remains sane.
            const docScroller = document.scrollingElement;
            if (docScroller && docScroller.scrollHeight > docScroller.clientHeight + 8) {
                const maxTop = docScroller.scrollHeight - docScroller.clientHeight;
                const step = Math.max(400, Math.floor(docScroller.clientHeight * 0.8));
                for (let top = 0, guard = 0; top < maxTop && guard < 80; guard += 1) {
                    top = Math.min(maxTop, top + step);
                    docScroller.scrollTop = top;
                    window.dispatchEvent(new Event('scroll'));
                    await sleep(35);
                }
                await sleep(80);
                docScroller.scrollTop = 0;
            }

            await sleep(250);

            // Re-tag every currently clamped descendant.  These attributes survive
            // cloneNode and are expanded by the standalone PDF page CSS.
            let tagged = 0;
            for (const el of [root, ...root.querySelectorAll('*')]) {
                if (!(el instanceof HTMLElement)) continue;
                const style = getComputedStyle(el);
                const overflow = `${style.overflow} ${style.overflowY}`;
                if (el.scrollHeight > el.clientHeight + 8 || /auto|scroll|hidden/.test(overflow)) {
                    el.setAttribute('data-pf-sync-expand-scroll', 'true');
                    tagged += 1;
                }
            }

            const sections = root.querySelectorAll('.print-section').length ||
                (root.matches('.print-section') ? 1 : 0);
            return {
                traversed,
                tagged,
                sectionCount: sections,
                clientHeight: Math.round(root.clientHeight || 0),
                scrollHeight: Math.round(root.scrollHeight || 0),
                textLength: (root.innerText || '').trim().length
            };
        }
        """
    )
    return result if isinstance(result, dict) else {}


def _snapshot_print_document_for_clone(target_frame) -> Dict[str, str]:
    """Serialize the confirmed PF print section and the CSS needed to render it.

    Practice Fusion renders the printable chart inside the existing EHR document.  A
    print-only visibility rule can still yield a blank PDF because PF's own ancestor
    layout/print rules continue to participate.  Instead, clone the confirmed
    ``data-pf-sync-print-document-root`` into a new blank page in the same authenticated
    browser context.
    """
    result = target_frame.evaluate(
        r"""
        () => {
            const root = document.querySelector('[data-pf-sync-print-document-root="true"]');
            if (!root) throw new Error('Marked print document root is missing');

            const sourceSectionCount = root.querySelectorAll('.print-section').length ||
                (root.matches('.print-section') ? 1 : 0);
            const sourceScrollHeight = Math.round(root.scrollHeight || 0);
            const sourceClientHeight = Math.round(root.clientHeight || 0);

            const clone = root.cloneNode(true);
            clone.removeAttribute('data-pf-sync-print-document-root');
            clone.setAttribute('data-pf-sync-cloned-print-root', 'true');

            // Explicitly remove modal/viewport clipping from every source element that
            // was identified as a scroll clamp.  Inline !important is used because PF's
            // Ember styles can otherwise win against the standalone-page stylesheet.
            const expandable = [clone, ...clone.querySelectorAll('[data-pf-sync-expand-scroll="true"]')];
            for (const el of expandable) {
                if (!(el instanceof HTMLElement)) continue;
                el.style.setProperty('height', 'auto', 'important');
                el.style.setProperty('min-height', '0', 'important');
                el.style.setProperty('max-height', 'none', 'important');
                el.style.setProperty('overflow', 'visible', 'important');
                el.style.setProperty('overflow-x', 'visible', 'important');
                el.style.setProperty('overflow-y', 'visible', 'important');
                if (['fixed', 'sticky'].includes(getComputedStyle(el).position)) {
                    el.style.setProperty('position', 'relative', 'important');
                    el.style.setProperty('inset', 'auto', 'important');
                }
            }

            const absolute = (value) => {
                try { return new URL(value, document.baseURI).href; }
                catch (_) { return value; }
            };

            for (const el of clone.querySelectorAll('*')) {
                for (const attr of ['src', 'href', 'poster']) {
                    if (el.hasAttribute(attr)) {
                        const value = el.getAttribute(attr);
                        if (value && !value.startsWith('#') && !value.startsWith('data:') &&
                            !value.startsWith('javascript:')) {
                            el.setAttribute(attr, absolute(value));
                        }
                    }
                }
                if (el.hasAttribute('srcset')) {
                    const converted = el.getAttribute('srcset').split(',').map(part => {
                        const bits = part.trim().split(/\s+/);
                        if (bits[0]) bits[0] = absolute(bits[0]);
                        return bits.join(' ');
                    }).join(', ');
                    el.setAttribute('srcset', converted);
                }
            }

            // Preserve current form values in case PF uses inputs in the print view.
            const sourceInputs = root.querySelectorAll('input, textarea, select');
            const clonedInputs = clone.querySelectorAll('input, textarea, select');
            sourceInputs.forEach((source, index) => {
                const dest = clonedInputs[index];
                if (!dest) return;
                if (source instanceof HTMLInputElement) {
                    if (source.type === 'checkbox' || source.type === 'radio') {
                        source.checked ? dest.setAttribute('checked', '') : dest.removeAttribute('checked');
                    } else {
                        dest.setAttribute('value', source.value || '');
                    }
                } else if (source instanceof HTMLTextAreaElement) {
                    dest.textContent = source.value || '';
                } else if (source instanceof HTMLSelectElement) {
                    Array.from(dest.options || []).forEach((option, optionIndex) => {
                        source.options[optionIndex]?.selected
                            ? option.setAttribute('selected', '')
                            : option.removeAttribute('selected');
                    });
                }
            });

            // Canvas pixels are not retained by cloneNode. Convert them to images.
            const sourceCanvases = root.querySelectorAll('canvas');
            const clonedCanvases = clone.querySelectorAll('canvas');
            sourceCanvases.forEach((source, index) => {
                const dest = clonedCanvases[index];
                if (!dest) return;
                try {
                    const img = document.createElement('img');
                    img.src = source.toDataURL('image/png');
                    img.width = source.width;
                    img.height = source.height;
                    img.style.cssText = dest.style.cssText;
                    dest.replaceWith(img);
                } catch (_) {}
            });

            const cssNodes = Array.from(document.querySelectorAll('link[rel="stylesheet"], style'));
            const css = cssNodes.map(node => {
                if (node.tagName === 'LINK') {
                    const href = node.getAttribute('href');
                    return href ? `<link rel="stylesheet" href="${absolute(href)}">` : '';
                }
                return `<style>${node.textContent || ''}</style>`;
            }).join('\n');

            return {
                baseUrl: document.baseURI || location.href,
                title: document.title || 'Practice Fusion chart',
                css,
                rootHtml: clone.outerHTML,
                sourceSectionCount,
                sourceScrollHeight,
                sourceClientHeight,
                sourceTextLength: (root.innerText || '').trim().length
            };
        }
        """
    )
    if not isinstance(result, dict) or not result.get("rootHtml"):
        raise RuntimeError("Could not serialize the Practice Fusion print document.")
    return {str(k): str(v or "") for k, v in result.items()}


def _isolated_print_html(snapshot: Dict[str, str]) -> str:
    """Build a standalone document whose body contains only the PF print section."""
    base_url = json.dumps(snapshot.get("baseUrl", ""))[1:-1]
    title = json.dumps(snapshot.get("title", "Practice Fusion chart"))[1:-1]
    return f"""<!doctype html>
<html>
<head>
<meta charset=\"utf-8\">
<base href=\"{base_url}\">
<title>{title}</title>
{snapshot.get('css', '')}
<style>
    html, body {{
        margin: 0 !important;
        padding: 0 !important;
        width: 100% !important;
        min-height: 100% !important;
        height: auto !important;
        overflow: visible !important;
        background: white !important;
    }}
    body {{ display: block !important; }}
    [data-pf-sync-cloned-print-root='true'] {{
        display: block !important;
        visibility: visible !important;
        position: static !important;
        inset: auto !important;
        float: none !important;
        width: 100% !important;
        min-width: 0 !important;
        max-width: none !important;
        height: auto !important;
        min-height: 0 !important;
        max-height: none !important;
        overflow: visible !important;
        margin: 0 !important;
        padding: 0 !important;
        transform: none !important;
        opacity: 1 !important;
        background: white !important;
    }}
    [data-pf-sync-cloned-print-root='true'] * {{
        visibility: visible !important;
        max-height: none !important;
    }}
    [data-pf-sync-expand-scroll='true'] {{
        height: auto !important;
        min-height: 0 !important;
        max-height: none !important;
        overflow: visible !important;
        overflow-x: visible !important;
        overflow-y: visible !important;
        position: relative !important;
        inset: auto !important;
    }}
    .print-section {{
        display: block !important;
        visibility: visible !important;
        height: auto !important;
        min-height: 0 !important;
        max-height: none !important;
        overflow: visible !important;
        break-inside: auto !important;
        page-break-inside: auto !important;
    }}
    .print-link, [title='Print'], .glyphicon-print,
    .glyphicon-remove, .icon-go-away {{ display: none !important; }}
    @media print {{
        html, body, [data-pf-sync-cloned-print-root='true'] {{
            display: block !important;
            visibility: visible !important;
            overflow: visible !important;
            height: auto !important;
            max-height: none !important;
            background: white !important;
        }}
    }}
</style>
</head>
<body>{snapshot.get('rootHtml', '')}</body>
</html>"""


def _save_pdf_debug_artifacts(
    debug_html: str,
    isolated_page,
    pdf_bytes: bytes,
    prefix: str,
) -> Path:
    debug_dir = Path("pf_sync_debug")
    debug_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = debug_dir / f"{prefix}_{stamp}"
    try:
        base.with_suffix(".html").write_text(debug_html, encoding="utf-8")
    except Exception:
        pass
    try:
        isolated_page.screenshot(path=str(base.with_suffix(".png")), full_page=True)
    except Exception:
        pass
    try:
        if pdf_bytes:
            base.with_suffix(".pdf").write_bytes(pdf_bytes)
    except Exception:
        pass
    return base


def generate_pdf(
    page: Page,
    config: SyncConfig,
    record: QueueRecord,
    downloads_dir: str,
    dry_run: bool,
) -> str:
    """Clone the confirmed PF SOAP print section into a clean page and print it.

    Printing the original EHR page either captured the Summary screen or produced a
    929-byte blank PDF because Practice Fusion's surrounding layout and print rules
    remained active.  The new page preserves PF styles but contains no EHR navigation,
    hidden modal ancestors, advertisements, or overlay backdrop.
    """
    if dry_run:
        return "DRY_RUN_NO_PDF"

    output_dir = Path(downloads_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_dir / format_pdf_name(record, config)

    marker = uuid.uuid4().hex
    existing_count = _mark_current_visible_print_links(
        page, config.printable_preview_ready_selector, marker
    )
    if existing_count:
        print(
            f"  preview guard: ignored {existing_count} pre-existing Summary-page print icon(s)",
            flush=True,
        )

    button = page.locator(config.generate_pdf_button_selector).first
    button.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    button.click()

    isolated_page = None
    cdp = None
    pdf_bytes = b""
    isolated_html = ""
    try:
        _new_visible_print_preview_link(
            page, config.printable_preview_ready_selector, marker, DEFAULT_TIMEOUT
        )

        modal_deadline = time.time() + 15
        while time.time() < modal_deadline:
            if first_visible_locator(page, config.print_modal_ready_selectors, 200) is None:
                break
            time.sleep(0.2)
        if first_visible_locator(page, config.print_modal_ready_selectors, 200) is not None:
            raise RuntimeError(
                "PRINT_PREVIEW_NOT_READY: the Print Chart options modal is still visible "
                "after clicking Print; PDF was not saved."
            )

        page.wait_for_timeout(800)
        try:
            page.wait_for_load_state("networkidle", timeout=15_000)
        except Exception:
            pass
        try:
            page.evaluate(
                "() => document.fonts && document.fonts.ready ? document.fonts.ready : Promise.resolve()"
            )
        except Exception:
            pass

        target_frame, print_root, root_info = _find_and_mark_print_document(
            page, record, config, DEFAULT_TIMEOUT
        )
        root_text = clean(print_root.inner_text(timeout=SHORT_TIMEOUT))
        print(
            "  confirmed printable SOAP document "
            f"({root_info.get('tag')}.{root_info.get('className')}, "
            f"{root_info.get('width')}x{root_info.get('height')}, "
            f"text={len(root_text)}, frame={root_info.get('url')})",
            flush=True,
        )

        expansion = _expand_and_materialize_print_modal(target_frame)
        print(
            "  expanded printable modal "
            f"(sections={expansion.get('sectionCount')}, "
            f"scroll={expansion.get('clientHeight')}->{expansion.get('scrollHeight')}, "
            f"scrollers={expansion.get('tagged')}, traversed={expansion.get('traversed')}, "
            f"text={expansion.get('textLength')})",
            flush=True,
        )

        snapshot = _snapshot_print_document_for_clone(target_frame)
        isolated_html = _isolated_print_html(snapshot)
        isolated_page = page.context.new_page()
        isolated_page.set_default_timeout(DEFAULT_TIMEOUT)
        isolated_page.set_viewport_size({"width": 1200, "height": 1600})
        isolated_page.set_content(isolated_html, wait_until="domcontentloaded")
        try:
            isolated_page.wait_for_load_state("networkidle", timeout=20_000)
        except Exception:
            pass
        try:
            isolated_page.evaluate(
                "() => document.fonts && document.fonts.ready ? document.fonts.ready : Promise.resolve()"
            )
        except Exception:
            pass
        try:
            isolated_page.wait_for_function(
                "() => Array.from(document.images || []).every(img => img.complete)",
                timeout=15_000,
            )
        except Exception:
            pass

        cloned_info = isolated_page.locator(
            "[data-pf-sync-cloned-print-root='true']"
        ).first.evaluate(
            """el => ({
                width: Math.round(el.getBoundingClientRect().width),
                height: Math.round(el.getBoundingClientRect().height),
                scrollHeight: Math.round(el.scrollHeight || 0),
                textLength: (el.innerText || '').trim().length,
                sectionCount: el.querySelectorAll('.print-section').length ||
                    (el.matches('.print-section') ? 1 : 0),
                display: getComputedStyle(el).display,
                visibility: getComputedStyle(el).visibility
            })"""
        )
        print(
            "  cloned SOAP document ready "
            f"({cloned_info.get('width')}x{cloned_info.get('height')}, "
            f"scrollHeight={cloned_info.get('scrollHeight')}, "
            f"sections={cloned_info.get('sectionCount')}, "
            f"text={cloned_info.get('textLength')}, "
            f"display={cloned_info.get('display')}, "
            f"visibility={cloned_info.get('visibility')})",
            flush=True,
        )
        source_section_count = int(snapshot.get("sourceSectionCount") or 0)
        cloned_section_count = int(cloned_info.get("sectionCount") or 0)
        if source_section_count and cloned_section_count < source_section_count:
            base = _save_pdf_debug_artifacts(
                isolated_html, isolated_page, b"", "cloned_print_sections_missing"
            )
            raise RuntimeError(
                "CLONED_PRINT_SECTIONS_MISSING: the isolated page retained only "
                f"{cloned_section_count} of {source_section_count} printable sections. "
                f"Diagnostics saved as {base}.*"
            )

        if int(cloned_info.get("textLength") or 0) < 300 or int(cloned_info.get("height") or 0) < 300:
            base = _save_pdf_debug_artifacts(
                isolated_html, isolated_page, b"", "cloned_print_document_invalid"
            )
            raise RuntimeError(
                "CLONED_PRINT_DOCUMENT_INVALID: the isolated page did not retain the "
                f"SOAP document. Diagnostics saved as {base}.*"
            )

        isolated_page.emulate_media(media="print")
        isolated_page.wait_for_timeout(300)
        try:
            pdf_bytes = isolated_page.pdf(
                format="Letter",
                print_background=True,
                prefer_css_page_size=True,
                display_header_footer=False,
                margin={
                    "top": "0.25in",
                    "bottom": "0.25in",
                    "left": "0.25in",
                    "right": "0.25in",
                },
            )
            print("  PDF engine: Playwright page.pdf() on cloned SOAP-only page", flush=True)
        except Exception as page_pdf_error:
            print(
                f"  page.pdf() unavailable ({type(page_pdf_error).__name__}); "
                "using Chrome Page.printToPDF fallback on cloned page",
                flush=True,
            )
            cdp = isolated_page.context.new_cdp_session(isolated_page)
            result = cdp.send(
                "Page.printToPDF",
                {
                    "landscape": False,
                    "displayHeaderFooter": False,
                    "printBackground": True,
                    "preferCSSPageSize": True,
                    "paperWidth": 8.5,
                    "paperHeight": 11,
                    "marginTop": 0.25,
                    "marginBottom": 0.25,
                    "marginLeft": 0.25,
                    "marginRight": 0.25,
                    "transferMode": "ReturnAsBase64",
                },
            )
            pdf_bytes = base64.b64decode(result.get("data", ""))

        if not isinstance(pdf_bytes, (bytes, bytearray)) or not pdf_bytes.startswith(b"%PDF"):
            base = _save_pdf_debug_artifacts(
                isolated_html, isolated_page, bytes(pdf_bytes or b""), "invalid_generated_pdf"
            )
            raise RuntimeError(
                f"The browser did not return a valid PDF. Diagnostics saved as {base}.*"
            )
        if len(pdf_bytes) < int(config.pdf_min_bytes or 1024):
            base = _save_pdf_debug_artifacts(
                isolated_html, isolated_page, bytes(pdf_bytes), "small_generated_pdf"
            )
            raise RuntimeError(
                f"Generated PDF is unexpectedly small: {len(pdf_bytes)} bytes. "
                f"Diagnostics saved as {base}.*"
            )

        destination.write_bytes(bytes(pdf_bytes))
        return str(destination.resolve())
    finally:
        if cdp is not None:
            try:
                cdp.detach()
            except Exception:
                pass
        if isolated_page is not None:
            try:
                isolated_page.close()
            except Exception:
                pass
        _cleanup_print_preview_markers(page)

def is_ignored(record: QueueRecord, config: SyncConfig) -> bool:
    return is_ignored_status(record.appointment_status, config)


def is_seen(record: QueueRecord, config: SyncConfig) -> bool:
    return is_seen_status(record.appointment_status, config)


def validate_patient_ready(record: QueueRecord) -> None:
    """Confirm the record can drive chart navigation.

    v5.4: this used to require patient_id as well as ehr_patient_guid, but 631 of the
    8288 rows in the PF patient export carry no PRN/record number. Those patients
    matched correctly on name + DOB, got a valid GUID, then hard-failed here with a
    message telling the operator to run match-patients -- which could never fix it,
    while resolve-patient refused to accept a blank --patient-id. The GUID is the only
    identifier chart navigation actually needs: patient_summary_url() takes the GUID
    and format_pdf_name() already falls back to it.
    """
    if not record.ehr_patient_guid:
        raise RuntimeError(
            "PATIENT_NOT_RESOLVED: no ehr_patient_guid; run match-patients or "
            "resolve-patient first."
        )


def process_one_record(
    page: Page,
    record: QueueRecord,
    config: SyncConfig,
    downloads_dir: str,
    scrape_run_id: str,
    exact_refresh: bool,
    dry_run: bool,
    all_rows: Sequence[QueueRecord] = (),
) -> None:
    started = time.perf_counter()
    original_status = record.status
    validate_patient_ready(record)
    record.attempt_count += 1
    if exact_refresh:
        record.refresh_count += 1
    record.status = "processing"
    record.status_reason = "refresh_started" if exact_refresh else "nightly_pdf_started"
    record.processing_started_at = now_iso()
    record.updated_at = now_iso()
    record.scrape_run_id = scrape_run_id
    record.error_message = ""

    summary_url = patient_summary_url(record.ehr_patient_guid)
    page.goto(summary_url, wait_until="domcontentloaded")
    try:
        page.locator(PATIENT_NAME_SELECTOR).first.wait_for(state="visible", timeout=10_000)
    except Exception:
        pass

    detected = find_encounter_for_appointment(
        page, config, record.ehr_patient_guid, record.appointment_date
    )
    record.encounter_key = detected.encounter_key
    record.encounter_date = detected.encounter_date
    record.encounter_type = detected.encounter_type
    record.encounter_code = detected.encounter_code
    record.encounter_chief_complaint = detected.chief_complaint
    record.encounter_source = detected.source

    if "/summary" not in (page.url or ""):
        page.goto(summary_url, wait_until="domcontentloaded")

    modal = open_print_chart(page, config)
    record.selected_sections = select_facesheet_sections(page, config, modal)
    notes_mode, record.selected_soap_note_text = select_notes_for_record(
        page, config, record, all_rows
    )
    record.notes_selection_mode = notes_mode
    record.pdf_path = generate_pdf(page, config, record, downloads_dir, dry_run)

    record.last_checked_at = now_iso()
    record.updated_at = now_iso()
    record.elapsed_seconds = round(time.perf_counter() - started, 3)
    if dry_run:
        record.status = original_status if original_status in {"ready", "review", "failed"} else "ready"
        record.status_reason = "dry_run_validated"
        record.message = "Encounter and SOAP note were found; PDF generation was skipped."
    else:
        record.status = "processed"
        record.status_reason = "refresh_processed" if exact_refresh else "nightly_pdf_processed"
        record.processed_at = now_iso()
        record.message = ""


def handle_process_error(record: QueueRecord, config: SyncConfig, exc: Exception) -> str:
    record.last_checked_at = now_iso()
    record.updated_at = now_iso()
    record.elapsed_seconds = 0.0
    record.error_message = f"{type(exc).__name__}: {exc}"
    if isinstance(exc, EncounterNotFoundError):
        if is_seen(record, config):
            record.status = "review"
            record.status_reason = "seen_appointment_missing_encounter"
            record.review_count += 1
            record.message = "Seen appointment has no matching encounter yet; it will be polled again."
        else:
            record.status = "ready"
            record.status_reason = "waiting_for_encounter"
            record.message = "Appointment has no matching encounter yet."
    elif isinstance(exc, SoapNoteNotFoundError):
        record.status = "review"
        record.status_reason = "soap_note_not_available_for_appointment_date"
        record.review_count += 1
        record.message = "Encounter exists but the SOAP note is not available in Print Chart yet."
    elif "PATIENT_NOT_RESOLVED" in str(exc):
        record.status = "needs_attention"
        record.status_reason = "patient_not_resolved"
        record.message = str(exc)
    else:
        record.status = "failed"
        record.status_reason = "pdf_worker_error"
        record.message = str(exc)
    return record.status


def process_records_on_page(
    page: Page,
    queue_json: str,
    config: SyncConfig,
    downloads_dir: str,
    candidates: Sequence[QueueRecord],
    all_rows: List[QueueRecord],
    store: Dict[str, Any],
    limit: int = 0,
    dry_run: bool = False,
    exact_refresh: bool = False,
) -> Dict[str, int]:
    if limit > 0:
        candidates = list(candidates)[:limit]
    counts = {"processed": 0, "validated": 0, "review": 0, "ready": 0, "failed": 0, "ignored": 0, "needs_attention": 0}
    scrape_run_id = str(uuid.uuid4())

    for index, record in enumerate(candidates, start=1):
        label = record.appointment_id or record.encounter_key or record.row_id
        print(f"[{index}/{len(candidates)}] {record.patient_name} | {record.appointment_date} | {label}")
        if is_ignored(record, config):
            record.status = "ignored"
            record.status_reason = f"ignored_appointment_status:{normalize_status(record.appointment_status)}"
            record.updated_at = now_iso()
            counts["ignored"] += 1
            save_store(queue_json, store, all_rows)
            print("  ignored")
            continue
        if record.patient_match_status != "matched" or not record.ehr_patient_guid:
            record.status = "needs_attention"
            record.status_reason = "patient_not_resolved"
            record.message = "Patient ID/GUID must be resolved before encounter processing."
            record.updated_at = now_iso()
            counts["needs_attention"] += 1
            save_store(queue_json, store, all_rows)
            print("  needs_attention: patient not resolved")
            continue
        try:
            process_one_record(
                page,
                record,
                config,
                downloads_dir,
                scrape_run_id,
                exact_refresh,
                dry_run,
                all_rows,
            )
            if dry_run:
                counts["validated"] += 1
                print(f"  validated in {record.elapsed_seconds:.3f}s")
            else:
                counts["processed"] += 1
                print(f"  processed in {record.elapsed_seconds:.3f}s -> {record.pdf_path}")
        except Exception as exc:
            state = handle_process_error(record, config, exc)
            counts[state] = counts.get(state, 0) + 1
            print(f"  {state}: {record.error_message}")
        finally:
            # v5.4: always tear the Print Chart modal down so a record that failed inside
            # the modal cannot leave it open over the next patient's chart.
            close_print_chart(page, config)
            save_store(queue_json, store, all_rows)
    return counts


def default_process_candidates(rows: Sequence[QueueRecord], include_failed: bool = False) -> List[QueueRecord]:
    statuses = {"ready", "review"}
    if include_failed:
        statuses.add("failed")
    return [record for record in rows if record.status in statuses]


def full_sync_on_page(
    page: Page,
    queue_json: str,
    config: SyncConfig,
    downloads_dir: str,
    patients_file: str,
    store: Dict[str, Any],
    rows: List[QueueRecord],
    limit_patients: int = 0,
    max_encounters_per_patient: int = 0,
    dry_run: bool = False,
    rescrape_all: bool = False,
) -> Dict[str, Any]:
    """Discover historical SOAP encounters and process only unprocessed dates.

    One queue row is created per patient/date. Practice Fusion's Notes menu can
    contain more than one SOAP note on the same date; selecting the date captures
    all matching notes in the resulting PDF without creating duplicate PDFs.
    """
    registry = load_patient_registry(patients_file)
    if limit_patients > 0:
        registry = registry[:limit_patients]

    by_row_id = {record.row_id: record for record in rows}
    processed_keys = {
        (record.ehr_patient_guid, record.encounter_date or (
            parse_date(record.appointment_date).isoformat()
            if parse_date(record.appointment_date) else ""
        ))
        for record in rows
        if record.status == "processed"
    }
    counts: Dict[str, int] = {
        "patients_scanned": 0,
        "patients_failed": 0,
        "encounter_dates_discovered": 0,
        "already_processed": 0,
        "processed": 0,
        "validated": 0,
        "review": 0,
        "ready": 0,
        "failed": 0,
    }
    patient_timings: List[Dict[str, Any]] = []

    for patient_index, patient in enumerate(registry, start=1):
        patient_started = time.perf_counter()
        patient_guid = clean(patient.get("ehr_patient_guid"))
        patient_id = clean(patient.get("patient_id"))
        patient_name = clean(patient.get("patient_name"))
        print(
            f"[patient {patient_index}/{len(registry)}] {patient_name} "
            f"({patient_id or patient_guid})"
        )
        if not patient_guid:
            counts["patients_failed"] += 1
            patient_timings.append(
                {
                    "patient_id": patient_id,
                    "patient_name": patient_name,
                    "seconds": round(time.perf_counter() - patient_started, 3),
                    "status": "missing_guid",
                }
            )
            continue
        try:
            summary_url = patient_summary_url(patient_guid)
            page.goto(summary_url, wait_until="domcontentloaded")
            try:
                page.locator(PATIENT_NAME_SELECTOR).first.wait_for(
                    state="visible", timeout=10_000
                )
            except Exception:
                pass
            discovered = all_patient_encounters(
                page, config, patient_guid, include_timeline=True
            )
            # One PDF per encounter date; the note picker selects all notes that
            # share that date.
            by_date: Dict[str, DetectedEncounter] = {}
            for encounter in discovered:
                by_date.setdefault(encounter.encounter_date, encounter)
            encounter_dates = list(by_date.values())
            encounter_dates.sort(
                key=lambda item: parse_date(item.encounter_date) or date.min,
                reverse=True,
            )
            if max_encounters_per_patient > 0:
                encounter_dates = encounter_dates[:max_encounters_per_patient]
            counts["patients_scanned"] += 1
            counts["encounter_dates_discovered"] += len(encounter_dates)

            for detected in encounter_dates:
                key = (patient_guid, detected.encounter_date)
                if key in processed_keys and not rescrape_all:
                    counts["already_processed"] += 1
                    continue
                row_id = (
                    f"full-sync|{patient_guid}|{detected.encounter_date}"
                )
                record = by_row_id.get(row_id)
                if record is None:
                    record = QueueRecord(
                        row_id=row_id,
                        practice="",
                        patient_id=patient_id,
                        ehr_patient_guid=patient_guid,
                        patient_name=patient_name,
                        patient_dob=clean(patient.get("dob")),
                        patient_phone=(patient.get("phones") or [""])[0],
                        patient_phone_normalized=(patient.get("phones") or [""])[0],
                        patient_match_status="matched",
                        patient_match_method="patient_registry_full_sync",
                        patient_match_score=1.0,
                        appointment_date=detected.encounter_date,
                        appointment_status="seen",
                        status="ready",
                        status_reason="full_sync_discovered",
                        encounter_key=detected.encounter_key,
                        encounter_date=detected.encounter_date,
                        encounter_type=detected.encounter_type,
                        encounter_code=detected.encounter_code,
                        encounter_chief_complaint=detected.chief_complaint,
                        encounter_source=detected.source,
                        created_at=now_iso(),
                        updated_at=now_iso(),
                        first_ready_at=now_iso(),
                    )
                    rows.append(record)
                    by_row_id[row_id] = record
                elif rescrape_all and record.status == "processed":
                    record.status = "ready"
                    record.status_reason = "full_sync_rescrape"
                    record.pdf_path = ""
                    record.processed_at = ""
                    record.updated_at = now_iso()

                # all_patient_encounters leaves the page on the timeline. Each
                # process call returns to the Summary URL and uses the normal flow.
                result = process_records_on_page(
                    page,
                    queue_json,
                    config,
                    downloads_dir,
                    [record],
                    rows,
                    store,
                    limit=1,
                    dry_run=dry_run,
                    exact_refresh=False,
                )
                for state in ("processed", "validated", "review", "ready", "failed"):
                    counts[state] += result.get(state, 0)
                if record.status == "processed":
                    processed_keys.add(key)
        except Exception as exc:
            counts["patients_failed"] += 1
            print(f"  patient ERROR: {type(exc).__name__}: {exc}")
        finally:
            patient_timings.append(
                {
                    "patient_id": patient_id,
                    "patient_name": patient_name,
                    "seconds": round(time.perf_counter() - patient_started, 3),
                }
            )
            save_store(queue_json, store, rows)

    counts["patient_timings"] = patient_timings
    return counts


# ---------------------------------------------------------------------------
# Refresh behavior
# ---------------------------------------------------------------------------


def newest_encounter_after_last_processed(
    page: Page,
    config: SyncConfig,
    patient_guid: str,
    rows: Sequence[QueueRecord],
) -> Optional[DetectedEncounter]:
    last_dates = [
        parse_date(record.encounter_date)
        for record in rows
        if record.ehr_patient_guid == patient_guid and record.status == "processed" and parse_date(record.encounter_date)
    ]
    last_date = max(last_dates) if last_dates else None
    summary_url = patient_summary_url(patient_guid)
    page.goto(summary_url, wait_until="domcontentloaded")
    encounters = all_patient_encounters(page, config, patient_guid, include_timeline=True)
    for encounter in encounters:
        encounter_date = parse_date(encounter.encounter_date)
        if encounter_date and (last_date is None or encounter_date > last_date):
            return encounter
    return None


def resolve_refresh_patient_template(
    store: Dict[str, Any],
    rows: Sequence[QueueRecord],
    patient_id: str = "",
    ehr_patient_guid: str = "",
) -> QueueRecord:
    """Resolve a refresh target by PRN or by the Practice Fusion chart GUID.

    ``patient_id`` remains the optional PRN/record number. ``ehr_patient_guid`` is
    the UUID used in the Practice Fusion chart URL and is sufficient by itself.
    """
    patient_id = clean(patient_id)
    ehr_patient_guid = clean(ehr_patient_guid)
    if not patient_id and not ehr_patient_guid:
        raise ValueError("Refresh requires --patient-id or --ehr-patient-guid.")

    patient_rows = [
        record
        for record in rows
        if record.ehr_patient_guid
        and (
            (ehr_patient_guid and record.ehr_patient_guid == ehr_patient_guid)
            or (patient_id and record.patient_id == patient_id)
        )
    ]
    if patient_rows:
        distinct_guids = {record.ehr_patient_guid for record in patient_rows}
        if len(distinct_guids) != 1:
            raise ValueError(
                f"Refresh selector matched multiple patient GUIDs: {sorted(distinct_guids)}"
            )
        return max(
            patient_rows,
            key=lambda record: parse_date(record.appointment_date) or date.min,
        )

    mappings = [
        mapping
        for mapping in store.get("patient_mappings", [])
        if (
            (ehr_patient_guid and clean(mapping.get("ehr_patient_guid")) == ehr_patient_guid)
            or (patient_id and clean(mapping.get("patient_id")) == patient_id)
        )
    ]
    distinct_mapping_guids = {
        clean(mapping.get("ehr_patient_guid")) for mapping in mappings
        if clean(mapping.get("ehr_patient_guid"))
    }
    if len(distinct_mapping_guids) > 1:
        raise ValueError(
            f"Refresh selector matched multiple saved patient GUIDs: {sorted(distinct_mapping_guids)}"
        )
    mapping = mappings[0] if mappings else {}
    resolved_guid = ehr_patient_guid or clean(mapping.get("ehr_patient_guid"))
    if not resolved_guid:
        raise ValueError(
            f"Could not resolve a unique patient GUID for patient_id={patient_id}"
        )

    resolved_patient_id = patient_id or clean(mapping.get("patient_id"))
    selector_value = resolved_patient_id or resolved_guid
    return QueueRecord(
        row_id=f"refresh|{selector_value}|{uuid.uuid4()}",
        practice="",
        patient_id=resolved_patient_id,
        ehr_patient_guid=resolved_guid,
        patient_name=clean(mapping.get("patient_name")),
        patient_dob=clean(mapping.get("dob")),
        patient_phone=clean(mapping.get("phone")),
        patient_phone_normalized=normalize_phone(clean(mapping.get("phone"))),
        patient_match_status="matched",
        patient_match_method="saved_mapping" if mapping else "direct_guid",
        status="ready",
        created_at=now_iso(),
        updated_at=now_iso(),
    )


def refresh_patient_latest_on_page(
    page: Page,
    queue_json: str,
    config: SyncConfig,
    downloads_dir: str,
    patient_id: str,
    ehr_patient_guid: str,
    dry_run: bool,
    store: Dict[str, Any],
    rows: List[QueueRecord],
) -> Dict[str, int]:
    template = resolve_refresh_patient_template(
        store, rows, patient_id=patient_id, ehr_patient_guid=ehr_patient_guid
    )
    detected = newest_encounter_after_last_processed(
        page, config, template.ehr_patient_guid, rows
    )
    if detected is None:
        print("No new encounter exists after the last processed encounter.")
        return {"no_new_encounter": 1}

    existing = next((record for record in rows if record.encounter_key == detected.encounter_key), None)
    if existing is None:
        identity_value = template.patient_id or template.ehr_patient_guid
        existing = QueueRecord(
            row_id=f"{template.practice}|refresh|{identity_value}|{detected.encounter_key[:20]}",
            practice=template.practice,
            patient_id=template.patient_id,
            ehr_patient_guid=template.ehr_patient_guid,
            patient_name=template.patient_name,
            patient_dob=template.patient_dob,
            patient_phone=template.patient_phone,
            patient_phone_normalized=template.patient_phone_normalized,
            patient_match_status="matched",
            patient_match_method="refresh_patient",
            appointment_date=detected.encounter_date,
            appointment_status="seen",
            status="ready",
            status_reason="patient_refresh_new_encounter",
            encounter_key=detected.encounter_key,
            encounter_date=detected.encounter_date,
            encounter_type=detected.encounter_type,
            encounter_code=detected.encounter_code,
            encounter_chief_complaint=detected.chief_complaint,
            encounter_source=detected.source,
            created_at=now_iso(),
            updated_at=now_iso(),
        )
        rows.append(existing)
    # Return to Summary because all_patient_encounters may leave us on timeline.
    page.goto(patient_summary_url(existing.ehr_patient_guid), wait_until="domcontentloaded")
    return process_records_on_page(
        page,
        queue_json,
        config,
        downloads_dir,
        [existing],
        rows,
        store,
        limit=1,
        dry_run=dry_run,
        exact_refresh=True,
    )


# ---------------------------------------------------------------------------
# Queue status/reset and local self-test
# ---------------------------------------------------------------------------


def queue_status(queue_json: str, show_limit: int = 20) -> Dict[str, Any]:
    store = load_store(queue_json)
    rows = store_rows(store)
    counts: Dict[str, int] = {}
    for record in rows:
        counts[record.status] = counts.get(record.status, 0) + 1
    print("Queue counts:")
    for status, count in sorted(counts.items()):
        print(f"  {status:18s} {count}")
    attention = [record for record in rows if record.status == "needs_attention"]
    review = [record for record in rows if record.status == "review"]
    if attention:
        print("\nNeeds attention:")
        for record in attention[:show_limit]:
            print(
                f"  appointment_id={record.appointment_id or '<none>'} "
                f"row_id={record.row_id} patient={record.patient_name} DOB={record.patient_dob} "
                f"phone={record.patient_phone} message={record.message or record.patient_match_message}"
            )
    if review:
        print("\nReview/poll again:")
        for record in review[:show_limit]:
            print(
                f"  appointment_id={record.appointment_id or '<none>'} patient={record.patient_name} "
                f"date={record.appointment_date} reason={record.status_reason}"
            )
    return {"counts": counts, "needs_attention": len(attention), "review": len(review)}


def reset_rows(
    queue_json: str,
    row_id: str = "",
    appointment_id: str = "",
    patient_id: str = "",
    all_processed: bool = False,
) -> int:
    store = load_store(queue_json)
    rows = store_rows(store)
    selected = select_queue_rows(
        rows, row_id=row_id, appointment_id=appointment_id, patient_id=patient_id
    )
    if all_processed:
        selected = [record for record in rows if record.status == "processed"]
    if not selected:
        raise ValueError("No rows matched the reset selector.")
    for record in selected:
        if record.status != "ignored":
            record.status = "ready"
            record.status_reason = "manually_reset_for_test"
            record.error_message = ""
            record.message = ""
            record.pdf_path = ""
            record.processed_at = ""
            record.updated_at = now_iso()
    save_store(queue_json, store, rows)
    return len(selected)


class _FakeCheckbox:
    """Minimal Locator stand-in for testing styled-checkbox handling without a browser."""

    def __init__(self, box: Dict[str, Any], kind: str = "input") -> None:
        self.box = box
        self.kind = kind

    def count(self) -> int:
        return 1

    def wait_for(self, **kwargs) -> None:
        return None

    def is_checked(self) -> bool:
        return bool(self.box["checked"])

    def is_visible(self) -> bool:
        if self.kind == "input":
            return bool(self.box["input_visible"])
        return bool(self.box.get(self.kind + "_visible", False))

    def check(self, force: bool = False, timeout: Any = None) -> None:
        if not self.box["force_works"]:
            raise RuntimeError("element is not visible")
        self.box["checked"] = True

    def uncheck(self, force: bool = False, timeout: Any = None) -> None:
        if not self.box["force_works"]:
            raise RuntimeError("element is not visible")
        self.box["checked"] = False

    def click(self, position: Any = None, timeout: Any = None) -> None:
        self.box["clicks"].append((self.kind, "glyph" if position else "center"))
        if self.kind != "sibling label":
            raise RuntimeError("not clickable")
        if position is None and self.box.get("center_hits_link"):
            self.box["dropdown_opened"] = True
            return
        self.box["checked"] = not self.box["checked"]

    def locator(self, selector: str) -> "_FakeCheckbox":
        kinds = {
            "xpath=following-sibling::label[1]": "sibling label",
            "xpath=../label": "parent label",
            "xpath=..": "parent container",
        }
        return _FakeCheckbox(self.box, kinds[selector])

    @property
    def first(self) -> "_FakeCheckbox":
        return self

    def evaluate(self, script: str, arg: Any = None) -> None:
        if not self.box["js_works"]:
            raise RuntimeError("evaluate blocked")
        self.box["checked"] = arg


def _fake_checkbox_box(**overrides: Any) -> Dict[str, Any]:
    state: Dict[str, Any] = {
        "checked": False,
        "input_visible": False,
        "force_works": False,
        "js_works": False,
        "sibling label_visible": False,
        "clicks": [],
        "dropdown_opened": False,
    }
    state.update(overrides)
    return state


def run_checkbox_self_test() -> None:
    """Styled-checkbox handling: PF hides the input and draws the control on the label."""
    # force=True on the hidden input succeeds without touching the label.
    box = _fake_checkbox_box(force_works=True)
    set_checkbox_state(_FakeCheckbox(box), True, "demographics")
    assert box["checked"] and not box["clicks"]

    # force fails: fall back to the label, clicking the glyph at its left edge.
    box = _fake_checkbox_box(**{"sibling label_visible": True})
    set_checkbox_state(_FakeCheckbox(box), True, "diagnoses")
    assert box["checked"] and box["clicks"] == [("sibling label", "glyph")]

    # Notes row: its label wraps the dropdown anchor, so a center click would open the
    # dropdown instead of toggling. The glyph position must be tried first.
    box = _fake_checkbox_box(center_hits_link=True, **{"sibling label_visible": True})
    set_checkbox_state(_FakeCheckbox(box), True, "notes group checkbox")
    assert box["checked"] and not box["dropdown_opened"]

    # No usable click target: DOM property plus input/change events.
    box = _fake_checkbox_box(js_works=True)
    set_checkbox_state(_FakeCheckbox(box), True, "goals")
    assert box["checked"]

    # Already in the desired state: never toggle.
    box = _fake_checkbox_box(checked=True, force_works=True)
    set_checkbox_state(_FakeCheckbox(box), True, "demographics")
    assert box["checked"] and not box["clicks"]

    # Unchecking works the same way.
    box = _fake_checkbox_box(checked=True, force_works=True)
    set_checkbox_state(_FakeCheckbox(box), False, "allergies")
    assert box["checked"] is False

    # Nothing works: raise, naming the control and every strategy tried.
    box = _fake_checkbox_box()
    try:
        set_checkbox_state(_FakeCheckbox(box), True, "chk-sia")
        raise AssertionError("set_checkbox_state should have raised.")
    except RuntimeError as exc:
        assert "CHECKBOX_NOT_SETTABLE" in str(exc)
        assert "chk-sia" in str(exc)

    # Interactivity must not depend on the input itself rendering.
    assert checkbox_is_interactive(_FakeCheckbox(_fake_checkbox_box(**{"sibling label_visible": True})))
    assert checkbox_is_interactive(_FakeCheckbox(_fake_checkbox_box(input_visible=True)))
    assert not checkbox_is_interactive(_FakeCheckbox(_fake_checkbox_box()))


def run_self_test() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        appointments = root / "appointments.csv"
        patients = root / "patients.csv"
        queue = root / "queue.json"

        # v5.4: this fixture used to invent headers (Appointment ID, PATIENT, DATE/TIME,
        # APPT. TYPE, SEEN BY PROVIDER) that no Practice Fusion export actually emits, then
        # printed "Practice Fusion DATE/TIME/APPT./provider column mapping: PASS". The test
        # asserted the exact thing that was broken and stayed green while real ingest failed.
        # The header set below is copied verbatim from a real PF CSV export.
        pf_export_headers = [
            "AppointmentTime", "Patient", "DOB", "MobilePhone", "HomePhone",
            "OfficePhone", "AppointmentType", "AppointmentStatus", "SeenBy",
            "Copay", "Eligibility", "Facility",
        ]

        def pf_row(
            time_value: str, name: str, dob: str, mobile: str,
            appointment_type: str, status: str, provider: str,
        ) -> Dict[str, str]:
            values = {
                "AppointmentTime": time_value,
                "Patient": name,
                "DOB": dob,
                "MobilePhone": mobile,
                "HomePhone": "",
                "OfficePhone": "",
                "AppointmentType": appointment_type,
                "AppointmentStatus": status,
                "SeenBy": provider,
                "Copay": "$35",
                "Eligibility": "Not available.",
                "Facility": "Self Test Clinic",
            }
            return {header: values[header] for header in pf_export_headers}

        write_csv(
            str(appointments),
            [
                pf_row(
                    "07/29/2026 09:15 AM", "Jane A Doe", "01/02/1980",
                    "(555) 111-2222", "Office Visit", "Checked out", "Test Provider M.D",
                ),
                pf_row(
                    "07/29/2026 10:00 AM", "John Smith", "04/05/1975",
                    "555-333-4444", "Follow-Up Visit", "Completed", "Test Provider M.D",
                ),
                pf_row(
                    "07/29/2026 10:30 AM", "Alex Lee", "03/03/1990",
                    "", "Lab Results", "Seen", "Test Provider M.D",
                ),
                pf_row(
                    "07/29/2026 11:00 AM", "Cancel Person", "02/02/1982",
                    "", "Wellness Exam", "Cancelled by patient", "Test Provider M.D",
                ),
                # Subset-name case: the report drops a middle name the chart carries, and
                # the DOB matches exactly. This must auto-resolve, not escalate.
                pf_row(
                    "07/29/2026 11:30 AM", "Maria Reyes Gomez", "06/07/1984",
                    "555-777-8888", "New Patient Visit", "Seen", "Test Provider M.D",
                ),
                # No PRN in the registry for this patient; GUID alone must be sufficient.
                pf_row(
                    "07/29/2026 01:00 PM", "Noel Prentiss", "09/09/1991",
                    "555-999-0000", "Sick", "Seen", "Test Provider M.D",
                ),
            ],
        )
        write_csv(
            str(patients),
            [
                {
                    "patient_id": "P1",
                    "ehr_patient_guid": "11111111-1111-1111-1111-111111111111",
                    "patient_name": "Jane Doe",
                    "dob": "01/02/1980",
                    "mobile_phone": "5551112222",
                    "status": "Active",
                    "raw_patient_json": "x" * 300_000,
                },
                {
                    # Same identity as Jane but explicitly inactive; it must be removed
                    # before matching so it cannot create a false ambiguity.
                    "patient_id": "P1-OLD",
                    "ehr_patient_guid": "11111111-1111-1111-1111-111111111199",
                    "patient_name": "Jane Doe",
                    "dob": "01/02/1980",
                    "mobile_phone": "5551112222",
                    "status": "Inactive",
                },
                {
                    "patient_id": "P2",
                    "ehr_patient_guid": "22222222-2222-2222-2222-222222222222",
                    "patient_name": "John Smith",
                    "dob": "04/05/1975",
                    "mobile_phone": "5553334444",
                },
                {
                    "patient_id": "P3",
                    "ehr_patient_guid": "33333333-3333-3333-3333-333333333333",
                    "patient_name": "Alex Lee",
                    "dob": "03/03/1990",
                    "mobile_phone": "5550001111",
                },
                {
                    "patient_id": "P4",
                    "ehr_patient_guid": "44444444-4444-4444-4444-444444444444",
                    "patient_name": "Alex Lee",
                    "dob": "03/03/1990",
                    "mobile_phone": "5550002222",
                },
                {
                    "patient_id": "P5",
                    "ehr_patient_guid": "55555555-5555-5555-5555-555555555555",
                    # Chart carries a middle name the appointment report omits.
                    "patient_name": "Maria Del Carmen Reyes Gomez",
                    "dob": "06/07/1984",
                    "mobile_phone": "5557778888",
                },
                {
                    # No PRN, as with 631 rows in the real PF export.
                    "patient_id": "",
                    "ehr_patient_guid": "66666666-6666-6666-6666-666666666666",
                    "patient_name": "Noel Prentiss",
                    "dob": "09/09/1991",
                    "mobile_phone": "5559990000",
                },
            ],
        )
        # Header normalization must handle the camelCase export form directly.
        assert normalize_header("AppointmentTime") == "appointment time"
        assert normalize_header("AppointmentStatus") == "appointment status"
        assert normalize_header("MobilePhone") == "mobile phone"
        assert normalize_header("SeenBy") == "seen by"
        assert normalize_header("DOB") == "dob"
        # The spaced/punctuated forms must still normalize to the same tokens.
        assert normalize_header("APPT. STATUS") == "appt status"
        assert normalize_header("Appointment Date") == "appointment date"

        config = SyncConfig()
        ingest_counts = ingest_appointments(
            str(appointments), str(queue), "Self Test Practice", config=config
        )
        assert ingest_counts["inserted"] == 6, ingest_counts

        def row_for(name: str) -> QueueRecord:
            rows_now = store_rows(load_store(str(queue)))
            return next(row for row in rows_now if row.patient_name == name)

        # Every field the PF export carries must survive mapping, not just date and name.
        pf_export_row = row_for("Jane A Doe")
        assert pf_export_row.appointment_date == "07/29/2026 09:15 AM"
        assert parse_date(pf_export_row.appointment_date) == date(2026, 7, 29)
        assert pf_export_row.appointment_type == "Office Visit"
        assert pf_export_row.appointment_status == "Checked out"
        assert pf_export_row.provider == "Test Provider M.D"
        assert pf_export_row.patient_phone_normalized == "5551112222"

        # A blank appointment_status must be a hard mapping failure, because an empty
        # status silently disables the ignored gate.
        unmapped = [{"SomeTime": "07/29/2026", "Who": "X", "Born": "01/01/1980"}]
        try:
            validate_appointment_report_mapping(
                unmapped, [map_appointment_row(row) for row in unmapped]
            )
            raise AssertionError("Unmapped report headers should have raised ValueError.")
        except ValueError as exc:
            assert "appointment status" in str(exc), str(exc)

        # The ignored gate must read the config, not a module-level constant.
        assert is_ignored_status("Cancelled by patient", config)
        narrowed = SyncConfig(ignored_statuses=["deleted"])
        assert not is_ignored_status("Cancelled by patient", narrowed)

        match_counts = match_patients(str(queue), str(patients))
        assert row_for("Cancel Person").status == "ignored"

        # Explicitly inactive patient charts are excluded before scoring.
        jane = row_for("Jane A Doe")
        assert jane.patient_match_status == "matched", jane.patient_match_message
        assert jane.patient_id == "P1"
        assert jane.ehr_patient_guid == "11111111-1111-1111-1111-111111111111"

        # Subset name plus exact DOB must auto-resolve rather than escalate.
        subset = row_for("Maria Reyes Gomez")
        assert subset.patient_match_status == "matched", subset.patient_match_message
        assert subset.ehr_patient_guid == "55555555-5555-5555-5555-555555555555"

        # A chart with no PRN must be usable end to end on the GUID alone.
        no_prn = row_for("Noel Prentiss")
        assert no_prn.patient_match_status == "matched", no_prn.patient_match_message
        assert no_prn.ehr_patient_guid == "66666666-6666-6666-6666-666666666666"
        assert not no_prn.patient_id
        validate_patient_ready(no_prn)
        assert format_pdf_name(no_prn, config).startswith("66666666-")

        # Genuine ambiguity (two charts, same name, same DOB) must still reach a human.
        assert row_for("Alex Lee").status == "needs_attention", match_counts

        resolve_patient_manually(
            str(queue),
            patient_id="P3",
            ehr_patient_guid="33333333-3333-3333-3333-333333333333",
            row_id=row_for("Alex Lee").row_id,
        )
        assert row_for("Alex Lee").patient_id == "P3"
        assert row_for("Alex Lee").status == "ready"

        # Manual resolution by GUID alone must be accepted.
        resolve_patient_manually(
            str(queue),
            patient_id="",
            ehr_patient_guid="66666666-6666-6666-6666-666666666666",
            row_id=row_for("Noel Prentiss").row_id,
        )
        assert row_for("Noel Prentiss").status == "ready"

        # Refresh can target the Practice Fusion chart GUID directly, independent of PRN.
        refresh_store = load_store(str(queue))
        refresh_rows = store_rows(refresh_store)
        guid_template = resolve_refresh_patient_template(
            refresh_store,
            refresh_rows,
            ehr_patient_guid="33333333-3333-3333-3333-333333333333",
        )
        assert guid_template.patient_id == "P3"
        assert guid_template.ehr_patient_guid == "33333333-3333-3333-3333-333333333333"
        direct_guid = resolve_refresh_patient_template(
            {"patient_mappings": []},
            [],
            ehr_patient_guid="77777777-7777-7777-7777-777777777777",
        )
        assert direct_guid.patient_id == ""
        assert direct_guid.ehr_patient_guid == "77777777-7777-7777-7777-777777777777"
        assert direct_guid.patient_match_method == "direct_guid"

        # Nightly appointment checks are Summary-only. A missing encounter must not
        # invoke Timeline navigation or any Page methods beyond the Summary reader.
        original_summary_reader = globals()["read_summary_encounters"]
        try:
            globals()["read_summary_encounters"] = lambda page, cfg, guid: []
            try:
                find_encounter_for_appointment(
                    object(), config, "33333333-3333-3333-3333-333333333333", "07/31/2026 10:45 AM"
                )
                raise AssertionError("Missing Summary encounter should have raised.")
            except EncounterNotFoundError as exc:
                assert "AFTER_SUMMARY_CHECK" in str(exc), str(exc)
                assert "TIMELINE" not in str(exc), str(exc)
        finally:
            globals()["read_summary_encounters"] = original_summary_reader

        # Note-date tokens must include the leading-zero-stripped month.
        tokens = note_date_tokens("07/27/2026 08:00 AM", config.note_date_formats)
        assert "07/27/2026" in tokens
        assert "7/27/2026" in tokens, tokens

        # Containment scoring: a dropped middle name or second surname must score high,
        # but a single shared given name must not. "Peyton Peyton" is a real malformed
        # registry row; scoring it 1.0 against "Peyton Hicks" attached the wrong chart.
        assert name_token_containment("Marlene Revilla Gomez", "Marlene Del Carmen Revilla Gomez") == 1.0
        assert name_token_containment("Elizabeth Vazquez Martinez", "Elizabeth Vazquez") == 1.0
        assert name_token_containment("Peyton Hicks", "Peyton Peyton") == 0.0
        assert name_token_containment("Ruben Tejada", "Ruben Alvarez") == 0.0
        assert identity_score("Peyton Hicks", "Peyton Peyton") < 0.85

        # v5.5: notes mode. First chart for a patient takes every note; once a PDF exists
        # for that patient, later appointments take only the appointment date's note.
        notes_cfg = SyncConfig()
        assert notes_cfg.notes_selection_mode == "auto"
        first = QueueRecord(
            row_id="r1", ehr_patient_guid="guid-a", appointment_date="07/29/2026 09:15 AM"
        )
        later = QueueRecord(
            row_id="r2", ehr_patient_guid="guid-a", appointment_date="08/05/2026 09:15 AM"
        )
        assert resolve_notes_mode(first, notes_cfg, [first, later]) == "all"
        # A retry of the same row must not count as its own prior history.
        first.status = "processed"
        first.pdf_path = "C:\\charts\\guid-a.pdf"
        assert resolve_notes_mode(first, notes_cfg, [first, later]) == "all"
        assert resolve_notes_mode(later, notes_cfg, [first, later]) == "date"
        # A different patient is unaffected.
        other = QueueRecord(row_id="r3", ehr_patient_guid="guid-b", appointment_date="08/05/2026")
        assert resolve_notes_mode(other, notes_cfg, [first, later, other]) == "all"
        # Explicit overrides win over auto.
        assert resolve_notes_mode(later, SyncConfig(notes_selection_mode="all"), [first, later]) == "all"
        assert resolve_notes_mode(first, SyncConfig(notes_selection_mode="date"), [first]) == "date"
        # A processed row with no pdf_path is not prior history.
        stub = QueueRecord(row_id="r4", ehr_patient_guid="guid-c", status="processed")
        target = QueueRecord(row_id="r5", ehr_patient_guid="guid-c")
        assert resolve_notes_mode(target, notes_cfg, [stub, target]) == "all"

        # The modal-ready selector list must prefer the real dialog over the hidden
        # carbon-content-modal-component wrapper that shares its data-element.
        assert notes_cfg.print_modal_ready_selectors[0].endswith(".content-modal")
        assert any(
            "not(.carbon-content-modal-component)" in selector
            for selector in notes_cfg.print_modal_ready_selectors
        )
        # Every configured section must appear in the known-option manifest, or the
        # exact-selection pass would clear it immediately after checking it.
        for selector in notes_cfg.facesheet_checkbox_selectors:
            assert selector in notes_cfg.facesheet_known_option_selectors, selector

        # The old singular print_modal_ready_selector key must still load.
        legacy = root / "legacy_config.json"
        atomic_write_json(
            str(legacy),
            {"print_modal_ready_selector": "[data-element='legacy-modal']"},
        )
        legacy_cfg = SyncConfig.load(str(legacy))
        assert legacy_cfg.print_modal_ready_selectors == ["[data-element='legacy-modal']"]

        run_checkbox_self_test()

        print("SELF-TEST PASSED")
        print("  Practice Fusion camelCase export header mapping: PASS")
        print("  all required columns mapped (date/status/type/provider): PASS")
        print("  unmapped-column failure raises before queueing: PASS")
        print("  ignored-status gate reads config in ingest and process: PASS")
        print("  ingest/upsert: PASS")
        print("  DOB-first matching incl. dropped middle name/second surname: PASS")
        print("  inactive patient registry rows excluded before matching: PASS")
        print("  GUID-only patients (no PRN) resolve and name PDFs: PASS")
        print("  refresh accepts Practice Fusion chart GUID directly: PASS")
        print("  phone/ambiguity handling: PASS")
        print("  ignored cancellation handling: PASS")
        print("  manual resolution by PRN or GUID + persistent mapping: PASS")
        print("  nightly appointment encounter lookup stays on Summary: PASS")
        print("  SOAP note date tokens incl. no-leading-zero month: PASS")
        print("  print modal selector prefers visible dialog over hidden wrapper: PASS")
        print("  exact facesheet selection manifest is self-consistent: PASS")
        print("  notes mode: all notes on first chart, date-scoped thereafter: PASS")
        print("  legacy singular print_modal_ready_selector still loads: PASS")
        print("  hidden styled checkboxes driven via label/glyph/property ladder: PASS")
        print("  large CSV/JSON field handling: PASS")
    return 0


# ---------------------------------------------------------------------------
# CLI orchestration
# ---------------------------------------------------------------------------


def add_browser_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--attach",
        action="store_true",
        help="Attach to Chrome already started with --remote-debugging-port.",
    )
    parser.add_argument(
        "--chrome-user-data-dir",
        default=os.path.join(os.getenv("USERPROFILE", ""), "pf_rpa_chrome")
        if os.getenv("USERPROFILE") else "",
        help=(
            "Dedicated reusable Practice Fusion Chrome profile directory. "
            "Defaults to %%USERPROFILE%%\\pf_rpa_chrome on Windows."
        ),
    )
    parser.add_argument(
        "--source-user-data-dir",
        default="",
        help="Real Chrome User Data root used only for the initial clone.",
    )
    parser.add_argument(
        "--source-profile",
        default="Profile 11",
        help="Chrome source profile folder, normally Profile 11.",
    )
    parser.add_argument(
        "--refresh-profile",
        action="store_true",
        help="Replace the dedicated profile with a fresh source profile copy.",
    )
    parser.add_argument("--chrome-exe", default="")
    parser.add_argument("--debug-port", default="9222")
    parser.add_argument(
        "--username",
        default=os.getenv("PF_USERNAME", ""),
        help="Practice Fusion username. Prefer the PF_USERNAME environment variable.",
    )
    parser.add_argument(
        "--password",
        default=os.getenv("PF_PASSWORD", ""),
        help="Practice Fusion password. Prefer PF_PASSWORD so it is not stored in shell history.",
    )
    parser.add_argument(
        "--typing-delay-ms",
        type=int,
        default=65,
        help="Delay between login keystrokes for reliable Ember input events (default 65 ms).",
    )
    parser.add_argument(
        "--login-timeout-seconds",
        type=int,
        default=900,
        help="Maximum time to wait for login/security verification (default 900).",
    )
    parser.add_argument("--keep-browser-open", action="store_true")


def add_report_dates(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--report-date", default="", help="Single report day; overrides start/end.")
    parser.add_argument("--start-date", default="")
    parser.add_argument("--end-date", default="")


def resolve_report_dates(args: argparse.Namespace) -> Tuple[date, date]:
    if getattr(args, "report_date", ""):
        value = require_date(args.report_date, "report date")
        return value, value
    today = practice_today()
    start = require_date(args.start_date, "start date") if getattr(args, "start_date", "") else today
    end = require_date(args.end_date, "end date") if getattr(args, "end_date", "") else today
    return start, end


def require_browser_args(args: argparse.Namespace) -> None:
    if not args.attach and not args.chrome_user_data_dir:
        raise ValueError("--chrome-user-data-dir is required unless --attach is used.")


def browser_command_wrapper(args: argparse.Namespace, callback):
    require_browser_args(args)
    playwright = context = page = None
    try:
        playwright, context, page = build_browser(args)
        page = wait_for_pf_login(context, page, args)
        return callback(page)
    finally:
        if playwright is not None and context is not None and page is not None:
            close_browser(args, playwright, context, page)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Practice Fusion appointment, encounter, and SOAP PDF sync"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("version", help="Print the exact worker build ID.")
    sub.add_parser("self-test", help="Run local synthetic tests without Practice Fusion.")

    doctor = sub.add_parser("doctor", help="Validate local files/config/browser prerequisites.")
    doctor.add_argument("--config-json", required=True)
    doctor.add_argument("--report-config-json", required=True)
    doctor.add_argument("--patients-file", default="")
    doctor.add_argument("--queue-json", default="")
    doctor.add_argument(
        "--appointments-file",
        default="",
        help="Validate appointment column mapping before ingest.",
    )
    add_browser_arguments(doctor)

    pull = sub.add_parser("pull-report", help="Pull Appointment & eligibility report.")
    pull.add_argument("--report-config-json", required=True)
    pull.add_argument("--output-csv", required=True)
    add_report_dates(pull)
    add_browser_arguments(pull)

    ingest = sub.add_parser("ingest", help="Upsert appointment CSV/XLSX/JSON into queue JSON.")
    ingest.add_argument("--appointments-file", "--appointments-csv", dest="appointments_file", required=True)
    ingest.add_argument("--queue-json", required=True)
    ingest.add_argument("--practice", required=True)
    ingest.add_argument("--source-report-name", default="")
    ingest.add_argument("--reset-existing", action="store_true")
    # v5.4: ingest now resolves the ignored-status gate through the same config that
    # process uses, so a customized ignored_statuses list applies to both.
    ingest.add_argument(
        "--config-json",
        default="",
        help="PDF sync config; supplies ignored_statuses so ingest and process share one gate.",
    )

    match = sub.add_parser("match-patients", help="Match appointment patients to patient registry.")
    match.add_argument("--queue-json", required=True)
    match.add_argument("--patients-file", "--patients-csv", dest="patients_file", required=True)
    match.add_argument("--fuzzy-threshold", type=float, default=0.82)
    match.add_argument(
        "--dob-match-threshold",
        type=float,
        default=0.85,
        help="Name threshold applied inside an exact-DOB bucket (default 0.85).",
    )
    match.add_argument("--rematch-all", action="store_true")

    resolve = sub.add_parser("resolve-patient", help="Manually resolve a needs_attention appointment.")
    resolve.add_argument("--queue-json", required=True)
    target = resolve.add_mutually_exclusive_group(required=True)
    target.add_argument("--row-id", default="")
    target.add_argument("--appointment-id", default="")
    # v5.4: --patient-id is no longer required. Supply either it (with --patients-file)
    # or --ehr-patient-guid; the GUID is what chart navigation needs.
    resolve.add_argument("--patient-id", default="")
    resolve.add_argument("--ehr-patient-guid", default="")
    resolve.add_argument("--patients-file", default="")
    resolve.add_argument("--resolved-patient-name", default="")

    process = sub.add_parser("process", help="Process ready/review appointments and create PDFs.")
    process.add_argument("--queue-json", required=True)
    process.add_argument("--config-json", required=True)
    process.add_argument("--downloads-dir", required=True)
    process.add_argument("--limit", type=int, default=0)
    process.add_argument("--dry-run", action="store_true")
    process.add_argument("--include-failed", action="store_true")
    add_browser_arguments(process)

    full_sync = sub.add_parser(
        "full-sync",
        help="Discover historical SOAP-note dates for patient registry rows and create PDFs.",
    )
    full_sync.add_argument("--queue-json", required=True)
    full_sync.add_argument("--config-json", required=True)
    full_sync.add_argument("--patients-file", required=True)
    full_sync.add_argument("--downloads-dir", required=True)
    full_sync.add_argument("--limit-patients", type=int, default=0)
    full_sync.add_argument("--max-encounters-per-patient", type=int, default=0)
    full_sync.add_argument("--dry-run", action="store_true")
    full_sync.add_argument("--rescrape-all", action="store_true")
    add_browser_arguments(full_sync)

    refresh = sub.add_parser("refresh", help="Refresh one appointment/encounter or newest patient encounter.")
    refresh.add_argument("--queue-json", required=True)
    refresh.add_argument("--config-json", required=True)
    refresh.add_argument("--downloads-dir", required=True)
    selector = refresh.add_mutually_exclusive_group(required=True)
    selector.add_argument("--row-id", default="")
    selector.add_argument("--appointment-id", default="")
    selector.add_argument("--encounter-id", default="")
    selector.add_argument(
        "--patient-id",
        default="",
        help="Practice Fusion PRN/record number (for example PE751838).",
    )
    selector.add_argument(
        "--ehr-patient-guid", "--patient-guid",
        dest="ehr_patient_guid",
        default="",
        help="Practice Fusion chart UUID from the patient URL.",
    )
    refresh.add_argument("--dry-run", action="store_true")
    add_browser_arguments(refresh)

    nightly = sub.add_parser("nightly", help="Pull -> ingest -> match -> process in one run.")
    nightly.add_argument("--queue-json", required=True)
    nightly.add_argument("--config-json", required=True)
    nightly.add_argument("--report-config-json", required=True)
    nightly.add_argument("--patients-file", required=True)
    nightly.add_argument("--downloads-dir", required=True)
    nightly.add_argument("--practice", required=True)
    nightly.add_argument("--appointments-file", default="", help="Use an existing report file instead of pulling PF.")
    nightly.add_argument("--report-output-csv", default="")
    nightly.add_argument("--limit", type=int, default=0)
    nightly.add_argument("--dry-run", action="store_true")
    nightly.add_argument("--include-failed", action="store_true")
    nightly.add_argument("--fuzzy-threshold", type=float, default=0.82)
    nightly.add_argument("--dob-match-threshold", type=float, default=0.85)
    add_report_dates(nightly)
    add_browser_arguments(nightly)

    status = sub.add_parser("status", help="Show queue counts and unresolved/review rows.")
    status.add_argument("--queue-json", required=True)
    status.add_argument("--show-limit", type=int, default=20)

    reset = sub.add_parser("reset", help="Reset rows to ready for a repeat test.")
    reset.add_argument("--queue-json", required=True)
    selector = reset.add_mutually_exclusive_group(required=True)
    selector.add_argument("--row-id", default="")
    selector.add_argument("--appointment-id", default="")
    selector.add_argument("--patient-id", default="")
    selector.add_argument("--all-processed", action="store_true")

    write_config = sub.add_parser("write-config", help="Write the complete PDF config.")
    write_config.add_argument("--config-json", required=True)
    write_report_config = sub.add_parser("write-report-config", help="Write the report config.")
    write_report_config.add_argument("--report-config-json", required=True)

    return parser


def run_doctor(args: argparse.Namespace) -> int:
    errors = []
    warnings = []
    print(f"PDF config path: {os.path.abspath(args.config_json)}")
    print(f"Report config path: {os.path.abspath(args.report_config_json)}")
    try:
        sync_config = SyncConfig.load(args.config_json)
        required = {
            "print_chart_button_selector": sync_config.print_chart_button_selector,
            "print_modal_ready_selectors": sync_config.print_modal_ready_selectors,
            "facesheet_checkbox_selectors": sync_config.facesheet_checkbox_selectors,
            "print_modal_select_none_selector": sync_config.print_modal_select_none_selector,
            "notes_group_checkbox_selector": sync_config.notes_group_checkbox_selector,
            "generate_pdf_button_selector": sync_config.generate_pdf_button_selector,
            "printable_preview_ready_selector": sync_config.printable_preview_ready_selector,
        }
        errors.extend([f"Missing sync config: {key}" for key, value in required.items() if not value])
    except Exception as exc:
        errors.append(f"Could not load PDF config: {exc}")
    try:
        report_config = AppointmentReportConfig.load(args.report_config_json)
        required = {
            "reports_menu_selector": report_config.reports_menu_selector,
            "appointment_report_link_selector": report_config.appointment_report_link_selector,
            "report_start_date_selector": report_config.report_start_date_selector,
            "report_end_date_selector": report_config.report_end_date_selector,
            "run_report_button_selectors": report_config.run_report_button_selectors,
        }
        errors.extend([f"Missing report config: {key}" for key, value in required.items() if not value])
    except Exception as exc:
        errors.append(f"Could not load report config: {exc}")
    # v5.4: doctor now maps the supplied appointment report. Previously it validated
    # configs, Chrome and the registry but never touched the report file, so a column
    # mapping failure only surfaced two steps later at ingest.
    if getattr(args, "appointments_file", ""):
        try:
            source_rows = read_tabular_rows(args.appointments_file)
            if not source_rows:
                warnings.append(f"Appointment report has no data rows: {args.appointments_file}")
            else:
                mapped_rows = [map_appointment_row(row) for row in source_rows]
                headers = [clean(key) for key in source_rows[0].keys() if clean(key)]
                print(f"Appointment report: {len(source_rows)} rows, headers {headers}")
                blank = [
                    label
                    for field_name, label in REQUIRED_APPOINTMENT_FIELDS.items()
                    if not any(clean(row.get(field_name, "")) for row in mapped_rows)
                ]
                if blank:
                    errors.append(
                        "Appointment column mapping produced no values for: "
                        + ", ".join(blank)
                        + f". Normalized headers: {sorted({normalize_header(h) for h in headers})}"
                    )
                else:
                    sample = mapped_rows[0]
                    print(
                        "  mapped sample: "
                        f"date={sample['appointment_date']!r}, status={sample['appointment_status']!r}, "
                        f"type={sample['appointment_type']!r}, provider={sample['provider']!r}, "
                        f"phone={sample['patient_phone']!r}"
                    )
                for field_name, label in OPTIONAL_APPOINTMENT_FIELDS.items():
                    if not any(clean(row.get(field_name, "")) for row in mapped_rows):
                        warnings.append(f"No source column mapped to {label}.")
        except Exception as exc:
            errors.append(f"Appointment report check failed: {type(exc).__name__}: {exc}")

    if args.patients_file:
        try:
            registry = load_patient_registry(args.patients_file)
            count = len(registry)
            if count == 0:
                warnings.append("Patient registry loaded but contains zero usable patient IDs/GUIDs.")
            else:
                without_prn = sum(1 for item in registry if not item["patient_id"])
                without_dob = sum(1 for item in registry if not item["dob"])
                print(f"Patient registry: {count} usable rows")
                if without_prn:
                    print(
                        f"  {without_prn} rows have no patient_id/PRN "
                        "(matched by GUID; PDFs will be named using the GUID)"
                    )
                if without_dob:
                    warnings.append(
                        f"{without_dob} registry rows have no parseable DOB and cannot use "
                        "DOB-first matching."
                    )
        except Exception as exc:
            errors.append(f"Patient registry failed: {exc}")
    if args.queue_json and os.path.exists(args.queue_json):
        try:
            print(f"Queue rows: {len(store_rows(load_store(args.queue_json)))}")
        except Exception as exc:
            errors.append(f"Queue JSON failed: {exc}")
    if not args.attach:
        if not args.chrome_user_data_dir:
            warnings.append("No --chrome-user-data-dir supplied; browser commands will require it.")
        try:
            chrome = find_chrome_exe(args.chrome_exe)
            print(f"Chrome: {chrome}")
        except SystemExit as exc:
            errors.append(str(exc))
    print(f"CSV field limit: {CSV_FIELD_LIMIT}")
    print(f"Practice timezone: {PRACTICE_TZ_NAME} (today = {practice_today().isoformat()})")
    print("Playwright import: OK")
    for warning in warnings:
        print(f"WARNING: {warning}")
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print("DOCTOR PASSED")
    return 0


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        if args.command == "version":
            print(BUILD_ID)
            return 0

        if args.command == "self-test":
            return run_self_test()

        if args.command == "doctor":
            return run_doctor(args)

        if args.command == "write-config":
            atomic_write_json(args.config_json, asdict(SyncConfig()))
            print(f"Complete PDF config written to {args.config_json}")
            return 0

        if args.command == "write-report-config":
            atomic_write_json(args.report_config_json, asdict(AppointmentReportConfig()))
            print(f"Appointment report config written to {args.report_config_json}")
            return 0

        if args.command == "ingest":
            counts = ingest_appointments(
                args.appointments_file,
                args.queue_json,
                args.practice,
                args.source_report_name,
                args.reset_existing,
                config=SyncConfig.load(args.config_json),
            )
            print(json.dumps(counts, indent=2))
            return 0

        if args.command == "match-patients":
            counts = match_patients(
                args.queue_json,
                args.patients_file,
                args.fuzzy_threshold,
                args.rematch_all,
                args.dob_match_threshold,
            )
            print(json.dumps(counts, indent=2))
            return 0

        if args.command == "resolve-patient":
            counts = resolve_patient_manually(
                args.queue_json,
                args.patient_id,
                args.ehr_patient_guid,
                args.row_id,
                args.appointment_id,
                args.patients_file,
                args.resolved_patient_name,
            )
            print(json.dumps(counts, indent=2))
            return 0

        if args.command == "status":
            queue_status(args.queue_json, args.show_limit)
            return 0

        if args.command == "reset":
            count = reset_rows(
                args.queue_json,
                args.row_id,
                args.appointment_id,
                args.patient_id,
                args.all_processed,
            )
            print(f"Reset {count} row(s) to ready.")
            return 0

        if args.command == "pull-report":
            start_date, end_date = resolve_report_dates(args)
            result = browser_command_wrapper(
                args,
                lambda page: pull_appointment_report_on_page(
                    page,
                    AppointmentReportConfig.load(args.report_config_json),
                    start_date,
                    end_date,
                    args.output_csv,
                ),
            )
            print(json.dumps(result, indent=2))
            return 0

        if args.command == "process":
            store = load_store(args.queue_json)
            rows = store_rows(store)
            candidates = default_process_candidates(rows, args.include_failed)
            if not candidates:
                print("No ready/review records to process.")
                return 0
            config = SyncConfig.load(args.config_json)
            run_id = append_run(store, "process", {"candidates": len(candidates), "dry_run": args.dry_run})

            def callback(page: Page):
                counts = process_records_on_page(
                    page,
                    args.queue_json,
                    config,
                    args.downloads_dir,
                    candidates,
                    rows,
                    store,
                    args.limit,
                    args.dry_run,
                    False,
                )
                finish_run(store, run_id, "success", counts)
                save_store(args.queue_json, store, rows)
                return counts

            counts = browser_command_wrapper(args, callback)
            print(json.dumps(counts, indent=2))
            return 1 if counts.get("failed", 0) else 0

        if args.command == "full-sync":
            store = load_store(args.queue_json)
            rows = store_rows(store)
            config = SyncConfig.load(args.config_json)
            run_id = append_run(
                store,
                "full-sync",
                {
                    "patients_file": str(Path(args.patients_file).resolve()),
                    "limit_patients": args.limit_patients,
                    "max_encounters_per_patient": args.max_encounters_per_patient,
                    "dry_run": args.dry_run,
                    "rescrape_all": args.rescrape_all,
                },
            )

            def callback(page: Page):
                counts = full_sync_on_page(
                    page,
                    args.queue_json,
                    config,
                    args.downloads_dir,
                    args.patients_file,
                    store,
                    rows,
                    args.limit_patients,
                    args.max_encounters_per_patient,
                    args.dry_run,
                    args.rescrape_all,
                )
                finish_run(store, run_id, "success", counts)
                save_store(args.queue_json, store, rows)
                return counts

            counts = browser_command_wrapper(args, callback)
            print(json.dumps(counts, indent=2))
            return 0

        if args.command == "refresh":
            store = load_store(args.queue_json)
            rows = store_rows(store)
            config = SyncConfig.load(args.config_json)

            def callback(page: Page):
                if args.patient_id or args.ehr_patient_guid:
                    return refresh_patient_latest_on_page(
                        page,
                        args.queue_json,
                        config,
                        args.downloads_dir,
                        args.patient_id,
                        args.ehr_patient_guid,
                        args.dry_run,
                        store,
                        rows,
                    )
                candidates = select_queue_rows(
                    rows,
                    row_id=args.row_id,
                    appointment_id=args.appointment_id,
                    encounter_id=args.encounter_id,
                )
                if not candidates:
                    raise ValueError("No queue row matched the refresh selector.")
                return process_records_on_page(
                    page,
                    args.queue_json,
                    config,
                    args.downloads_dir,
                    [max(candidates, key=lambda record: parse_date(record.appointment_date) or date.min)],
                    rows,
                    store,
                    1,
                    args.dry_run,
                    True,
                )

            counts = browser_command_wrapper(args, callback)
            print(json.dumps(counts, indent=2))
            return 1 if counts.get("failed", 0) else 0

        if args.command == "nightly":
            start_date, end_date = resolve_report_dates(args)
            report_file = args.appointments_file
            if not report_file:
                report_file = args.report_output_csv or (
                    f"appointments_{start_date.isoformat()}_to_{end_date.isoformat()}.csv"
                )

            config = SyncConfig.load(args.config_json)
            report_config = AppointmentReportConfig.load(args.report_config_json)

            def callback(page: Page):
                if not args.appointments_file:
                    pull_result = pull_appointment_report_on_page(
                        page, report_config, start_date, end_date, report_file
                    )
                    print("Report pull:", json.dumps(pull_result, indent=2))
                ingest_counts = ingest_appointments(
                    report_file, args.queue_json, args.practice, config=config
                )
                print("Ingest:", json.dumps(ingest_counts, indent=2))
                match_counts = match_patients(
                    args.queue_json,
                    args.patients_file,
                    args.fuzzy_threshold,
                    dob_match_threshold=args.dob_match_threshold,
                )
                print("Patient matching:", json.dumps(match_counts, indent=2))

                # Reload after ingest/match because those functions persist a new store.
                store = load_store(args.queue_json)
                rows = store_rows(store)
                candidates = default_process_candidates(rows, args.include_failed)
                counts = process_records_on_page(
                    page,
                    args.queue_json,
                    config,
                    args.downloads_dir,
                    candidates,
                    rows,
                    store,
                    args.limit,
                    args.dry_run,
                    False,
                )
                return {
                    "report_file": report_file,
                    "ingest": ingest_counts,
                    "match": match_counts,
                    "process": counts,
                }

            result = browser_command_wrapper(args, callback)
            print(json.dumps(result, indent=2))
            return 0

        parser.error(f"Unknown command: {args.command}")
        return 2
    except KeyboardInterrupt:
        print("Cancelled by user.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
