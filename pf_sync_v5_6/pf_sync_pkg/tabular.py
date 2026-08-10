"""Tabular (CSV/XLSX/JSON) file reading and CSV writing helpers."""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

from pf_sync_pkg.utils import clean


def read_tabular_rows(path: str) -> List[Dict[str, Any]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".json":
        with open(path, "r", encoding="utf-8") as handle:
            raw = json.load(handle)
        if isinstance(raw, dict):
            raw = raw.get("rows", raw.get("data", []))
        if not isinstance(raw, list):
            raise ValueError(f"JSON tabular input must contain a list of rows: {path}")
        return [dict(item) for item in raw if isinstance(item, dict)]

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read XLSX appointment reports") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean(value) or f"column_{index + 1}" for index, value in enumerate(rows[0])]
        return [
            {headers[index]: clean(value) for index, value in enumerate(row)}
            for row in rows[1:]
            if any(clean(value) for value in row)
        ]

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"Tabular file has no header row: {path}")
        return [dict(row) for row in reader]


def write_csv(path: str, rows: Sequence[Dict[str, Any]]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            key = clean(key)
            if key and key not in seen:
                headers.append(key)
                seen.add(key)
    if not headers:
        headers = ["message"]
        rows = [{"message": "No results"}]
    with target.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: clean(row.get(key)) for key in headers})
