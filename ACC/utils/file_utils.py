# ==========================================================
# utils/file_utils.py  (Merged & Refactored)
# ==========================================================
"""
Generic File Utilities for ACC RPA
----------------------------------
Purpose:
    Provide carrier-agnostic helpers for reading/writing files and extracting text.

Responsibilities:
    • Read .txt, .csv, .xlsx, .pdf files
    • Append rows to templates (Excel/CSV)
    • Extract dates/text from PDFs (E&O, certificates)
    • Cleanup temporary or downloaded files
    • Generate Success/Error logs
    • Upload to Azure Blob Storage

Excludes:
    • Carrier-specific parsing or matrix logic
"""

import os
import re
import csv
import json
import pdfplumber
import pandas as pd
import tempfile
from pathlib import Path
from datetime import datetime
from typing import List, Optional
from openpyxl import load_workbook
from pdf2image import convert_from_path
import pytesseract
import shutil
from utils.azure_blob_utils import authenticate_blob_storage, upload_file_to_blob
from utils.logger_utils import safe_log
from utils import drive_utils
from utils import config

SUCCESS_BLOB_PATH = "raw/agent_contract_request_carrier/success/"
ERROR_BLOB_PATH = "raw/agent_contract_request_carrier/error/"


# ==========================================================
# 1️⃣ PDF DATE EXTRACTION
# ==========================================================
def extract_dates_from_text(text: str) -> List[datetime]:
    """Find all date strings and return as datetime objects."""
    patterns = [
        #r"\b(0?[1-9]|1 *[0-2])[\/\-\.](0?[1-9]|[12][0-9]|3[01])[\/\-\.](\d{2,4})\b",
        #r"\b(\d{4})[\/\-\.](0?[1-9]|1[0-2])[\/\-\.](0?[1-9]|[12][0-9]|3[01])\b",
        # Input error margin of 2 spaces between characters:
        r"\b(0? {0,2}[1-9]|1 {0,2}[0-2]) {0,2}[\/\-\.] {0,2}(0? {0,2}[1-9] {0,2}|[12] {0,2}[0-9]|3 {0,2}[01]) {0,2}[\/\-\.] {0,2}(\d{2,4})\b",
        r"\b(\d{4}) {0,2}[\/\-\.] {0,2}(0? {0,2}[1-9]|1 {0,2}[0-2]) {0,2}[\/\-\.] {0,2}(0? {0,2}[1-9]|[12] {0,2}[0-9]|3 {0,2}[01])\b",
    ]
    dates = []
    for pat in patterns:
        for m in re.findall(pat, text):
            try:
                y, mth, d = _normalize_date_parts(m)
                dt = datetime(int(y), int(mth), int(d))
                if 2000 <= dt.year <= 2100:
                    dates.append(dt)
            except Exception:
                continue
    return dates


def _normalize_date_parts(parts: tuple) -> tuple:
    parts = (str(parts[0]).replace(' ',''),str(parts[1]).replace(' ',''),str(parts[2]).replace(' ',''))
    if len(parts[0]) == 4:
        y, mth, d = parts
    elif len(parts[2]) == 4:
        mth, d, y = parts
    else:
        mth, d, y = parts
        y = f"20{y}" if len(y) == 2 else y
    return y, mth, d


def extract_latest_date_from_pdf(pdf_path: str) -> Optional[datetime]:
    """
    Smart E&O expiration extractor.

    Strategy:
      1. Extract all PDF text line-by-line
      2. Find lines containing "policy", "expiration", "policy period", etc.
      3. Extract dates from those lines + the next line
      4. If none found → fallback to OCR first 2 pages
      5. If still none → fallback to entire-PDF date scan
      6. Return the max (true expiration date)
    """
    if not pdf_path or not Path(pdf_path).exists():
        return None

    # --------------------------------------------------
    # Step 1: Load and split text by lines
    # --------------------------------------------------
    try:
        with pdfplumber.open(pdf_path) as pdf:
            lines = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                lines.extend(t.splitlines())
    except Exception:
        lines = []

    # OCR fallback if PDF has no text
    if not lines:
        print("==OCR Fallback")
        ocr_text = extract_text_via_ocr(pdf_path)
        lines = ocr_text.splitlines()

    if not lines:
        return None

    # --------------------------------------------------
    # Step 2: Identify "policy"-related lines
    # --------------------------------------------------
    KEYWORDS = [
        "policy",
        "policy period",
        "policy term",
        "policy expiration",
        "policy expires",
        "expiration",
        "expiry",
        "effective",
        "coverage period",
    ]

    candidate_text_blocks = []

    for i, line in enumerate(lines):
        L = line.lower()
        if any(k in L for k in KEYWORDS):
            block = line

            # Also include next line if present (common in PDFs)
            if i + 1 < len(lines):
                block += " " + lines[i + 1]

            candidate_text_blocks.append(block)

    # --------------------------------------------------
    # Step 3: Extract dates from policy blocks
    # --------------------------------------------------
    extracted_dates = []
    for block in candidate_text_blocks:
        extracted_dates.extend(extract_dates_from_text(block))

    # --------------------------------------------------
    # Step 4: Fallback → extract from full PDF text
    # --------------------------------------------------
    if not extracted_dates:
        full_text = "\n".join(lines)
        extracted_dates = extract_dates_from_text(full_text)

    if not extracted_dates:
        return None

    # --------------------------------------------------
    # Step 5: Return latest date
    # --------------------------------------------------
    return max(extracted_dates).date()



# ==========================================================
# 2️⃣ OCR FALLBACK
# ==========================================================
def extract_text_via_ocr(pdf_path: str) -> str:
    """Extract text from a scanned PDF using pytesseract."""
    try:
        text = ""
        with tempfile.TemporaryDirectory() as tmp:
            pages = convert_from_path(pdf_path, dpi=800, first_page=1, last_page=2, output_folder=tmp)
            for img in pages:
                text += pytesseract.image_to_string(img)
        return text
    except Exception as e:
        print(f'[OCR] Error: {e}')
        return ""

# ==========================================================
# 3️⃣ APPEND ROWS (Preserve Excel Formatting)
# ==========================================================
def append_rows_preserve_format(file_path: str, df_rows: pd.DataFrame, skip_rows: int = 1):
    """
    Append rows to Excel template while preserving logos and formatting.
    Starts inserting data *after* header rows defined by skip_rows.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active

    next_row = ws.max_row + 1 if ws.max_row > skip_rows else skip_rows + 1
    for _, rec in df_rows.iterrows():
        for col_idx, val in enumerate(rec.values, start=1):
            v = None if pd.isna(val) or val == "" else val
            ws.cell(row=next_row, column=col_idx, value=v)
        next_row += 1

    wb.save(file_path)
    print(f"✅ Appended {len(df_rows)} record(s) to {os.path.basename(file_path)} (format preserved)")

# ==========================================================
# 3️⃣ APPEND TO TEMPLATE (Mapped Columns)
# ==========================================================
def append_to_template(
    df: pd.DataFrame,
    carrier_row: dict,
    base_template_path: str = None,
    mapping: dict = None,
    header_row: int = 1,
    header_column: int = 0,
    summary: dict = None
) -> Optional[str]:
    try:
        import json
        import pandas as pd
        import os
        from openpyxl import load_workbook

        carrier_name = carrier_row.get("carrier_name", "UnknownCarrier")
        if df is None or df.empty:
            print(f"ℹ️ No rows to append for {carrier_name}.")
            return None
        print("==Received dataframe:")
        print(df.to_string())
        # ----------------------------------------------------------
        # 1️⃣ Resolve mapping (priority: arg → summary → carrier_row)
        # ----------------------------------------------------------
        if not mapping:
            if summary and summary.get("template_mapping"):
                mapping = summary["template_mapping"]
            else:
                mapping_raw = carrier_row.get("template_field_map")
                if mapping_raw:
                    mapping = json.loads(mapping_raw) if isinstance(mapping_raw, str) else mapping_raw

        if not mapping:
            print(f"⚠️ No template mapping found for {carrier_name}.")
            return None

        # ----------------------------------------------------------
        # 2️⃣ Build export DataFrame with correct columns
        # ----------------------------------------------------------
        export_cols = list(mapping.keys())
        mapped_df = pd.DataFrame(columns=export_cols)

        for tpl_col, df_col in mapping.items():
            if df_col and isinstance(df_col, str) and df_col.strip():
                if df_col in df.columns:
                    mapped_df[tpl_col] = df[df_col]
                else:
                    # treat as literal value
                    mapped_df[tpl_col] = [df_col] * len(df)
            else:
                mapped_df[tpl_col] = ""

        mapped_df.drop_duplicates(keep='first',inplace=True)
        # ----------------------------------------------------------
        # 3️⃣ Locate and open the template
        # ----------------------------------------------------------
        template_path = base_template_path
        if not template_path:
            template_name = carrier_row.get("carrier_template")
            download_dir = carrier_row.get("download_path") or os.getenv("TEMP", "/tmp")
            template_path = os.path.join(download_dir, template_name)

        if not template_path or not os.path.exists(template_path):
            print(f"⚠️ Template not found: {template_path}")
            return None

        wb = load_workbook(template_path)
        ws = wb.active

        # ----------------------------------------------------------
        # 4️⃣ Determine where to append (after header_row)
        # ----------------------------------------------------------
        max_row = ws.max_row
        if max_row > header_row:
            ws.delete_rows(header_row + 1, max_row - header_row)

        # 4B — now the next writable row is always header_row + 1
        next_row = header_row + 1

        # 4C — append the mapped rows
        for _, record in mapped_df.iterrows():
            for col_idx, val in enumerate(record.values, start=1 + header_column):
                v = None if pd.isna(val) or val == "" else val
                ws.cell(row=next_row, column=col_idx, value=v)
            next_row += 1

        wb.save(template_path)
        print(f"✅ Appended {len(mapped_df)} row(s) to template → {os.path.basename(template_path)}")
        return template_path

    except Exception as e:
        from utils.logger_utils import safe_log
        safe_log(
            "ACC_RPA_FILEUTILS",
            f"Template append failed for {carrier_row.get('carrier_name')}: {e}",
            code="TEMPLATE_APPEND_ERROR",
        )
        print(f"❌ Template append failed for {carrier_row.get('carrier_name')}: {e}")
        return None



# ==========================================================
# 4️⃣ UPLOAD FINAL TEMPLATE (EOD)
# ==========================================================
def upload_template_to_success(template_path: str, carrier_row: dict):
    """
    Timestamp and upload filled template to success blob after EOD email send.
    """
    try:
        carrier_name = carrier_row.get("carrier_name", "UnknownCarrier")
        if not template_path or not os.path.exists(template_path):
            print(f"⚠️ No valid template to upload for {carrier_name}")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{Path(template_path).stem}_{timestamp}{Path(template_path).suffix}"
        renamed_path = os.path.join(os.path.dirname(template_path), new_name)
        os.rename(template_path, renamed_path)

        blob_service = authenticate_blob_storage()
        blob_name = f"{SUCCESS_BLOB_PATH}{os.path.basename(renamed_path)}"
        url = upload_file_to_blob(blob_service, renamed_path, blob_name, container_name="834analytics-dev")
        print(f"☁️ Uploaded final template for {carrier_name} → {url}")
        return url

    except Exception as e:
        safe_log("ACC_RPA_FILEUTILS", f"Template upload failed for {carrier_row.get('carrier_name')}: {e}",
                 code="TEMPLATE_UPLOAD_FAIL")
        return None


# ==========================================================
# 5️⃣ OUTPUT FILE GENERATION (Success / Error .txt)
# ==========================================================
def generate_output_files(df_success: pd.DataFrame,
                          df_error: pd.DataFrame,
                          carrier_row: dict,
                          run_id: str) -> dict:
    """Generate Success + Error TXT logs."""
    carrier_name = carrier_row.get("carrier_name", "UnknownCarrier")
    download_dir = carrier_row.get("download_path") or os.getenv("TEMP", "/tmp")
    os.makedirs(download_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    success_file, error_file = None, None

    # ✅ Explicit empty checks to avoid ambiguous DataFrame truth value
    if isinstance(df_success, pd.DataFrame) and not df_success.empty:
        success_file = os.path.join(download_dir, f"{carrier_name}_Success_{timestamp}.txt")
        df_success.to_csv(success_file, sep="\t", index=False)
        print(f"✅ Success TXT generated: {success_file}")
    else:
        print("ℹ️ No success records — skipping success file creation.")

    if isinstance(df_error, pd.DataFrame) and not df_error.empty:
        error_file = os.path.join(download_dir, f"{carrier_name}_Error_{timestamp}.txt")
        df_error.to_csv(error_file, sep="\t", index=False)
        print(f"⚠️ Error TXT generated: {error_file}")
    else:
        print("ℹ️ No errors — skipping error file creation.")

    return {"success_file": success_file, "error_file": error_file}



# ==========================================================
# 6️⃣ UPLOAD OUTPUT FILES
# ==========================================================
def upload_output_files_to_blob(file_dict: dict) -> dict:
    """Upload generated .txt success/error files."""
    uploaded = {}
    blob_service = authenticate_blob_storage()

    for key, local_path in file_dict.items():
        if not local_path or not os.path.exists(local_path):
            print(f"[DEBUG] Checking upload path → key={key}, local_path={local_path}")
            if not local_path or not os.path.exists(local_path):
                print(f"[DEBUG] SKIP → Path does not exist: {local_path}")
                continue
        try:
            folder = SUCCESS_BLOB_PATH if "success" in key.lower() else ERROR_BLOB_PATH
            blob_name = f"{folder}{os.path.basename(local_path)}"
            url = upload_file_to_blob(blob_service, local_path, blob_name, container_name="834analytics-dev")
            uploaded[key] = url
            print(f"☁️ Uploaded {os.path.basename(local_path)} → {url}")
        except Exception as e:
            safe_log("ACC_RPA_FILEUTILS", f"Blob upload failed for {local_path}: {e}", code="BLOB_UPLOAD_ERROR")

    return uploaded

def _norm(s: str):
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())

def _like(name: str, keys: list):
    return any(k in _norm(name) for k in [_norm(k) for k in keys])

def pick_best_file(files, keys):
    """
    Generic picker:
      • Filters by name keywords
      • Returns the file with latest modifiedTime
    (Used for W9 and Contract only – no E&O logic here.)
    """
    if not files:
        return None

    # Filter by keyword match
    matches = [
        f for f in files
        if _like(f.get("name", ""), keys) and not f.get("mimeType", "").endswith("folder")
    ]
    if not matches:
        return None

    # Convert modifiedTime and pick latest
    for f in matches:
        f["_mtime"] = pd.to_datetime(
            f.get("modifiedTime", None), utc=True, errors="coerce"
        )

    latest = max(f["_mtime"] for f in matches)
    same_ts = [f for f in matches if f["_mtime"] == latest]

    # Return one file (we’ll download it later into agent folder)
    return same_ts[0]

def pick_best_eo_file(files, agent_dir: str, npn: str = ""):
    """
    E&O-specific picker:
      • Only considers PDFs
      • Uses latest *date* (YYYY-MM-DD) of modifiedTime
      • For that latest date, downloads *all* candidates into agent_dir
      • Extracts expiry from each and picks the best:
            - Prefer non-expired, with the furthest expiry
            - If all expired, pick the latest expiry
            - If no expiry parsed, fall back to first successfully downloaded file
    Returns:
      (best_eo_path: str or None, best_expiry: date or None)
    """
    if not files:
        return None, None

    pdfs = [f for f in files if f.get("mimeType") == "application/pdf"]
    if not pdfs:
        return None, None

    # Compute modified *date* for same-day logic
    for f in pdfs:
        dt = pd.to_datetime(f.get("modifiedTime", None), utc=True, errors="coerce")
        f["_mdate"] = dt.date() if pd.notnull(dt) else None

    valid_pdfs = [f for f in pdfs if f.get("_mdate") is not None]
    if not valid_pdfs:
        return None, None

    latest_date = max(f["_mdate"] for f in valid_pdfs)
    candidates = [f for f in valid_pdfs if f["_mdate"] == latest_date]

    best_path = None
    best_expiry = None
    today = datetime.utcnow().date()

    for f in candidates:
        normalized_filename = str(f["name"]).replace('/','_')
        normalized_filename = normalized_filename.replace('\\','_')
        local_path = os.path.join(agent_dir, normalized_filename)
        try:
            drive_utils.download_file(f["id"], local_path)
        except Exception as e:
            print(f"⚠️ Download failed for E&O {f.get('name')}: {e}")
            continue

        expiry = extract_latest_date_from_pdf(local_path)

        # First one with any expiry
        if best_expiry is None and expiry:
            best_expiry = expiry
            best_path = local_path
            continue

        if expiry:
            # Prefer non-expired over expired; within each group, pick furthest expiry
            if expiry >= today:
                if best_expiry is None or best_expiry < today or expiry > best_expiry:
                    best_expiry = expiry
                    best_path = local_path
            else:
                if best_expiry is not None and best_expiry < today and expiry > best_expiry:
                    best_expiry = expiry
                    best_path = local_path

        # Still nothing with a parsed expiry → allow a no-expiry file as fallback
        if not expiry and best_path is None and best_expiry is None:
            best_path = local_path

    if best_path is None:
        # All downloads failed
        return None, None

    return best_path, best_expiry

# ==========================================================
# 7️⃣ VALIDATION DOCS
# ==========================================================
def validate_priority_docs(
    gdrive_url: str,
    npn: str,
    base_download_path: str,
    carrier_name: str = "",
    skip_non_contract: bool = False
) -> dict:

    """
    Crawl the agent's Drive folder, find latest E&O, W9, and Contract files.
    Uses normalized carrier variants (from config + auto-generated) to detect contracts.
    """
    import pandas as pd
    from datetime import datetime
    import re, shutil, os
    from utils import config  # ✅ access CARRIER_VARIANT_MAP

    # ==========================================================
    # 1️⃣ Define folder keywords
    # ==========================================================
    EO_PARENT_FOLDERS = ["eo", "e&o", "e-o", "errors", "omission", "omissions"]
    EOKW = ["eo", "e&o", "e-o", "errors", "error", "ommission", "ommissions", "certificate", "policy"]
    W9KW = ["w9", "w-9"]
    W9FOLD = ["financial", "financials", "finance"]
    CONTRACT_PARENT_FOLDERS = ["contract", "contracts", "contracting"]

    # ==========================================================
    # 2️⃣ Build carrier variants (core + config-based)
    # ==========================================================
    carrier_variants = []
    if carrier_name:
        cname = carrier_name.strip()
        c_lower = cname.lower()
        base_norm = _norm(cname)

        # Core normalized variants
        carrier_variants = [
            cname,
            cname.replace(" ", ""),       # remove spaces
            cname.split()[0],             # first token (e.g. "Priority" from "Priority Health")
            base_norm,                    # normalized full name
        ]

        # Add shorter variants if compound names
        if "health" in base_norm:
            carrier_variants.append(base_norm.replace("health", ""))
        if "insurance" in base_norm:
            carrier_variants.append(base_norm.replace("insurance", ""))

        # Add mapped variants dynamically from config
        for key, variants in getattr(config, "CARRIER_VARIANT_MAP", {}).items():
            if key in c_lower:
                carrier_variants.extend(variants)

    print(f"[DEBUG] Contract variants for {carrier_name}: {carrier_variants}")

    # ==========================================================
    # 3️⃣ Drive traversal logic
    # ==========================================================
    folder_id = None
    if gdrive_url and "/folders/" in gdrive_url:
        folder_id = gdrive_url.split("/folders/")[-1].split("?")[0]

    try:
        files = drive_utils.list_all_files_recursive(folder_id)
        print(f"📂 Recursively found {len(files)} file(s)/folder(s) under {folder_id}")
    except Exception as e:
        print(f"⚠️ Drive listing failed for {npn}: {e}")
        files = []

    # Local folder setup
    agent_dir = os.path.join(base_download_path, "single", npn)
    if os.path.isdir(agent_dir):
        shutil.rmtree(agent_dir, ignore_errors=True)
    os.makedirs(agent_dir, exist_ok=True)

    # ==========================================================
    # 4️⃣ Discover E&O + W9 - Multiple Folders
    # ==========================================================
    eo_path = None
    eo_valid_until = None
    f_w9 = None

    if not skip_non_contract:
        # 🔍 E&O candidates
        eo_parents = (f for f in files if _like(f.get("name", ""), EO_PARENT_FOLDERS)
             and f.get("mimeType", "").endswith("folder"))

        for folder in eo_parents:
            eo_candidates = []
            if folder:
                sub_files = drive_utils.list_all_files_recursive(folder["id"])
                eo_candidates = [f for f in sub_files if f.get("mimeType") == "application/pdf"]
                if eo_candidates:
                    print(f"📄 Found {len(eo_candidates)} E&O candidate(s) in E&O folder")
            else:
                eo_candidates = [
                    f for f in files
                    if f.get("mimeType") == "application/pdf" and _like(f.get("name", ""), EOKW)
                ]

            if eo_candidates:
                eo_path, eo_valid_until = pick_best_eo_file(eo_candidates, agent_dir, npn)
                if eo_path:
                    break
            else:
                print(f"ℹ️ No E&O candidates found for {npn}")

        # 🔍 W-9 (still single best file by modifiedTime)
        f_w9 = pick_best_file(files, W9KW + W9FOLD)


    # ==========================================================
    # 6️⃣ Discover Contracts - Multiple folders
    # ==========================================================
    f_ctr = None
    contract_parents = (f for f in files if _like(f.get("name", ""), CONTRACT_PARENT_FOLDERS)
         and f.get("mimeType", "").endswith("folder"))

    for folder in contract_parents:
        sub_files = drive_utils.list_all_files_recursive(folder["id"])

        carrier_folder = next(
            (sf for sf in sub_files
             if sf.get("mimeType", "").endswith("folder") and _like(sf.get("name", ""), carrier_variants)),
            None,
        )

        pdf_candidates = []
        if carrier_folder:
            carrier_files = drive_utils.list_all_files_recursive(carrier_folder["id"])
            pdf_candidates = [f for f in carrier_files if f.get("mimeType") == "application/pdf"]
        else:
            pdf_candidates = [
                f for f in sub_files
                if f.get("mimeType") == "application/pdf" and _like(f.get("name", ""), carrier_variants)
            ]

        carrier_variants.append(str(npn))
        if pdf_candidates:
            f_ctr = pick_best_file(pdf_candidates, carrier_variants)
            if not f_ctr:
                print(f"ℹ️ No Contract PDF found for {npn}")
            else:
                print(f"📄 Found contract PDF for {carrier_name}: {f_ctr.get('name')}")
                break
        pass


    # ==========================================================
    # 7️⃣ Download selected files (W9 + Contract ONLY)
    # ==========================================================
    def _dl(f):
        if not f:
            return None
        try:
            out_path = os.path.join(agent_dir, f["name"])
            return drive_utils.download_file(f["id"], out_path)
        except Exception as e:
            print(f"⚠️ Download failed for {f.get('name')}: {e}")
            return None

    # eo_path already downloaded in pick_best_eo_file (if any)
    w9_path = _dl(f_w9)
    ctr_path = _dl(f_ctr)

    # ==========================================================
    # 8️⃣ Validation summary
    # ==========================================================
    print(f"[E&O] Extracted E&O expiry for {npn}: {eo_valid_until}")
    today = datetime.utcnow().date()
    eo_valid = bool(eo_valid_until and eo_valid_until >= today)
    attachments = [p for p in [ctr_path, w9_path, eo_path] if p and os.path.isfile(p)]

    return {
        "eo_path": eo_path,
        "w9_path": w9_path,
        "contract_path": ctr_path,
        "eo_valid_until": eo_valid_until,
        "eo_valid": eo_valid,
        "attachments": attachments,
    }



# ==========================================================
# 8️⃣ CLEANUP DOWNLOAD PATH
# ==========================================================
def cleanup_download_path(download_path: str, active_templates: list = None):
    """
    Clean up per-carrier download directory safely.

    Behavior:
      • Keeps base folder (e.g., acc/)
      • Removes all contents of 'single/' including subfolders
      • Cleans files in all other subfolders but keeps their structure
      • Skips active template files if provided
    """
    import os, shutil

    if not download_path or not os.path.exists(download_path):
        print(f"⚠️ Download path not found: {download_path}")
        return

    active_templates = [t.lower() for t in (active_templates or [])]

    for root, dirs, files in os.walk(download_path):
        for file in files:
            file_path = os.path.join(root, file)
            # Skip template files
            if any(t in file.lower() for t in active_templates):
                continue
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"⚠️ Failed to delete file {file_path}: {e}")

        # Clean each subfolder
        for subdir in dirs:
            subdir_path = os.path.join(root, subdir)

            # 🚮 Special handling for 'single/' folder — delete everything inside it
            if os.path.basename(subdir_path).lower() == "single":
                try:
                    shutil.rmtree(subdir_path, ignore_errors=True)
                    os.makedirs(subdir_path, exist_ok=True)
                    print(f"🧹 Cleared entire 'single/' folder: {subdir_path}")
                except Exception as e:
                    print(f"⚠️ Failed to clean 'single/' folder {subdir_path}: {e}")
                continue

            # Default cleanup for all other folders (preserve structure)
            try:
                for sub_file in os.listdir(subdir_path):
                    sub_file_path = os.path.join(subdir_path, sub_file)
                    if os.path.isfile(sub_file_path):
                        if any(t in sub_file.lower() for t in active_templates):
                            continue
                        os.remove(sub_file_path)
            except Exception as e:
                print(f"⚠️ Failed to clean subfolder {subdir_path}: {e}")

    print(f"🧹 Cleaned all files inside {download_path} (preserved folders + templates).")



from utils.azure_blob_utils import authenticate_blob_storage, download_file_from_blob

from typing import Optional

def download_base_template_from_blob(carrier_row: dict) -> Optional[str]:
    """
    Downloads the carrier's base template (.xlsx or .csv) from Azure Blob Storage
    into its local download folder.
    Uses [base_blob_url] + [carrier_template] from process matrix.
    """
    from utils.azure_blob_utils import authenticate_blob_storage, download_file_from_blob

    base_blob = carrier_row.get("base_blob_url")
    template_name = carrier_row.get("carrier_template")
    download_dir = carrier_row.get("download_path") or os.getenv("TEMP", "/tmp")

    if not base_blob or not template_name:
        print(f"⚠️ Missing base_blob_url or carrier_template for {carrier_row.get('carrier_name')}.")
        return None

    # Normalize extension
    if not (template_name.endswith(".xlsx") or template_name.endswith(".csv")):
        template_name = f"{template_name}.xlsx"

    blob_path = f"{base_blob.rstrip('/')}/templates/{template_name}"
    local_path = os.path.join(download_dir, template_name)

    try:
        blob_service = authenticate_blob_storage()
        print(f"☁️ Fetching base template from blob: {blob_path}")
        return download_file_from_blob(blob_service, blob_path, local_path)
    except Exception as e:
        from utils.logger_utils import safe_log
        safe_log(
            "ACC_RPA_FILEUTILS",
            f"Template download failed for {carrier_row.get('carrier_name')}: {e}",
            code="AZURE_DOWNLOAD_ERROR"
        )
        return None


